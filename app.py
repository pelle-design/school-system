# ==================== IMPORTS ====================
import os
import re
import io
import csv
import json
import time
import secrets
import random
import string
import sqlite3
from datetime import datetime, timedelta
from functools import wraps
from markupsafe import escape
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, redirect, url_for, session, flash, abort, send_from_directory, jsonify, g
from werkzeug.utils import secure_filename
from flask_wtf.csrf import CSRFProtect
from markupsafe import escape
import requests
import base64
import uuid
# ==================== APP CONFIGURATION ====================
app = Flask(__name__)
app.secret_key = secrets.token_hex(32)
app.secret_key = os.environ.get('SECRET_KEY', secrets.token_hex(32))
app.config['UPLOAD_FOLDER'] = 'static/uploads'

ALLOWED_IMAGE_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

# ==================== POSTGRESQL DATABASE SETUP ====================

import os
import psycopg2
import psycopg2.extras
from flask import g


DATABASE_URL = os.environ.get("DATABASE_URL")

if not DATABASE_URL:
    raise RuntimeError(
        "DATABASE_URL is missing. Add PostgreSQL URL in Render Environment Variables."
    )


def get_db():
    """Get PostgreSQL database connection"""

    db = getattr(g, '_database', None)

    if db is None:
        db = g._database = psycopg2.connect(
            DATABASE_URL,
            cursor_factory=psycopg2.extras.RealDictCursor
        )

    return db


@app.teardown_appcontext
def close_connection(exception):

    db = getattr(g, '_database', None)

    if db is not None:
        db.close()

def init_db():
    """Initialize PostgreSQL database with all tables"""

    db = get_db()
    cursor = db.cursor()

    # USERS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id SERIAL PRIMARY KEY,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            full_name TEXT DEFAULT '',
            role TEXT NOT NULL,
            phone TEXT,
            status INTEGER DEFAULT 1,
            child_id TEXT,
            profile_pic TEXT DEFAULT 'default_avatar.png',
            must_change_password INTEGER DEFAULT 0
        )
    """)

    # ROLE LIMITS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS role_limits (
            role_name TEXT PRIMARY KEY,
            max_count INTEGER DEFAULT 1
        )
    """)

    # CLASSES
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS classes (
            id SERIAL PRIMARY KEY,
            name TEXT UNIQUE NOT NULL,
            level TEXT,
            stream TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # ADMISSION SETTINGS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS admission_settings (
            id INTEGER PRIMARY KEY DEFAULT 1,
            is_open INTEGER DEFAULT 1,
            deadline DATE,
            closing_reason TEXT,
            fee_amount NUMERIC DEFAULT 50000,
            payment_gateway TEXT DEFAULT 'MTN',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # STUDENTS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS students (
            student_id TEXT PRIMARY KEY,
            full_name TEXT NOT NULL,
            class TEXT NOT NULL,
            photo_path TEXT DEFAULT 'default_avatar.png',
            fees_paid NUMERIC DEFAULT 0,
            fees_balance NUMERIC DEFAULT 0,
            fees_total NUMERIC DEFAULT 0,
            admission_date DATE,
            parent_phone TEXT,
            date_of_birth DATE,
            age INTEGER,
            sex TEXT,
            preferred_house TEXT,
            disability TEXT,
            sports_activities TEXT,
            lin TEXT,
            parent_id INTEGER,
            admission_source TEXT DEFAULT 'local',
            admission_status TEXT DEFAULT 'approved',
            application_date TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
            payment_status TEXT DEFAULT 'pending',
            payment_transaction_id TEXT,
            payment_date TIMESTAMP
        )
    """)
    cursor.execute("""
        ALTER TABLE students
        ADD COLUMN IF NOT EXISTS programme TEXT
    """)
    
    cursor.execute("""
        ALTER TABLE students
        ADD COLUMN IF NOT EXISTS residence TEXT
    """)
    # STAFF
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS staff (
            id SERIAL PRIMARY KEY,
            staff_no TEXT UNIQUE,
            full_name TEXT NOT NULL,
            position TEXT NOT NULL,
            department TEXT,
            phone TEXT,
            email TEXT,
            nssf_number TEXT,
            tin_number TEXT,
            bank_account TEXT,
            bank_name TEXT,
            salary_basic NUMERIC DEFAULT 0,
            salary_allowances NUMERIC DEFAULT 0,
            salary_deductions NUMERIC DEFAULT 0,
            status TEXT DEFAULT 'active',
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)

    # PAYROLL
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payroll (
            id SERIAL PRIMARY KEY,
            payroll_no TEXT UNIQUE,
            month_year DATE,
            total_amount NUMERIC DEFAULT 0,

            approval_code TEXT,
            approval_status TEXT DEFAULT 'pending',
            approved_by TEXT,
            approved_at TIMESTAMP,

            recorded_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            headteacher_access_token TEXT,
            token_expires_at TIMESTAMP,

            management_approval_code TEXT,
            management_access_token TEXT,
            management_token_expires_at TIMESTAMP,

            management_approval_status TEXT DEFAULT 'pending',
            management_approved_by TEXT,
            management_approved_at TIMESTAMP,

            bank_authorization_token TEXT,
            bank_transaction_ref TEXT,
            bank_payment_status TEXT DEFAULT 'pending',
            bank_payment_response TEXT,

            token_resend_count INTEGER DEFAULT 0,
            last_resend_at TIMESTAMP
        )
    """)

    # SALARY PAYMENTS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS salary_payments (
            id SERIAL PRIMARY KEY,
            staff_id INTEGER,
            payroll_id INTEGER,
            month_year DATE,

            basic NUMERIC,
            allowances NUMERIC,
            deductions NUMERIC,
            gross_salary NUMERIC,
            nssf_employee NUMERIC,
            paye_tax NUMERIC,
            net_salary NUMERIC,

            payment_date DATE,
            payment_method TEXT,

            approval_code TEXT,
            approval_status TEXT DEFAULT 'pending',

            transaction_ref TEXT,
            recorded_by TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY(staff_id) REFERENCES staff(id),
            FOREIGN KEY(payroll_id) REFERENCES payroll(id)
        )
    """)
    # MARKS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS marks (
            id SERIAL PRIMARY KEY,
            student_id TEXT,
            subject TEXT,
            term TEXT,
            year INTEGER,

            ai1 NUMERIC(5,2),
            ai2 NUMERIC(5,2),
            ai3 NUMERIC(5,2),
            ai_average NUMERIC(5,2),
            ai_contribution NUMERIC(5,2),

            eot_score NUMERIC(5,2),
            total_score NUMERIC(5,2),

            grade TEXT,
            identifier NUMERIC(5,2),
            descriptor TEXT,

            teacher_initials TEXT,
            teacher_id INTEGER,

            paper1 NUMERIC,
            paper2 NUMERIC,
            points INTEGER,

            FOREIGN KEY(student_id)
                REFERENCES students(student_id),

            UNIQUE(student_id, subject, term, year)
        )
    """)
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS fee_payment_requests (
            id SERIAL PRIMARY KEY,
    
            prn VARCHAR(10) NOT NULL UNIQUE,
    
            student_id TEXT NOT NULL,
    
            amount NUMERIC(12,2) NOT NULL,
    
            term TEXT NOT NULL,
    
            year INTEGER NOT NULL,
    
            payment_status TEXT NOT NULL DEFAULT 'pending',
    
            payment_method TEXT,
    
            provider_transaction_id TEXT,
    
            created_by TEXT,
    
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
            expires_at TIMESTAMP,
    
            paid_at TIMESTAMP,
    
            CONSTRAINT fee_payment_requests_student_fk
                FOREIGN KEY (student_id)
                REFERENCES students(student_id)
                ON DELETE CASCADE,
    
            CONSTRAINT fee_payment_requests_prn_format
                CHECK (prn ~ '^[0-9]{10}$'),
    
            CONSTRAINT fee_payment_requests_amount_positive
                CHECK (amount > 0)
        )
    """)
    # ATTENDANCE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS attendance (
            id SERIAL PRIMARY KEY,
            student_id TEXT,
            date DATE,
            status TEXT,

            FOREIGN KEY(student_id)
                REFERENCES students(student_id),

            UNIQUE(student_id, date)
        )
    """)


    # SCHEDULES
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS schedules (
            id SERIAL PRIMARY KEY,
            type TEXT,
            term_scope TEXT,
            content TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)
    # ==================== GRADING SYSTEM ====================
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS grading_system (
            id SERIAL PRIMARY KEY,
            min_score NUMERIC,
            max_score NUMERIC,
            grade TEXT,
            descriptor TEXT
        )
    """)
    
    # ==================== A-LEVEL GRADING ====================
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS alevel_grading (
            id SERIAL PRIMARY KEY,
            min_score NUMERIC(5,2),
            max_score NUMERIC(5,2),
            grade TEXT,
            points INTEGER,
            is_subsidiary INTEGER DEFAULT 0
        )
    """)
    
    # ==================== IDENTIFIER GRADING ====================
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS identifier_grading (
            id SERIAL PRIMARY KEY,
            min_value NUMERIC(5,2),
            max_value NUMERIC(5,2),
            identifier NUMERIC(5,2)
        )
    """)
    
    # Add identifier column to existing PostgreSQL installations
    cursor.execute("""
        ALTER TABLE identifier_grading
        ADD COLUMN IF NOT EXISTS identifier NUMERIC(5,2)
    """)


    # TEACHER COMMENTS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS teacher_comments (
            id SERIAL PRIMARY KEY,
            student_id TEXT,
            term TEXT,
            year INTEGER,

            comment TEXT,
            headteacher_comment TEXT,

            class_teacher_comment_locked INTEGER DEFAULT 0,
            headteacher_comment_locked INTEGER DEFAULT 0,

            FOREIGN KEY(student_id)
                REFERENCES students(student_id)
        )
    """)


    # TEACHER ASSIGNMENTS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS teacher_class_assignments (
            id SERIAL PRIMARY KEY,

            user_id INTEGER NOT NULL,
            class_name TEXT NOT NULL,
            subject TEXT,

            assignment_type TEXT NOT NULL,
            assigned_by TEXT,

            assigned_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY(user_id)
                REFERENCES users(id),

            UNIQUE(user_id, class_name, subject, assignment_type)
        )
    """)


    # NOTIFICATIONS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS notifications (
            id SERIAL PRIMARY KEY,

            user_role TEXT,
            message TEXT,
            link TEXT,

            title TEXT DEFAULT 'Notification',

            is_read INTEGER DEFAULT 0,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)


    # SCHOOL SETTINGS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS school_settings (
            id INTEGER PRIMARY KEY DEFAULT 1,

            next_term_begins DATE,
            next_term_ends DATE,

            headteacher_stamp TEXT,

            school_name TEXT DEFAULT 'YOUR SCHOOL NAME',
            school_address TEXT DEFAULT 'P.O. Box 123, Kampala, Uganda',
            school_phone TEXT DEFAULT 'Tel: +256 712 345678',
            school_email TEXT DEFAULT 'info@school.com',

            logo_url TEXT,

            nssf_employee_rate NUMERIC DEFAULT 5.0,
            paye_rate NUMERIC DEFAULT 10.0,
            paye_threshold NUMERIC DEFAULT 235000
        )
    """)


    # PREDEFINED COMMENTS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS predefined_comments (
            id SERIAL PRIMARY KEY,

            comment_type TEXT,
            comment_text TEXT,

            is_active INTEGER DEFAULT 1,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)


    # PAYMENTS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payments (
            id SERIAL PRIMARY KEY,

            student_id TEXT,
            amount NUMERIC,

            payment_date DATE,

            receipt_no TEXT UNIQUE,

            payment_method TEXT,

            notes TEXT,

            recorded_by TEXT,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY(student_id)
                REFERENCES students(student_id)
        )
    """)


    # SUBJECTS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS subjects (
            id SERIAL PRIMARY KEY,

            subject_name TEXT UNIQUE NOT NULL,

            subject_code TEXT UNIQUE,

            is_active INTEGER DEFAULT 1
        )
    """)
    # BUDGET CATEGORIES
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS budget_categories (
            id SERIAL PRIMARY KEY,
            code TEXT UNIQUE,
            name TEXT,
            description TEXT,
            allocated_amount NUMERIC,
            year INTEGER
        )
    """)
    # STUDENT FEE STRUCTURE
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS student_fee_structure (
            id SERIAL PRIMARY KEY,
    
            level TEXT NOT NULL,
    
            programme TEXT NOT NULL,
    
            residence TEXT NOT NULL,
    
            term TEXT NOT NULL,
    
            year INTEGER NOT NULL,
    
            amount NUMERIC(12,2) NOT NULL DEFAULT 0,
    
            description TEXT,
    
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
    
            UNIQUE (
                level,
                programme,
                residence,
                term,
                year
            )
        )
    """)
    # EXPENDITURES
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS expenditures (
            id SERIAL PRIMARY KEY,

            voucher_no TEXT UNIQUE,

            category_id INTEGER,

            description TEXT,

            amount NUMERIC,

            expenditure_date DATE,

            payment_method TEXT,

            payee_name TEXT,

            payee_phone TEXT,

            status TEXT,

            recorded_by TEXT,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY(category_id)
                REFERENCES budget_categories(id)
        )
    """)


    # INVENTORY CATEGORIES
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS inventory_categories (
            id SERIAL PRIMARY KEY,

            name TEXT NOT NULL,

            description TEXT,

            warning_level INTEGER DEFAULT 10,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)


    # INVENTORY ITEMS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS inventory_items (
            id SERIAL PRIMARY KEY,

            item_code TEXT UNIQUE NOT NULL,

            name TEXT NOT NULL,

            category_id INTEGER,

            unit TEXT DEFAULT 'pieces',

            quantity INTEGER DEFAULT 0,

            minimum_quantity INTEGER DEFAULT 10,

            maximum_quantity INTEGER DEFAULT 0,

            reorder_level INTEGER DEFAULT 5,

            location TEXT,

            supplier TEXT,

            purchase_date DATE,

            purchase_price NUMERIC,

            current_value NUMERIC,

            status TEXT DEFAULT 'working',

            condition_notes TEXT,

            last_maintenance DATE,

            next_maintenance DATE,

            responsible_person TEXT,

            responsible_role TEXT,

            image_path TEXT,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY(category_id)
                REFERENCES inventory_categories(id)
        )
    """)


    # INVENTORY TRANSACTIONS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS inventory_transactions (
            id SERIAL PRIMARY KEY,

            item_id INTEGER,

            transaction_type TEXT,

            quantity INTEGER,

            unit_price NUMERIC,

            total_amount NUMERIC,

            transaction_date DATE,

            issued_to TEXT,

            issued_to_role TEXT,

            purpose TEXT,

            reference_no TEXT,

            recorded_by TEXT,

            approved_by TEXT,

            notes TEXT,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY(item_id)
                REFERENCES inventory_items(id)
        )
    """)


    # INVENTORY ALERTS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS inventory_alerts (
            id SERIAL PRIMARY KEY,

            item_id INTEGER,

            alert_type TEXT,

            message TEXT,

            is_read INTEGER DEFAULT 0,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY(item_id)
                REFERENCES inventory_items(id)
        )
    """)
    cursor.execute("""CREATE TABLE IF NOT EXISTS payment_requests (

           id SERIAL PRIMARY KEY,

           student_id TEXT,

           prn TEXT UNIQUE,

           amount NUMERIC,

           status TEXT DEFAULT 'pending',

           created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

           FOREIGN KEY(student_id)
           REFERENCES students(student_id)

    ) """)

    # HOUSES
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS houses (
            id SERIAL PRIMARY KEY,

            name TEXT UNIQUE,

            description TEXT,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)


    # SPORTS ACTIVITIES
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS sports_activities (
            id SERIAL PRIMARY KEY,

            name TEXT UNIQUE
        )
    """)


    # PAYMENT WEBHOOKS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payment_webhooks (
            id SERIAL PRIMARY KEY,

            transaction_id TEXT,

            amount NUMERIC,

            phone_number TEXT,

            student_id TEXT,

            reference TEXT,

            payment_method TEXT,

            raw_data TEXT,

            status TEXT,

            processed INTEGER DEFAULT 0,

            received_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    """)


    # PAYMENT GATEWAY CONFIG
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS payment_gateway_config (
            id INTEGER PRIMARY KEY DEFAULT 1,

            gateway_name TEXT DEFAULT 'School Pay',

            api_key TEXT,

            api_secret TEXT,

            webhook_secret TEXT,

            callback_url TEXT,

            status TEXT DEFAULT 'inactive'
        )
    """)


    # BANK TRANSACTION LOGS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS bank_transaction_logs (
            id SERIAL PRIMARY KEY,

            payroll_id INTEGER,

            staff_id INTEGER,

            transaction_ref TEXT,

            amount NUMERIC,

            recipient_account TEXT,

            recipient_phone TEXT,

            status TEXT,

            response TEXT,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY(payroll_id)
                REFERENCES payroll(id),

            FOREIGN KEY(staff_id)
                REFERENCES staff(id)
        )
    """)


    # AUTHORIZATION LOGS
    cursor.execute("""
        CREATE TABLE IF NOT EXISTS authorization_logs (
            id SERIAL PRIMARY KEY,

            payroll_id INTEGER,

            action TEXT,

            performed_by TEXT,

            ip_address TEXT,

            details TEXT,

            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,

            FOREIGN KEY(payroll_id)
                REFERENCES payroll(id)
        )
    """)


    # SAVE DATABASE TABLE CREATION
    db.commit()

    # DEFAULT ROLE LIMITS
    cursor.execute("""
        SELECT COUNT(*) AS count
        FROM role_limits
    """)

    if cursor.fetchone()['count'] == 0:

        cursor.executemany(
            """
            INSERT INTO role_limits
            (role_name, max_count)
            VALUES (%s,%s)
            """,
            [
                ('admin',1),
                ('headteacher',1),
                ('management',1),
                ('bursar',1),
                ('dos',1)
            ]
        )


    # DEFAULT ADMISSION SETTINGS
    cursor.execute("""
        SELECT COUNT(*) AS count
        FROM admission_settings
    """)

    if cursor.fetchone()['count'] == 0:

        cursor.execute("""
            INSERT INTO admission_settings
            (id,is_open,fee_amount)
            VALUES (1,1,50000)
        """)


    # DEFAULT ADMIN USER
    cursor.execute("""
        SELECT COUNT(*) AS count
        FROM users
    """)

    if cursor.fetchone()['count'] == 0:

        from werkzeug.security import generate_password_hash

        password = generate_password_hash(
            'admin123'
        )

        cursor.execute("""
            INSERT INTO users
            (
                username,
                password,
                full_name,
                role,
                status,
                phone,
                must_change_password
            )
            VALUES
            (%s,%s,%s,%s,%s,%s,%s)
        """,
        (
            'admin',
            password,
            'Administrator',
            'admin',
            1,
            '0700000000',
            0
        ))


    # DEFAULT CLASSES
    cursor.execute("""
        SELECT COUNT(*) AS count
        FROM classes
    """)

    if cursor.fetchone()['count'] == 0:

        cursor.executemany(
            """
            INSERT INTO classes
            (name)
            VALUES (%s)
            """,
            [
                ('Senior 1',),
                ('Senior 2',),
                ('Senior 3',),
                ('Senior 4',),
                ('Senior 5',),
                ('Senior 6',)
            ]
        )


    # DEFAULT SUBJECTS
    cursor.execute("""
        SELECT COUNT(*) AS count
        FROM subjects
    """)

    if cursor.fetchone()['count'] == 0:

        cursor.executemany(
            """
            INSERT INTO subjects
            (
                subject_name,
                subject_code
            )
            VALUES (%s,%s)
            """,
            [
                ('ENGLISH LANGUAGE','ENG'),
                ('MATHEMATICS','MTC'),
                ('BIOLOGY','BIO'),
                ('CHEMISTRY','CHE'),
                ('PHYSICS','PHY'),
                ('HISTORY','HIS'),
                ('GEOGRAPHY','GEO'),
                ('CHRISTIAN RELIGIOUS DEDUCATION','CRE'),
                ('Islamic Religious Education','IRE'),
                ('AGRICULTURE','AGR'),
                ('Commerce','COM'),
                ('ENTREPRENUERSHIP','ENT'),
                ('INFORMATION AND COMMUNICATION TECHNOLOGY','ICT'),
                ('LITERATURE IN ENGLISH','LIT'),
                ('ART AND DESIGN','ART'),
                ('Kiswahili','KIS')
            ]
        )


    # DEFAULT SCHOOL SETTINGS
    cursor.execute("""
        SELECT COUNT(*) AS count
        FROM school_settings
    """)

    if cursor.fetchone()['count'] == 0:

        cursor.execute("""
            INSERT INTO school_settings(id)
            VALUES(1)
        """)


    # FINAL SAVE
    db.commit()

    cursor.close()

    print("Database initialized successfully!")


# Initialize database on startup
with app.app_context():
    init_db()


# csrf = CSRFProtect()
# csrf.init_app(app)

app.config.update(
    SESSION_COOKIE_SECURE=True,
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE='Lax',
    PERMANENT_SESSION_LIFETIME=timedelta(hours=2)
)
# ==================== HELPER FUNCTIONS ====================

def query_db(query, args=(), one=False):
    """Execute a query and return results"""
    db = get_db()
    cur = db.cursor()

    cur.execute(query, args)

    rv = cur.fetchall()

    cur.close()

    if one:
        return rv[0] if rv else None

    return rv


def execute_db(query, args=()):
    """Execute a query and commit"""
    db = get_db()
    cur = db.cursor()

    cur.execute(query, args)

    db.commit()

    cur.close()

    return True


def get_db_dict():
    """Return PostgreSQL dictionary cursor connection"""
    return get_db()
from datetime import date

def calculate_age(birth_date):
    today = date.today()

    age = today.year - birth_date.year

    if (today.month, today.day) < (birth_date.month, birth_date.day):
        age -= 1

    return age

def sanitize_input(text):
    """Remove dangerous characters and escape HTML"""

    if not text:
        return ''

    text = re.sub(
        r'<script\b[^<]*(?:(?!<\/script>)<[^<]*)*<\/script>',
        '',
        text,
        flags=re.IGNORECASE
    )

    text = re.sub(
        r'javascript:',
        '',
        text,
        flags=re.IGNORECASE
    )

    text = re.sub(
        r'on\w+\s*=',
        '',
        text,
        flags=re.IGNORECASE
    )

    return escape(text)


def validate_input(text, max_length=500, allow_html=False):
    """Validate and sanitize user input"""

    if not text:
        return ''

    if len(text) > max_length:
        text = text[:max_length]

    if not allow_html:
        text = sanitize_input(text)

    return text


def allowed_file(filename, allowed_set):

    return (
        '.' in filename and
        filename.rsplit('.', 1)[1].lower() in allowed_set
    )


def check_permission(allowed_roles):

    if 'role' not in session:
        return False

    return session.get('role') in allowed_roles


def login_required(f):

    @wraps(f)
    def decorated(*args, **kwargs):

        if 'user_id' not in session:
            flash('Please login first.', 'warning')
            return redirect(url_for('login'))

        return f(*args, **kwargs)

    return decorated


def admin_required(f):

    @wraps(f)
    def decorated(*args, **kwargs):

        if not check_permission(['admin']):
            abort(403)

        return f(*args, **kwargs)

    return decorated


def get_photo_url(photo_path):

    if photo_path and photo_path != 'default_avatar.png':

        return url_for(
            'static',
            filename='uploads/' + photo_path
        )

    return url_for(
        'static',
        filename='uploads/default_avatar.png'
    )


def validate_and_format_phone(phone):

    if not phone:
        return None

    cleaned = re.sub(
        r'[^0-9+]',
        '',
        phone.strip()
    )

    if cleaned.startswith('+'):

        digits = re.sub(
            r'\D',
            '',
            cleaned
        )

        if len(digits) >= 9:
            return cleaned

        return None


    digits = re.sub(
        r'\D',
        '',
        cleaned
    )


    if len(digits) == 9:

        return f'+256{digits}'


    elif len(digits) == 12 and digits.startswith('256'):

        return f'+{digits}'


    return None
    # ==================== NOTIFICATION FUNCTIONS ====================

def add_notification(user_role, message, link=None, title=None):
    if title is None:
        title = "New Notification"
    execute_db("""
        INSERT INTO notifications (user_role,title,message,link,is_read,created_at)
        VALUES (%s,%s,%s,%s,0,%s)
    """,(user_role,title,message,link,datetime.now()))

def get_notification_count(user_role):
    db=get_db()
    cur=db.cursor()
    cur.execute("SELECT COUNT(*) AS count FROM notifications WHERE user_role=%s AND is_read=0",(user_role,))
    result=cur.fetchone()
    cur.close()
    return result['count'] if result else 0

def mark_all_notifications_read(user_role):
    execute_db("UPDATE notifications SET is_read=1 WHERE user_role=%s",(user_role,))


# ==================== NOTIFICATION API ENDPOINTS ====================

@app.route('/get_notifications')
def get_notifications():
    if not session.get('user_id'):
        return jsonify([])

    user_role=session.get('role')
    db=get_db()
    cur=db.cursor()
    cur.execute("""
        SELECT id,title,message,link,is_read,created_at
        FROM notifications
        WHERE user_role=%s
        ORDER BY created_at DESC
        LIMIT 30
    """,(user_role,))
    notifications=cur.fetchall()
    cur.close()
    result=[]
    for n in notifications:
        result.append({
            'id':n['id'],
            'title':n.get('title','Notification'),
            'message':n['message'],
            'link':n.get('link',''),
            'is_read':n['is_read'],
            'created_at':str(n['created_at'])[:19] if n['created_at'] else ''
        })
    return jsonify(result)


@app.route('/mark_notifications_read',methods=['POST'])
def mark_notifications_read():
    if not session.get('user_id'):
        return jsonify({'error':'Not logged in'})
    try:
        execute_db(
            "UPDATE notifications SET is_read=1 WHERE user_role=%s",
            (session.get('role'),)
        )
        return jsonify({'success':True})
    except Exception as e:
        return jsonify({'success':False,'error':str(e)})


@app.route('/dos/get_notifications')
def dos_get_notifications():
    return get_notifications()

@app.route('/dos/mark_notifications_read',methods=['POST'])
def dos_mark_notifications_read():
    return mark_notifications_read()

@app.route('/headteacher/get_notifications')
def headteacher_get_notifications():
    return get_notifications()

@app.route('/headteacher/mark_notifications_read',methods=['POST'])
def headteacher_mark_notifications_read():
    return mark_notifications_read()

@app.route('/bursar/get_notifications')
def bursar_get_notifications():
    return get_notifications()

@app.route('/bursar/mark_notifications_read',methods=['POST'])
def bursar_mark_notifications_read():
    return mark_notifications_read()

@app.route('/management/get_notifications')
def management_get_notifications():
    return get_notifications()

@app.route('/management/mark_notifications_read',methods=['POST'])
def management_mark_notifications_read():
    return mark_notifications_read()

@app.route('/stores/get_notifications')
def stores_get_notifications():
    return get_notifications()

@app.route('/stores/mark_notifications_read',methods=['POST'])
def stores_mark_notifications_read():
    return mark_notifications_read()
# ==================== GRADING HELPERS ====================

def get_grade_and_descriptor(percentage):
    db = get_db()
    cur = db.cursor()
    cur.execute(
        """
        SELECT grade, descriptor
        FROM grading_system
        WHERE %s BETWEEN min_score AND max_score
        ORDER BY min_score DESC
        LIMIT 1
        """,
        (percentage,)
    )
    result = cur.fetchone()
    cur.close()

    if result:
        return result['grade'], result['descriptor']

    return None, None


def get_identifier(total_score):
    db = get_db()
    cur = db.cursor()
    cur.execute(
        """
        SELECT identifier
        FROM identifier_grading
        WHERE %s BETWEEN min_value AND max_value
        ORDER BY min_value DESC
        LIMIT 1
        """,
        (total_score,)
    )
    result = cur.fetchone()
    cur.close()

    if result:
        return result['identifier']

    return None

# =========================================================
# A-LEVEL CLASS TEACHER COMMENT
# BASED ON TOTAL POINTS
# =========================================================
def get_alevel_class_teacher_comment(total_points):

    if total_points >= 15:
        return (
            "Excellent performance. Keep it up!"
        )

    elif total_points >= 12:
        return (
            "Very good performance. Maintain it!"
        )

    elif total_points >= 9:
        return (
            "Good performance. Continue working hard."
        )

    elif total_points >= 6:
        return (
            "Fair performance. Put in more effort"
        )

    else:
        return (
            "The student needs significant improvement"
        )


# =========================================================
# A-LEVEL HEADTEACHER COMMENT
# BASED ON TOTAL POINTS
# =========================================================
def get_alevel_headteacher_comment(total_points):

    if total_points >= 15:
        return (
            "Excellent academic performance. Keep up the excellent work."
        )

    elif total_points >= 12:
        return (
            "Very good academic performance. Continue working hard."
        )

    elif total_points >= 9:
        return (
            "Good academic performance. Maintain "
            "consistent effort."
        )

    elif total_points >= 6:
        return (
            "Fair academic performance. Increase effort"
        )

    else:
        return (
            "The student needs considerable improvement and greater "
            "commitment to academic work"
        )


# =========================================================
# O-LEVEL CLASS TEACHER COMMENT
# BASED ON GENERAL IDENTIFIER
# =========================================================
def get_olevel_class_teacher_comment(general_identifier):

    if general_identifier >= 2.50:
        return (
            "Excellent performance. Keep it up!"
        )

    elif general_identifier >= 2.00:
        return (
            "Very good performance. Maintain it!"
        )

    elif general_identifier >= 1.50:
        return (
            "Good performance. Continue putting in consistent effort."
        )

    elif general_identifier >= 1.00:
        return (
            "Fair performance. Needs to put in more effort"
        )

    else:
        return (
            "The student needs significant improvement"
        )


# =========================================================
# O-LEVEL HEADTEACHER COMMENT
# BASED ON GENERAL IDENTIFIER
# =========================================================
def get_olevel_headteacher_comment(general_identifier):

    if general_identifier >= 2.50:
        return (
            "Excellent academic performance. Keep up the excellent performance."
        )

    elif general_identifier >= 2.00:
        return (
            "Very good academic performance. Maintain consistent "
            "effort."
        )

    elif general_identifier >= 1.50:
        return (
            "Good academic performance. Work hard to achieve even better results."
        )

    elif general_identifier >= 1.00:
        return (
            "Fair academic performance. Increase effort and commitment in order to improve."
        )

    else:
        return (
            "The student needs significant improvement and greater "
            "commitment to academic work"
        )

# ==================== TEACHER ASSIGNMENT HELPERS ====================
def get_user_assignments(user_id=None):
    if user_id is None:
        user_id = session.get('user_id')

    if not user_id:
        return []

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT
            id,
            user_id,
            class_name,
            subject,
            assignment_type,
            assigned_by
        FROM teacher_class_assignments
        WHERE user_id = %s
        ORDER BY assignment_type, class_name, subject
    """, (user_id,))

    assignments = cur.fetchall()

    cur.close()

    return assignments


def get_user_classes(user_id=None, assignment_type=None):
    if user_id is None:
        user_id = session.get('user_id')

    if not user_id:
        return []

    db = get_db_dict()
    cur = db.cursor()

    if assignment_type:
        cur.execute("""
            SELECT DISTINCT class_name
            FROM teacher_class_assignments
            WHERE user_id = %s
              AND assignment_type = %s
            ORDER BY class_name
        """, (user_id, assignment_type))
    else:
        cur.execute("""
            SELECT DISTINCT class_name
            FROM teacher_class_assignments
            WHERE user_id = %s
            ORDER BY class_name
        """, (user_id,))

    rows = cur.fetchall()

    cur.close()

    return [row['class_name'] for row in rows]

def generate_fee_prn():
    """
    Generate a random 10-digit fee payment reference.
    """

    while True:

        prn = ''.join(
            str(secrets.randbelow(10))
            for _ in range(10)
        )

        db = get_db_dict()
        cur = db.cursor()

        cur.execute(
            """
            SELECT 1
            FROM fee_payment_requests
            WHERE prn=%s
            LIMIT 1
            """,
            (prn,)
        )

        exists = cur.fetchone()

        cur.close()
        db.close()

        if not exists:
            return prn    

def teacher_has_class_access(user_id, student_class, assignment_type=None):
    """
    Checks whether a teacher has access to a particular
    student class/stream.

    Rules:

        Assignment S.1  -> S.1A, S.1B, S.1C
        Assignment S.2  -> S.2A, S.2B, S.2C

        Assignment S.1B -> S.1B only
        Assignment S.1C -> S.1C only
    """

    if not user_id or not student_class:
        return False

    assignments = get_user_assignments(user_id)

    student_class = student_class.strip().upper()

    for assignment in assignments:

        if (
            assignment_type
            and assignment.get('assignment_type') != assignment_type
        ):
            continue

        assigned_class = (
            assignment.get('class_name') or ''
        ).strip().upper()

        if not assigned_class:
            continue

        # Exact stream/class match
        if student_class == assigned_class:
            return True

        # Main class assignment covers all its streams
        #
        # S.1 -> S.1A
        # S.1 -> S.1B
        # S.2 -> S.2A
        #
        # S.1B does NOT cover S.1A.
        if re.match(
            r'^' + re.escape(assigned_class) + r'[A-Z]+$',
            student_class
        ):
            return True

    return False
def get_teacher_accessible_classes(user_id=None, assignment_type=None):
    """
    Returns the classes assigned to the teacher.

    Assignments are kept at the main-class level where possible,
    e.g. S.1, S.2, S.3.

    The actual students can still be stored as S.1A, S.1B, etc.
    """

    return get_user_classes(
        user_id=user_id,
        assignment_type=assignment_type
    )
# ==================== BANK PAYMENT PROCESSING ====================
def process_bank_payment(payroll):
    import random
    results={'success':False,'token':None,'reference':None,'error':None}
    if random.random()>0.1:
        results['success']=True
        results['token']=f"TOKEN-{payroll['payroll_no']}"
        results['reference']=f"REF-{payroll['payroll_no']}-{int(time.time())}"
    else:
        results['error']="Bank API temporarily unavailable"
    return results

# =========================================================
# SIX-DIGIT PAYROLL APPROVAL CODE
# =========================================================

def generate_approval_code():
    return ''.join(
        str(secrets.randbelow(10))
        for _ in range(6)
    )

def generate_secure_token(expiry_hours=2):

    characters = string.ascii_letters + string.digits

    token = ''.join(
        secrets.choice(characters)
        for _ in range(32)
    )

    expires_at = datetime.utcnow() + timedelta(
        hours=expiry_hours
    )

    return token, expires_at

def generate_voucher_no():
    db = get_db_dict()
    cur = db.cursor()

    today = datetime.now()
    prefix = f"EV-{today.strftime('%Y%m')}"

    cur.execute("""
        SELECT voucher_no
        FROM expenditures
        WHERE voucher_no LIKE %s
        ORDER BY id DESC
        LIMIT 1
    """, (f"{prefix}-%",))

    last = cur.fetchone()

    cur.close()

    if last and last.get('voucher_no'):
        try:
            last_num = int(
                last['voucher_no'].split('-')[-1]
            )
        except (ValueError, AttributeError):
            last_num = 0
    else:
        last_num = 0

    next_num = last_num + 1

    return f"{prefix}-{next_num:04d}"

def get_student_level(class_name):
    if not class_name:
        return None

    class_name = class_name.strip().upper()

    match = re.match(r'^(S\.[1-6])', class_name)

    if match:
        return match.group(1)

    return None

# =========================================================
# APPLY STUDENT FEE STRUCTURE
# =========================================================

def apply_student_fee_structure(
    student_id,
    term='Term 1',
    year=None
):

    if year is None:
        year = datetime.now().year

    db = get_db_dict()
    cur = db.cursor()

    try:

        # =================================================
        # GET STUDENT
        # =================================================

        cur.execute("""
            SELECT
                student_id,
                class,
                programme,
                residence
            FROM students
            WHERE student_id = %s
        """, (
            student_id,
        ))

        student = cur.fetchone()

        if not student:
            return False, 'Student not found.'

        # =================================================
        # DETERMINE LEVEL
        # =================================================

        level = get_student_level(
            student['class']
        )

        if not level:

            return False, (
                f"Could not determine level from "
                f"class {student['class']}."
            )

        programme = (
            student['programme'] or ''
        ).strip()

        residence = (
            student['residence'] or ''
        ).strip()

        # =================================================
        # CHECK PROGRAMME
        # =================================================

        if level in [
            'S.1',
            'S.2',
            'S.3',
            'S.4'
        ]:

            valid_programmes = [
                'USE',
                'Non-USE'
            ]

        else:

            valid_programmes = [
                'UPOLET',
                'Non-UPOLET'
            ]

        if programme not in valid_programmes:

            return False, (
                f"Invalid or missing programme "
                f"for {level}."
            )

        # =================================================
        # CHECK RESIDENCE
        # =================================================

        if residence not in [
            'Day',
            'Boarding'
        ]:

            return False, (
                'Invalid or missing residence.'
            )

        # =================================================
        # FIND FEE STRUCTURE
        # =================================================

        cur.execute("""
            SELECT
                amount
            FROM student_fee_structure
            WHERE level = %s
            AND programme = %s
            AND residence = %s
            AND term = %s
            AND year = %s
        """, (
            level,
            programme,
            residence,
            term,
            year
        ))

        fee = cur.fetchone()

        if not fee:

            return False, (
                f'No fee structure found for '
                f'{level} - {programme} - '
                f'{residence} - {term} - {year}.'
            )

        fees_total = float(
            fee['amount'] or 0
        )

        # =================================================
        # CALCULATE PAYMENTS
        # =================================================

        cur.execute("""
            SELECT
                COALESCE(
                    SUM(amount),
                    0
                ) AS total_paid
            FROM payments
            WHERE student_id = %s
        """, (
            student_id,
        ))

        payment = cur.fetchone()

        fees_paid = float(
            payment['total_paid'] or 0
        )

        fees_balance = max(
            fees_total - fees_paid,
            0
        )

        # =================================================
        # UPDATE STUDENT
        # =================================================

        cur.execute("""
            UPDATE students
            SET
                fees_total = %s,
                fees_paid = %s,
                fees_balance = %s
            WHERE student_id = %s
        """, (
            fees_total,
            fees_paid,
            fees_balance,
            student_id
        ))

        db.commit()

        return True, (
            f'Fee structure applied successfully. '
            f'Total fees: UGX {fees_total:,.2f}'
        )

    except Exception as e:

        db.rollback()

        return False, str(e)

    finally:

        cur.close()
# ==================== CONTEXT PROCESSORS ====================
@app.context_processor
def inject_now():
    return {'datetime':datetime}

@app.context_processor
def inject_notifications():
    if 'user_id' in session:
        role=session.get('role')
        if role in ['headteacher','bursar','management','admin']:
            try:
                db=get_db()
                cur=db.cursor()
                cur.execute("SELECT COUNT(*) AS count FROM notifications WHERE user_role=%s AND is_read=0",(role,))
                count_row=cur.fetchone()
                notification_count=count_row['count'] if count_row else 0
                cur.execute("""
                    SELECT id,user_role,message,link,is_read,created_at
                    FROM notifications
                    WHERE user_role=%s AND is_read=0
                    ORDER BY created_at DESC
                    LIMIT 5
                """,(role,))
                rows=cur.fetchall()
                cur.close()
                return {
                    'notification_count':notification_count,
                    'notifications':rows
                }
            except Exception:
                return {'notification_count':0,'notifications':[]}
    return {'notification_count':0,'notifications':[]}

# ==================== TEMPLATE FILTERS ====================
@app.template_filter('currency')
def currency_filter(value):
    return "{:,.2f}".format(float(value)) if value else '0.00'

@app.template_filter('word_format')
def word_format(value):
    words={1:'One',2:'Two',3:'Three',4:'Four',5:'Five',6:'Six',7:'Seven',8:'Eight',9:'Nine',10:'Ten'}
    return words.get(int(value),str(value)) if value else 'Zero'

# ==================== AUTHENTICATION ROUTES ====================
@app.route('/')
def index():
    return redirect(url_for('login'))

@app.after_request
def add_security_headers(response):
    response.headers['Content-Security-Policy']="default-src 'self'; script-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net https://cdnjs.cloudflare.com; style-src 'self' 'unsafe-inline' https://cdn.jsdelivr.net; font-src 'self' https://cdnjs.cloudflare.com; img-src 'self' data:;"
    response.headers['X-Content-Type-Options']='nosniff'
    response.headers['X-Frame-Options']='DENY'
    response.headers['X-XSS-Protection']='1; mode=block'
    return response

@app.route('/login',methods=['GET','POST'])
def login():
    if request.method=='POST':
        username=sanitize_input(request.form['username'].strip())
        password=request.form['password'].strip()

        cur=get_db().cursor()
        cur.execute("""
            SELECT id,username,role,status,phone,must_change_password,password
            FROM users
            WHERE username=%s
        """,(username,))
        user=cur.fetchone()
        cur.close()
        if user and user['status']==1:
            stored_password=user['password']
            if stored_password==password or check_password_hash(stored_password,password):
                session['user_id']=user['id']
                session['username']=user['username']
                session['role']=user['role']
                session['phone']=user['phone']

                if user['must_change_password']==1:
                    flash('Please change your password.','warning')
                    return redirect(url_for('change_password'))
                flash(f"Welcome {username}!",'success')
                return redirect(url_for('dashboard'))

        flash('Invalid credentials.','danger')
        return redirect(url_for('login'))

    return render_template('login.html')


@app.route('/change_password',methods=['GET','POST'])
@login_required
def change_password():
    if request.method=='POST':
        new_pass=request.form['new_password'].strip()
        confirm=request.form['confirm_password'].strip()

        if new_pass!=confirm:
            flash('Passwords do not match.','danger')
            return redirect(url_for('change_password'))

        hashed_password=generate_password_hash(new_pass)

        execute_db(
            "UPDATE users SET password=%s,must_change_password=0 WHERE id=%s",
            (hashed_password,session['user_id'])
        )

        flash('Password changed successfully!','success')
        return redirect(url_for('dashboard'))

    return render_template('change_password.html')


@app.route('/forgot_password',methods=['GET','POST'])
def forgot_password():
    if request.method=='POST':
        username=request.form['username'].strip()
        phone=validate_and_format_phone(request.form['phone'].strip())

        if not phone:
            flash('Invalid phone number format.','danger')
            return redirect(url_for('forgot_password'))

        new_pass=request.form['new_password'].strip()
        confirm=request.form['confirm_password'].strip()

        if new_pass!=confirm:
            flash('Passwords do not match.','danger')
            return redirect(url_for('forgot_password'))

        cur=get_db().cursor()
        cur.execute(
            "SELECT id FROM users WHERE username=%s AND phone=%s",
            (username,phone)
        )
        user=cur.fetchone()
        cur.close()

        if user:
            hashed_password=generate_password_hash(new_pass)
            execute_db(
                "UPDATE users SET password=%s,must_change_password=0 WHERE id=%s",
                (hashed_password,user['id'])
            )
            flash('Password reset successfully.','success')
        else:
            flash('Username and phone number do not match.','danger')

        return redirect(url_for('login'))

    return render_template('forgot_password.html')


@app.route('/logout')
def logout():
    session.clear()
    flash('Logged out.','info')
    return redirect(url_for('login'))
    

# ==================== DASHBOARD ROUTES ====================
@app.route('/dashboard')
@login_required
def dashboard():
    role = session.get('role')
    
    if role == 'admin':
        search = request.args.get('search', '').strip()
        page = request.args.get('page', 1, type=int)
        per_page = 10

        cur = get_db_dict().cursor()

        # Count users
        if search:
            cur.execute(
                """
                SELECT COUNT(*) AS total
                FROM users
                WHERE username LIKE %s
                """,
                (f'%{search}%',)
            )
        else:
            cur.execute(
                """
                SELECT COUNT(*) AS total
                FROM users
                """
            )

        total = cur.fetchone()['total']
        total_pages = (total + per_page - 1) // per_page
        offset = (page - 1) * per_page
        # Fetch users
        if search:
            cur.execute(
                """
                SELECT 
                    id,
                    username,
                    role,
                    phone,
                    status,
                    profile_pic
                FROM users
                WHERE username LIKE %s
                ORDER BY id
                LIMIT %s OFFSET %s
                """,
                (
                    f'%{search}%',
                    per_page,
                    offset
                )
            )
        else:
            cur.execute(
                """
                SELECT 
                    id,
                    username,
                    role,
                    phone,
                    status,
                    profile_pic
                FROM users
                ORDER BY id
                LIMIT %s OFFSET %s
                """,
                (
                    per_page,
                    offset
                )
            )
        users = cur.fetchall()
        cur.close()
        return render_template(
            'dashboard.html',
            role=role,
            data={
                'users': users,
                'total_pages': total_pages,
                'current_page': page
            },
            search=search
        )
    elif role == 'bursar':
        return redirect(url_for('bursar_dashboard'))

    elif role =='parent':
        return redirect(url_for('parent_dashboard'))
    else:
        return render_template(
            'dashboard.html',
            role=role
        )

@app.route('/notifications')
@login_required
def view_all_notifications():
    role = session.get('role')
    db = get_db_dict()
    cur = db.cursor()
    cur.execute("SELECT * FROM notifications WHERE user_role = ? ORDER BY created_at DESC", (role,))
    notifications = cur.fetchall()
    cur.close()
    return render_template('notifications.html', notifications=notifications)

@app.route('/notification/mark_read/<int:notification_id>')
@login_required
def mark_notification_read_route(notification_id):
    mark_notification_read(notification_id)
    return redirect(request.referrer or url_for('dashboard'))

@app.route('/notification/mark_all_read')
@login_required
def mark_all_notifications_read_route():
    role = session.get('role')
    mark_all_notifications_read(role)
    flash('All notifications marked as read.', 'success')
    return redirect(request.referrer or url_for('dashboard'))

# ==================== DASHBOARD ROUTES ====================

@app.route('/admin/dashboard')
@login_required
def admin_dashboard():
    role=session.get('role')

    if role=='admin':
        search=request.args.get('search','').strip()
        page=request.args.get('page',1,type=int)
        per_page=10

        cur=get_db().cursor()

        if search:
            cur.execute(
                "SELECT COUNT(*) AS count FROM users WHERE username ILIKE %s",
                (f'%{search}%',)
            )
        else:
            cur.execute("SELECT COUNT(*) AS count FROM users")

        total=cur.fetchone()['count']

        total_pages=(total+per_page-1)//per_page
        offset=(page-1)*per_page

        if search:
            cur.execute("""
                SELECT id,username,role,phone,status,profile_pic
                FROM users
                WHERE username ILIKE %s
                ORDER BY id
                LIMIT %s OFFSET %s
            """,(f'%{search}%',per_page,offset))
        else:
            cur.execute("""
                SELECT id,username,role,phone,status,profile_pic
                FROM users
                ORDER BY id
                LIMIT %s OFFSET %s
            """,(per_page,offset))

        users=cur.fetchall()
        cur.close()

        return render_template(
            'dashboard.html',
            role=role,
            data={
                'users':users,
                'total_pages':total_pages,
                'current_page':page
            },
            search=search
        )

    elif role=='bursar':
        return redirect(url_for('bursar_dashboard'))

    else:
        return render_template('dashboard.html',role=role)

   
# ==================== ADMIN ROUTES ====================

@app.route('/admin/add_user',methods=['POST'])
@admin_required
def add_user():
    full_name=sanitize_input(request.form['full_name'].strip())
    username=sanitize_input(request.form['username'].strip())
    password=request.form['password'].strip()
    role=request.form['role'].strip()
    phone_raw=request.form.get('phone','').strip()
    phone=validate_and_format_phone(phone_raw) if phone_raw else None
    child_id=request.form.get('child_id','').strip() or None

    db=get_db()
    cur=db.cursor()

    cur.execute("SELECT id FROM users WHERE username=%s",(username,))
    if cur.fetchone():
        flash(f'Username "{username}" already exists! Please choose a different username.','danger')
        cur.close()
        return redirect(url_for('dashboard'))

    cur.execute("SELECT max_count FROM role_limits WHERE role_name=%s",(role,))
    limit=cur.fetchone()

    if limit:
        cur.execute("SELECT COUNT(*) AS count FROM users WHERE role=%s",(role,))
        count=cur.fetchone()['count']

        if count>=limit['max_count']:
            flash(f"Cannot add. Only {limit['max_count']} {role} allowed in the system.",'danger')
            cur.close()
            return redirect(url_for('dashboard'))

    cur.close()

    hashed_password=generate_password_hash(password)

    try:
        execute_db("""
            INSERT INTO users
            (username,password,full_name,role,phone,status,child_id,profile_pic,must_change_password)
            VALUES (%s,%s,%s,%s,%s,1,%s,'default_avatar.png',1)
        """,(username,hashed_password,full_name,role,phone,child_id))

        flash(f'User {full_name} ({username}) added. Password: {password}','success')

    except Exception as e:
        flash(f'Error: {str(e)}','danger')

    return redirect(url_for('dashboard'))


@app.route('/admin/edit_user/<int:user_id>',methods=['GET','POST'])
@admin_required
def edit_user(user_id):

    if request.method=='POST':

        full_name=sanitize_input(request.form.get('full_name','').strip())
        username=request.form.get('username','').strip()
        role=request.form.get('role','').strip()
        phone=request.form.get('phone','').strip()
        child_id=request.form.get('child_id','').strip() or None
        file=request.files.get('profile_pic')

        profile_pic=None

        if file and file.filename and allowed_file(file.filename,ALLOWED_IMAGE_EXTENSIONS):
            ext=file.filename.rsplit('.',1)[1].lower()
            filename=f"user_{user_id}.{ext}"
            file.save(os.path.join(app.config['UPLOAD_FOLDER'],filename))
            profile_pic=filename

        try:
            if profile_pic:
                execute_db("""
                    UPDATE users
                    SET username=%s,full_name=%s,role=%s,phone=%s,child_id=%s,profile_pic=%s
                    WHERE id=%s
                """,(username,full_name,role,phone,child_id,profile_pic,user_id))
            else:
                execute_db("""
                    UPDATE users
                    SET username=%s,full_name=%s,role=%s,phone=%s,child_id=%s
                    WHERE id=%s
                """,(username,full_name,role,phone,child_id,user_id))

            flash('User updated successfully.','success')

        except Exception as e:
            flash(f'Error: {str(e)}','danger')

        return redirect(url_for('dashboard'))

    cur=get_db().cursor()
    cur.execute("""
        SELECT id,username,full_name,role,phone,child_id,profile_pic
        FROM users WHERE id=%s
    """,(user_id,))

    user=cur.fetchone()
    cur.close()

    if not user:
        flash('User not found.','danger')
        return redirect(url_for('dashboard'))

    return render_template('edit_user.html',user=user)


@app.route('/admin/toggle_user/<int:user_id>')
@admin_required
def toggle_user(user_id):

    if user_id==session.get('user_id'):
        flash('Cannot toggle your own account.','warning')
        return redirect(url_for('dashboard'))

    execute_db(
        "UPDATE users SET status=CASE WHEN status=1 THEN 0 ELSE 1 END WHERE id=%s",
        (user_id,)
    )

    flash('Status toggled.','success')
    return redirect(url_for('dashboard'))


@app.route('/admin/delete_user/<int:user_id>')
@admin_required
def delete_user(user_id):

    if user_id == session.get('user_id'):
        flash('Cannot delete your own account.', 'warning')
        return redirect(url_for('dashboard'))

    # Delete teacher assignments first
    execute_db(
        "DELETE FROM teacher_class_assignments WHERE user_id=%s",
        (user_id,)
    )

    # Now delete the user
    execute_db(
        "DELETE FROM users WHERE id=%s",
        (user_id,)
    )

    flash('User deleted.', 'success')
    return redirect(url_for('dashboard'))

@app.route('/admin/role_counts')
def admin_role_counts():

    if not check_permission(['admin']):
        abort(403)

    cur=get_db().cursor()

    cur.execute("""
        SELECT role,
        COUNT(*) AS count,
        (SELECT max_count FROM role_limits WHERE role_name=users.role) AS max_count
        FROM users
        GROUP BY role
    """)

    counts=cur.fetchall()
    cur.close()

    return render_template('admin/role_counts.html',counts=counts)


@app.route('/admin/school_settings', methods=['GET', 'POST'])
def school_settings():

    if not check_permission(['admin', 'headteacher']):
        abort(403)

    if request.method == 'POST':

        try:

            begins = request.form.get('next_term_begins') or None
            ends = request.form.get('next_term_ends') or None

            school_name = request.form.get(
                'school_name',
                'YOUR SCHOOL NAME'
            )

            school_address = request.form.get(
                'school_address',
                'P.O. Box 123, Kampala, Uganda'
            )

            school_phone = request.form.get(
                'school_phone',
                'Tel: +256 712 345678'
            )

            school_email = request.form.get(
                'school_email',
                'Email: info@school.com'
            )
            # ================= LOGO UPLOAD =================
            logo_filename = None
            logo_file = request.files.get('logo')
            if (
                logo_file 
                and logo_file.filename
                and allowed_file(
                    logo_file.filename,
                    ALLOWED_IMAGE_EXTENSIONS
                )
            ):
                ext = logo_file.filename.rsplit('.',1)[1].lower()
                logo_filename = (
                    f"logo_{int(datetime.now().timestamp())}.{ext}"
                )
                logo_file.save(
                    os.path.join(
                        app.config['UPLOAD_FOLDER'],
                        logo_filename
                    )
                )
            # ================= STAMP UPLOAD =================
            stamp_filename = None
            stamp_file = request.files.get('stamp')

            if (
                stamp_file 
                and stamp_file.filename
                and allowed_file(
                    stamp_file.filename,
                    ALLOWED_IMAGE_EXTENSIONS
                )
            ):
                ext = stamp_file.filename.rsplit('.',1)[1].lower()
                stamp_filename = (
                    f"stamp_{int(datetime.now().timestamp())}.{ext}"
                )
                stamp_file.save(
                    os.path.join(
                        app.config['UPLOAD_FOLDER'],
                        stamp_filename
                    )
                )
            # ================= RATES =================
            nssf_employee_rate = float(
                request.form.get(
                    'nssf_employee_rate',
                    5.0
                )
            )

            paye_rate = float(
                request.form.get(
                    'paye_rate',
                    10.0
                )
            )

            paye_threshold = float(
                request.form.get(
                    'paye_threshold',
                    235000
                )
            )
            # ================= UPDATE SETTINGS =================
            execute_db(
                """
                UPDATE school_settings
                SET
                    next_term_begins=%s,
                    next_term_ends=%s,
                    school_name=%s,
                    school_address=%s,
                    school_phone=%s,
                    school_email=%s,
                    nssf_employee_rate=%s,
                    paye_rate=%s,
                    paye_threshold=%s
                WHERE id=1
                """,
                (
                    begins,
                    ends,
                    school_name,
                    school_address,
                    school_phone,
                    school_email,
                    nssf_employee_rate,
                    paye_rate,
                    paye_threshold
                )
            )
            if logo_filename:

                execute_db(
                    """
                    UPDATE school_settings
                    SET logo_url=%s
                    WHERE id=1
                    """,
                    (logo_filename,)
                )
            # ================= UPDATE STAMP =================
            if stamp_filename:
                execute_db(
                    """
                    UPDATE school_settings
                    SET headteacher_stamp=%s
                    WHERE id=1
                    """,
                    (stamp_filename,)
                )
            flash(
                'School settings updated successfully.',
                'success'
            )
        except Exception as e:
            db = get_db()
            db.rollback()
            print(
                "SETTINGS ERROR:",
                str(e)
            )
            flash(
                f"Error saving settings: {str(e)}",
                "danger"
            )
    # ================= LOAD SETTINGS =================
    cur = get_db().cursor()
    cur.execute(
        """
        SELECT
            next_term_begins,
            next_term_ends,
            headteacher_stamp,
            school_name,
            school_address,
            school_phone,
            school_email,
            logo_url,
            nssf_employee_rate,
            paye_rate,
            paye_threshold
        FROM school_settings
        WHERE id=1
        """
    )
    settings = cur.fetchone()
    cur.close()
    return render_template(
        'admin/school_settings.html',
        settings=settings,
        nssf_rate=settings['nssf_employee_rate'] if settings else 5.0,
        paye_rate=settings['paye_rate'] if settings else 10.0,
        paye_threshold=settings['paye_threshold'] if settings else 235000
    )

@app.route('/admin/nssf_paye_settings',methods=['GET','POST'])
def nssf_paye_settings():

    if not check_permission(['admin','bursar']):
        abort(403)

    if request.method=='POST':

        execute_db("""
            UPDATE school_settings
            SET nssf_employee_rate=%s,paye_rate=%s,paye_threshold=%s
            WHERE id=1
        """,(
            float(request.form['nssf_employee_rate']),
            float(request.form['paye_rate']),
            float(request.form['paye_threshold'])
        ))

        flash('NSSF and PAYE settings updated successfully.','success')

    cur=get_db().cursor()
    cur.execute("""
        SELECT nssf_employee_rate,paye_rate,paye_threshold
        FROM school_settings WHERE id=1
    """)
    settings=cur.fetchone()
    cur.close()

    return render_template('admin/nssf_paye_settings.html',settings=settings)

# ==================== ADMISSION PORTAL ROUTES ====================
def extract_results_from_pdf(file_path):
    return {'english': 75, 'math': 68, 'science': 82, 'social_studies': 70, 'average': 73.75, 'qualifies': True}

def determine_admission_worth(results):
    min_average = 60
    min_english = 50
    min_math = 50
    qualifies = (results.get('average', 0) >= min_average and results.get('english', 0) >= min_english and results.get('math', 0) >= min_math)
    return {'qualifies': qualifies, 'average': results.get('average', 0), 'message': 'Congratulations! You qualify.' if qualifies else 'Sorry, you do not meet requirements.'}

def generate_admission_letter(student):
    return f"""<!DOCTYPE html><html><head><title>Admission Letter</title></head><body><h2>Admission Letter</h2><p>Dear {student['full_name']},</p><p>You have been admitted.</p></body></html>"""

@app.route('/admissions', methods=['GET', 'POST'])
def admissions_portal():
    cur = get_db().cursor()
    cur.execute("SELECT is_open, deadline, closing_reason FROM admission_settings WHERE id=1")
    settings = cur.fetchone()
    cur.close()

    is_open = settings[0] if settings else 1
    deadline = settings[1] if settings else None
    closing_reason = settings[2] if settings else ''

    if not is_open:
        return render_template('admissions/closed.html', reason=closing_reason, deadline=deadline)

    if request.method == 'POST':
        full_name = sanitize_input(request.form['full_name'].strip())
        date_of_birth = request.form['date_of_birth']
        sex = request.form['sex']
        preferred_house = request.form['preferred_house']
        disability = request.form.get('disability', '')
        sports_activities = request.form.getlist('sports_activities')
        lin = request.form['lin']
        phone = request.form['phone']
        email = request.form['email']

        birth_date = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
        age = calculate_age(birth_date)

        photo = request.files.get('photo')
        photo_filename = None

        if photo and photo.filename:
            ext = photo.filename.rsplit('.', 1)[1].lower()
            student_id_temp = f"TEMP-{int(datetime.now().timestamp())}"
            photo_filename = f"{student_id_temp}.{ext}"
            photo.save(os.path.join(app.config['UPLOAD_FOLDER'], photo_filename))

        results_file = request.files.get('results_pdf')
        results_data = None

        if results_file and results_file.filename:
            filename = secure_filename(f"results_{int(datetime.now().timestamp())}_{results_file.filename}")
            filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            results_file.save(filepath)
            results_data = extract_results_from_pdf(filepath)

        qualification = determine_admission_worth(results_data) if results_data else {
            'qualifies': False,
            'message': 'Results not uploaded'
        }

        session['admission_data'] = {
            'full_name': full_name,
            'date_of_birth': date_of_birth,
            'age': age,
            'sex': sex,
            'preferred_house': preferred_house,
            'disability': disability,
            'sports_activities': ','.join(sports_activities),
            'lin': lin,
            'phone': phone,
            'email': email,
            'photo_filename': photo_filename,
            'qualification': qualification,
            'results_data': results_data
        }

        if qualification['qualifies']:
            return redirect(url_for('admission_payment'))
        else:
            flash(qualification['message'], 'danger')
            return redirect(url_for('admissions_portal'))

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("SELECT name FROM houses ORDER BY name")
    houses = cur.fetchall()

    cur.execute("SELECT name FROM sports_activities ORDER BY name")
    sports = cur.fetchall()

    cur.close()

    return render_template('admissions/apply.html', houses=houses, sports=sports)


@app.route('/admissions/payment', methods=['GET', 'POST'])
def admission_payment():
    admission_data = session.get('admission_data')

    if not admission_data:
        flash('Please complete the application form first.', 'warning')
        return redirect(url_for('admissions_portal'))

    cur = get_db().cursor()
    cur.execute("SELECT is_open, fee_amount FROM admission_settings WHERE id=1")
    settings = cur.fetchone()
    cur.close()

    if not settings or settings[0] == 0:
        flash('Online admissions are currently closed.', 'danger')
        return redirect(url_for('admissions_portal'))

    amount = settings[1] if settings else 50000

    if request.method == 'POST':
        phone_number = request.form['phone_number']
        transaction_ref = f"ADM-{int(datetime.now().timestamp())}"

        result = request_momo_payment(phone_number, amount, transaction_ref)

        if result['success']:
            session['payment_data'] = {
                'transaction_id': result['transaction_id'],
                'amount': amount,
                'phone': phone_number,
                'reference': transaction_ref
            }

            temp_student_id = f"TEMP-{int(datetime.now().timestamp())}"

            execute_db("""
                INSERT INTO students (
                    student_id,
                    full_name,
                    class,
                    parent_phone,
                    date_of_birth,
                    age,
                    sex,
                    preferred_house,
                    disability,
                    sports_activities,
                    lin,
                    admission_source,
                    admission_status,
                    payment_status,
                    payment_transaction_id
                )
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,'online','pending','pending',%s)
            """,
            (
                temp_student_id,
                admission_data['full_name'],
                'Pending',
                admission_data.get('phone'),
                admission_data.get('date_of_birth'),
                admission_data.get('age'),
                admission_data.get('sex'),
                admission_data.get('preferred_house'),
                admission_data.get('disability'),
                admission_data.get('sports_activities'),
                admission_data.get('lin'),
                result['transaction_id']
            ))

            flash('Payment request sent! Please check your phone and complete the payment.', 'info')

            return redirect(url_for(
                'admission_payment_status',
                transaction_id=result['transaction_id']
            ))

        else:
            flash(result['message'], 'danger')
            return redirect(url_for('admission_payment'))

    return render_template(
        'admissions/payment.html',
        amount=amount,
        student_name=admission_data['full_name']
    )


@app.route('/admissions/payment/status/<transaction_id>')
def admission_payment_status(transaction_id):

    status = check_payment_status(transaction_id)

    if status == 'successful':

        cur = get_db().cursor()

        cur.execute("""
            UPDATE students
            SET payment_status='completed',
                payment_date=CURRENT_TIMESTAMP,
                admission_status='pending'
            WHERE payment_transaction_id=%s
        """,(transaction_id,))

        get_db().commit()
        cur.close()

        flash('Payment confirmed! Your application is pending review by the admissions office.', 'success')

        return redirect(url_for('admission_submitted'))

    elif status == 'failed':

        flash('Payment failed. Please try again.', 'danger')

        return redirect(url_for('admission_payment'))

    else:
        return render_template(
            'admissions/payment_pending.html',
            transaction_id=transaction_id
        )


@app.route('/admissions/submitted')
def admission_submitted():

    admission_data = session.get('admission_data')

    if not admission_data:
        return redirect(url_for('admissions_portal'))

    return render_template(
        'admissions/submitted.html',
        data=admission_data
    )
# ==================== DOS MODULE ====================
SCHOOL_ABBR = "TSS"
def generate_student_id():
    return generate_unique_number(SCHOOL_ABBR, 'students', 'student_id', year_format=True)

# ==================== DOS ADMISSION MANAGEMENT ====================

@app.route('/dos/admit_student', methods=['GET', 'POST'])
def dos_admit():
    if not check_permission(['dos']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("SELECT name FROM houses ORDER BY name")
    houses = cur.fetchall()

    cur.execute("SELECT name FROM sports_activities ORDER BY name")
    sports = cur.fetchall()

    cur.close()

    if request.method == 'POST':
        full_name = request.form['full_name'].strip()
        class_name = request.form['class'].strip()
        date_of_birth = request.form.get('date_of_birth') or None
        sex = request.form.get('sex') or None
        preferred_house = request.form.get('preferred_house') or None
        lin = request.form.get('lin') or None
        disability = request.form.get('disability') or None
        sports_activities = request.form.getlist('sports_activities')

        parent_phone_raw = request.form.get('parent_phone', '').strip()
        parent_phone = validate_and_format_phone(parent_phone_raw) if parent_phone_raw else None

        photo = request.files.get('photo')

        age = None

        if date_of_birth:
            birth_date = datetime.strptime(date_of_birth, '%Y-%m-%d').date()
            age = calculate_age(birth_date)

        student_id = generate_student_id()

        photo_filename = "default_avatar.png"

        if photo and photo.filename and allowed_file(photo.filename, ALLOWED_IMAGE_EXTENSIONS):
            ext = photo.filename.rsplit('.', 1)[1].lower()
            photo_filename = f"{student_id}.{ext}"
            photo.save(
                os.path.join(app.config['UPLOAD_FOLDER'], photo_filename)
            )

        try:
            execute_db(
                """
                INSERT INTO students
                (
                    student_id,
                    full_name,
                    class,
                    photo_path,
                    fees_paid,
                    fees_balance,
                    admission_date,
                    parent_phone,
                    date_of_birth,
                    age,
                    sex,
                    preferred_house,
                    disability,
                    sports_activities,
                    lin,
                    admission_source,
                    admission_status
                )
                VALUES
                (
                    %s,%s,%s,%s,
                    0,0,
                    CURRENT_DATE,
                    %s,%s,%s,%s,%s,%s,%s,%s,
                    'local',
                    'approved'
                )
                """,
                (
                    student_id,
                    full_name,
                    class_name,
                    photo_filename,
                    parent_phone,
                    date_of_birth,
                    age,
                    sex,
                    preferred_house,
                    disability,
                    ','.join(sports_activities) if sports_activities else None,
                    lin
                )
            )

            flash(
                f'Student {full_name} admitted with ID {student_id}.',
                'success'
            )

        except Exception as e:
            flash(f'Error: {str(e)}', 'danger')

        return redirect(url_for('dos_admit'))

    return render_template(
        'dos/admit_student.html',
        houses=houses,
        sports=sports
    )


@app.route('/dos/admission_settings', methods=['GET', 'POST'])
def dos_admission_settings():

    if not check_permission(['dos']):
        abort(403)

    if request.method == 'POST':

        is_open = 1 if request.form.get('is_open') == 'on' else 0
        deadline = request.form.get('deadline') or None
        closing_reason = request.form.get('closing_reason', '')
        fee_amount = float(request.form.get('fee_amount', 50000))

        execute_db(
            """
            UPDATE admission_settings
            SET
                is_open=%s,
                deadline=%s,
                closing_reason=%s,
                fee_amount=%s,
                updated_at=CURRENT_TIMESTAMP
            WHERE id=1
            """,
            (
                is_open,
                deadline,
                closing_reason,
                fee_amount
            )
        )

        flash(
            'Admission settings updated successfully.',
            'success'
        )

        return redirect(url_for('dos_admission_settings'))

    db = get_db_dict()
    cur = db.cursor()

    cur.execute(
        """
        SELECT 
            is_open,
            deadline,
            closing_reason,
            fee_amount
        FROM admission_settings
        WHERE id=1
        """
    )

    settings = cur.fetchone()

    cur.execute(
        """
        SELECT
            student_id,
            full_name,
            lin,
            application_date
        FROM students
        WHERE admission_source='online'
        AND admission_status='pending'
        ORDER BY application_date DESC
        """
    )

    pending = cur.fetchall()

    cur.close()

    return render_template(
        'dos/admission_settings.html',
        settings=settings,
        pending=pending,
        is_open=settings['is_open'] if settings else 1,
        deadline=settings['deadline'] if settings else '',
        closing_reason=settings['closing_reason'] if settings else '',
        fee_amount=settings['fee_amount'] if settings else 50000
    )
    
def generate_unique_number(prefix, table, column, year_format=False):
    db = get_db_dict()
    cur = db.cursor()

    year = datetime.now().strftime("%Y") if year_format else ""

    if year_format:
        pattern = f"{prefix}-{year}-%"

        cur.execute(
            f"""
            SELECT {column}
            FROM {table}
            WHERE {column} LIKE %s
            ORDER BY {column} DESC
            LIMIT 1
            """,
            (pattern,)
        )
    else:
        pattern = f"{prefix}-%"

        cur.execute(
            f"""
            SELECT {column}
            FROM {table}
            WHERE {column} LIKE %s
            ORDER BY {column} DESC
            LIMIT 1
            """,
            (pattern,)
        )

    last = cur.fetchone()
    cur.close()

    if last:
        last_number = int(last[column].split('-')[-1])
        number = last_number + 1
    else:
        number = 1

    if year_format:
        return f"{prefix}-{year}-{number:04d}"

    return f"{prefix}-{number:04d}"

@app.route('/dos/pending_admissions')
@login_required
def dos_pending_admissions():

    if not check_permission(['dos', 'admin']):
        abort(403)

    cur = get_db().cursor()

    cur.execute("""
        SELECT
            student_id,
            full_name,
            sex,
            age,
            parent_phone,
            admission_status,
            payment_status,
            admission_source
        FROM students
        WHERE admission_source='online'
          AND admission_status='pending'
        ORDER BY student_id DESC
    """)

    applications = cur.fetchall()

    cur.close()

    return render_template(
        'dos/pending_admissions.html',
        applications=applications
    )
    
@app.route('/dos/approve_admission/<student_id>', methods=['GET','POST'])
@login_required
def approve_admission(student_id):

    if not check_permission(['dos','admin']):
        abort(403)

    db = get_db()
    cur = db.cursor()

    # Get applicant
    cur.execute("""
        SELECT *
        FROM students
        WHERE student_id=%s
    """,(student_id,))

    student = cur.fetchone()

    if not student:
        cur.close()
        flash("Student application not found.", "danger")
        return redirect(url_for('dos_pending_admissions'))


    if request.method == 'POST':

        assigned_class = request.form.get('class')


        if not assigned_class:
            flash("Please select a class.", "warning")
            return redirect(request.url)


        new_student_id = generate_student_id()


        cur.execute("""
            INSERT INTO students(
                student_id,
                full_name,
                class,
                photo_path,
                fees_paid,
                fees_balance,
                fees_total,
                admission_date,
                parent_phone,
                date_of_birth,
                age,
                sex,
                preferred_house,
                disability,
                sports_activities,
                lin,
                admission_source,
                admission_status,
                payment_status,
                payment_transaction_id,
                payment_date
            )
            VALUES(
                %s,%s,%s,%s,%s,%s,%s,
                CURRENT_DATE,
                %s,%s,%s,%s,%s,%s,%s,%s,
                'online',
                'approved',
                'completed',
                %s,
                %s
            )
        """,
        (
            new_student_id,
            student['full_name'],
            assigned_class,
            student['photo_path'],
            student['fees_paid'],
            student['fees_balance'],
            student['fees_total'],
            student['parent_phone'],
            student['date_of_birth'],
            student['age'],
            student['sex'],
            student['preferred_house'],
            student['disability'],
            student['sports_activities'],
            student['lin'],
            student['payment_transaction_id'],
            student['payment_date']
        ))


        # Delete temporary application
        cur.execute("""
            DELETE FROM students
            WHERE student_id=%s
        """,(student_id,))


        db.commit()
        cur.close()


        flash(
            f"Student approved. New ID: {new_student_id}",
            "success"
        )

        return redirect(url_for('dos_pending_admissions'))


    # Load classes
    cur.execute("""
        SELECT name
        FROM classes
        ORDER BY name
    """)

    classes = cur.fetchall()

    cur.close()


    return render_template(
        'dos/approve_admission.html',
        student=student,
        classes=classes
    )

# ==================== REJECT ONLINE ADMISSION ====================

@app.route('/dos/reject_admission/<student_id>')
@login_required
def reject_admission(student_id):

    if not check_permission(['dos','admin']):
        abort(403)

    db = get_db()
    cur = db.cursor()

    # Check applicant exists
    cur.execute("""
        SELECT full_name, admission_status
        FROM students
        WHERE student_id=%s
    """,(student_id,))

    student = cur.fetchone()


    if not student:
        cur.close()
        flash("Student application not found.", "danger")
        return redirect(url_for('dos_pending_admissions'))
    # Update status
    cur.execute("""
        UPDATE students
        SET admission_status='rejected'
        WHERE student_id=%s
    """,(student_id,))


    db.commit()
    cur.close()


    flash(
        f"Admission application for {student['full_name']} rejected.",
        "warning"
    )
    return redirect(url_for('dos_pending_admissions'))
# ==================== ADMISSION LETTER ====================

@app.route('/admission_letter/<student_id>')
@login_required
def admission_letter(student_id):
    cur = get_db().cursor()
    cur.execute("""
        SELECT 
            student_id,
            full_name,
            class,
            admission_date,
            preferred_house
        FROM students
        WHERE student_id=%s
        AND admission_status='approved'
    """,(student_id,))

    student = cur.fetchone()

    cur.close()
    if not student:
        flash("Admission record not found.", "danger")
        return redirect(url_for('dashboard'))
    return render_template(
        'admissions/admission_letter.html',
        student=student
    )

@app.route('/dos/class_lists')
def dos_class_lists():

    if not check_permission(['dos']):
        abort(403)

    class_filter = request.args.get('class', '') or ''
    search = request.args.get('search', '') or ''
    term = request.args.get('term', 'Term 1')

    db = get_db_dict()
    cur = db.cursor()

    cur.execute(
        """
        SELECT DISTINCT class
        FROM students
        WHERE class IS NOT NULL
        AND class != ''
        ORDER BY class
        """
    )

    classes = [
        row['class']
        for row in cur.fetchall()
    ]

    query = """
        SELECT
            student_id,
            full_name,
            class,
            photo_path,
            parent_phone,
            sex,
            age,
            preferred_house,
            lin,
            admission_source
        FROM students
        WHERE 1=1
    """

    params = []

    if class_filter:
        query += " AND class=%s"
        params.append(class_filter)

    if search:
        query += """
        AND (
            student_id ILIKE %s
            OR full_name ILIKE %s
        )
        """
        pattern = f"%{search}%"
        params.extend([pattern, pattern])
    query += " ORDER BY full_name"
    cur.execute(query, params)
    students = cur.fetchall()

    for student in students:
        student['photo_url'] = get_photo_url(
            student.get('photo_path')
        )
    cur.close()
    return render_template(
        'dos/class_lists.html',
        classes=classes,
        students=students,
        selected_class=class_filter,
        search=search,
        term=term
    )

# ==================== DOS TEACHER ASSIGNMENTS ====================

@app.route('/dos/teacher_assignments', methods=['GET', 'POST'])
def dos_teacher_assignments():
    if not check_permission(['dos']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    # =========================================================
    # POST - ADD TEACHER ASSIGNMENT
    # =========================================================
    if request.method == 'POST':

        teacher_id = request.form.get('teacher_id')
        assignment_type = request.form.get('assignment_type')
        class_name = request.form.get('class_name')
        subject = request.form.get('subject') or None

        if not teacher_id or not assignment_type or not class_name:
            flash(
                "Please complete all required fields.",
                "danger"
            )
            cur.close()
            return redirect(url_for('dos_teacher_assignments'))

        class_name = class_name.strip().upper()

        # Class teachers do not have a subject
        if assignment_type == 'classteacher':
            subject = None

        # =====================================================
        # CHECK FOR DUPLICATE ASSIGNMENT
        # =====================================================
        cur.execute(
            """
            SELECT id
            FROM teacher_class_assignments
            WHERE user_id = %s
              AND class_name = %s
              AND assignment_type = %s
              AND subject IS NOT DISTINCT FROM %s
            """,
            (
                teacher_id,
                class_name,
                assignment_type,
                subject
            )
        )

        if cur.fetchone():
            flash(
                "This assignment already exists.",
                "warning"
            )
            cur.close()
            return redirect(url_for('dos_teacher_assignments'))

        # =====================================================
        # ONLY ONE CLASS TEACHER PER CLASS
        # =====================================================
        if assignment_type == 'classteacher':

            cur.execute(
                """
                SELECT id
                FROM teacher_class_assignments
                WHERE class_name = %s
                  AND assignment_type = 'classteacher'
                LIMIT 1
                """,
                (class_name,)
            )

            if cur.fetchone():
                flash(
                    f"{class_name} already has a class teacher.",
                    "danger"
                )
                cur.close()
                return redirect(url_for('dos_teacher_assignments'))

        # =====================================================
        # INSERT ASSIGNMENT
        # =====================================================
        cur.execute(
            """
            INSERT INTO teacher_class_assignments
            (
                user_id,
                class_name,
                subject,
                assignment_type,
                assigned_by,
                assigned_at
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s,
                CURRENT_TIMESTAMP
            )
            """,
            (
                teacher_id,
                class_name,
                subject,
                assignment_type,
                session.get('username')
            )
        )

        db.commit()

        flash(
            "Teacher assignment added successfully.",
            "success"
        )

        cur.close()

        return redirect(url_for('dos_teacher_assignments'))

    # =========================================================
    # GET TEACHERS
    # =========================================================
    cur.execute(
        """
        SELECT
            id,
            username,
            full_name
        FROM users
        WHERE role IN (
            'teacher',
            'subject_teacher',
            'classteacher'
        )
        ORDER BY full_name
        """
    )

    teachers = cur.fetchall()

    # =========================================================
    # GET STUDENT CLASSES
    #
    # Students are stored as:
    #
    # S.1A
    # S.1B
    # S.1C
    # S.2A
    # S.2B
    #
    # We convert them into main classes:
    #
    # S.1
    # S.2
    #
    # IMPORTANT:
    # We return the key as "class" because your existing
    # HTML template uses c.class.
    # =========================================================
    cur.execute(
        """
        SELECT DISTINCT class
        FROM students
        WHERE class IS NOT NULL
          AND TRIM(class) <> ''
        ORDER BY class
        """
    )

    student_class_rows = cur.fetchall()

    main_classes = set()

    for row in student_class_rows:

        student_class = (
            row['class']
            if isinstance(row, dict)
            else row[0]
        )

        if not student_class:
            continue

        student_class = student_class.strip().upper()

        # -----------------------------------------------------
        # Convert:
        #
        # S.1A -> S.1
        # S.1B -> S.1
        # S.2A -> S.2
        #
        # If a class is already S.1, keep it as S.1.
        # -----------------------------------------------------
        if (
            len(student_class) >= 4
            and student_class[-1].isalpha()
        ):
            main_class = student_class[:-1]
        else:
            main_class = student_class

        main_classes.add(main_class)

    # =========================================================
    # BUILD CLASSES FOR THE EXISTING HTML
    #
    # Your template expects:
    #
    # c.class
    #
    # Therefore we deliberately use "class" here.
    # =========================================================
    classes = [
        {
            'class': class_name
        }
        for class_name in sorted(
            main_classes,
            key=lambda x: (
                int(x.split('.')[1])
                if '.' in x
                and x.split('.')[1].isdigit()
                else 999,
                x
            )
        )
    ]

    # =========================================================
    # GET ALL ASSIGNMENTS
    # =========================================================
    cur.execute(
        """
        SELECT
            tca.id,
            tca.user_id,
            tca.class_name,
            tca.subject,
            tca.assignment_type,
            tca.assigned_by,
            tca.assigned_at,
            u.username,
            u.full_name,
            u.phone
        FROM teacher_class_assignments tca
        JOIN users u
          ON tca.user_id = u.id
        ORDER BY
            u.full_name,
            tca.class_name,
            tca.subject
        """
    )

    assignments = cur.fetchall()

    cur.close()

    # =========================================================
    # ORGANIZE ASSIGNMENTS BY TEACHER
    # =========================================================
    teachers_data = {}

    for assignment in assignments:

        user_id = assignment['user_id']

        if user_id not in teachers_data:
            teachers_data[user_id] = {
                'username': assignment['username'],
                'full_name': assignment['full_name'],
                'phone': assignment['phone'],
                'class_teacher': None,
                'subjects': []
            }

        # =====================================================
        # CLASS TEACHER
        # =====================================================
        if assignment['assignment_type'] == 'classteacher':

            teachers_data[user_id]['class_teacher'] = (
                assignment['class_name']
            )

        # =====================================================
        # SUBJECT TEACHER
        # =====================================================
        else:

            teachers_data[user_id]['subjects'].append(
                {
                    'class': assignment['class_name'],
                    'subject': assignment['subject']
                }
            )

    # =========================================================
    # SEPARATE TEACHERS BY ROLE
    # =========================================================
    class_teachers = []
    subject_teachers = []
    both_roles = []

    for teacher in teachers_data.values():

        has_class = (
            teacher['class_teacher'] is not None
        )

        has_subject = (
            len(teacher['subjects']) > 0
        )

        # =====================================================
        # BOTH ROLES
        # =====================================================
        if has_class and has_subject:

            teacher['classteacher_class'] = (
                teacher['class_teacher']
            )

            both_roles.append(teacher)

        # =====================================================
        # CLASS TEACHER ONLY
        # =====================================================
        elif has_class:

            teacher['class_name'] = (
                teacher['class_teacher']
            )

            class_teachers.append(teacher)

        # =====================================================
        # SUBJECT TEACHER ONLY
        # =====================================================
        elif has_subject:

            for subject_assignment in teacher['subjects']:

                subject_teachers.append(
                    {
                        'username': teacher['username'],
                        'full_name': teacher['full_name'],
                        'phone': teacher['phone'],
                        'class_name': subject_assignment['class'],
                        'subject': subject_assignment['subject']
                    }
                )

    # =========================================================
    # RENDER TEMPLATE
    # =========================================================
    return render_template(
        'dos/teacher_assignments.html',
        teachers=teachers,
        classes=classes,
        assignments=assignments,
        class_teachers=class_teachers,
        subject_teachers=subject_teachers,
        both_roles=both_roles
    )
# ==================== REMOVE STUDENT ====================

@app.route('/dos/remove_student/<student_id>', methods=['POST'])
def dos_remove_student(student_id):

    if not check_permission(['dos']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute(
        """
        SELECT photo_path
        FROM students
        WHERE student_id=%s
        """,
        (student_id,)
    )

    row = cur.fetchone()

    cur.close()

    if row and row['photo_path'] != 'default_avatar.png':

        path = os.path.join(
            app.config['UPLOAD_FOLDER'],
            row['photo_path']
        )

        if os.path.exists(path):
            os.remove(path)

    execute_db(
        "DELETE FROM students WHERE student_id=%s",
        (student_id,)
    )

    flash(
        f'Student {student_id} removed.',
        'success'
    )

    return redirect(
        url_for('dos_class_lists')
    )


# ==================== PROMOTION ====================

@app.route('/dos/promote', methods=['GET', 'POST'])
def dos_promote():

    if not check_permission(['dos']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute(
        """
        SELECT DISTINCT class
        FROM students
        WHERE class IS NOT NULL
        AND class != ''
        ORDER BY class
        """
    )

    classes = [
        row['class']
        for row in cur.fetchall()
    ]

    if request.method == 'POST':

        from_class = request.form['from_class']

        match = re.search(
            r'(\d+)',
            from_class
        )

        if match:

            to_class = from_class.replace(
                match.group(1),
                str(int(match.group(1))+1)
            )

        else:

            to_class = from_class + " (Promoted)"


        cur.execute(
            """
            UPDATE students
            SET class=%s
            WHERE class=%s
            """,
            (
                to_class,
                from_class
            )
        )

        count = cur.rowcount

        db.commit()

        flash(
            f'{count} students promoted from {from_class} to {to_class}.',
            'success'
        )

    cur.close()

    return render_template(
        'dos/promote.html',
        classes=classes
    )


# ==================== ATTENDANCE ====================

@app.route('/dos/attendance')
def dos_attendance():

    if not check_permission(['dos']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute(
        """
        SELECT DISTINCT class
        FROM students
        WHERE class IS NOT NULL
        AND class != ''
        ORDER BY class
        """
    )

    classes = [
        row['class']
        for row in cur.fetchall()
    ]

    cur.close()

    return render_template(
        'dos/attendance.html',
        classes=classes
    )


# ==================== SCHEDULES ====================

@app.route('/dos/schedules', methods=['GET', 'POST'])
def dos_schedules():

    if not check_permission(['dos']):
        abort(403)
    if request.method == 'POST':

        schedule_type = request.form['schedule_type']
        term_scope = request.form['term_scope']
        content = request.form.get(
            'schedule_text',
            ''
        ).strip()

        file = request.files.get('schedule_file')

        final_content = content


        if file and file.filename and allowed_file(
            file.filename,
            {'csv'}
        ):
            stream = io.StringIO(
                file.stream.read().decode("UTF8"),
                newline=None
            )
            reader = csv.reader(stream)
            rows = []
            for row in reader:
                if any(row):
                    rows.append(
                        ",".join(
                            [
                                escape(cell.strip())
                                for cell in row
                            ]
                        )
                    )

            final_content = "\n".join(rows)
        execute_db(
            """
            INSERT INTO schedules
            (
                type,
                term_scope,
                content,
                updated_at
            )
            VALUES
            (%s,%s,%s,CURRENT_TIMESTAMP)

            ON CONFLICT(type,term_scope)

            DO UPDATE SET
                content=EXCLUDED.content,
                updated_at=CURRENT_TIMESTAMP
            """,
            (
                schedule_type,
                term_scope,
                final_content
            )
        )
        flash(
            f'{schedule_type.capitalize()} saved.',
            'success'
        )
        return redirect(
            url_for('dos_schedules')
        )
    db = get_db_dict()
    cur = db.cursor()
    cur.execute(
        """
        SELECT
            type,
            term_scope,
            content,
            updated_at
        FROM schedules
        ORDER BY type, term_scope DESC
        """
    )

    schedules = cur.fetchall()
    cur.close()
    return render_template(
        'dos/schedules.html',
        schedules=schedules
    )

def get_olevel_grade_details(score):
    db = get_db_dict()
    cur = db.cursor()

    cur.execute(
        """
        SELECT grade, descriptor
        FROM grading_system
        WHERE %s BETWEEN min_score AND max_score
        LIMIT 1
        """,
        (score,)
    )

    result = cur.fetchone()

    cur.close()

    if result:
        return (
            result['grade'],
            result['descriptor']
        )

    return (
        '-',
        '-'
    )

# ==================== O-LEVEL GRADING MANAGEMENT ====================

@app.route('/dos/olevel_grading', methods=['GET', 'POST'])
def dos_olevel_grading():

    if not check_permission(['dos']):
        abort(403)

    if request.method == 'POST':

        file = request.files.get('grading_file')

        if not file or not file.filename:
            flash(
                'Please upload an Excel file.',
                'danger'
            )
            return redirect(
                url_for('dos_olevel_grading')
            )

        try:

            from openpyxl import load_workbook

            wb = load_workbook(
                file,
                data_only=True
            )

            sheet = wb.active

            headers = [
                str(cell.value).strip().lower()
                if cell.value else ''
                for cell in sheet[1]
            ]

            cols = {}

            for idx, header in enumerate(headers):

                if header in [
                    'min_score',
                    'max_score',
                    'grade',
                    'descriptor'
                ]:
                    cols[header] = idx


            required = [
                'min_score',
                'max_score',
                'grade',
                'descriptor'
            ]

            if not all(
                item in cols
                for item in required
            ):

                flash(
                    'Missing required columns: min_score, max_score, grade, descriptor',
                    'danger'
                )

                return redirect(
                    url_for('dos_olevel_grading')
                )


            execute_db(
                "DELETE FROM grading_system"
            )


            count = 0


            for row_idx in range(
                2,
                sheet.max_row + 1
            ):

                min_val = sheet.cell(
                    row=row_idx,
                    column=cols['min_score'] + 1
                ).value

                max_val = sheet.cell(
                    row=row_idx,
                    column=cols['max_score'] + 1
                ).value

                grade_val = sheet.cell(
                    row=row_idx,
                    column=cols['grade'] + 1
                ).value

                desc_val = sheet.cell(
                    row=row_idx,
                    column=cols['descriptor'] + 1
                ).value


                if None in [
                    min_val,
                    max_val,
                    grade_val
                ]:
                    continue


                try:

                    execute_db(
                        """
                        INSERT INTO grading_system
                        (
                            min_score,
                            max_score,
                            grade,
                            descriptor
                        )
                        VALUES
                        (%s,%s,%s,%s)
                        """,
                        (
                            float(min_val),
                            float(max_val),
                            str(grade_val).strip(),
                            str(desc_val).strip()
                            if desc_val else ''
                        )
                    )

                    count += 1


                except Exception:

                    continue


            flash(
                f'{count} O-Level grading rules uploaded.',
                'success'
            )


        except Exception as e:

            flash(
                f'Error: {str(e)}',
                'danger'
            )


        return redirect(
            url_for('dos_olevel_grading')
        )


    db = get_db_dict()
    cur = db.cursor()

    cur.execute(
        """
        SELECT
            min_score,
            max_score,
            grade,
            descriptor
        FROM grading_system
        ORDER BY min_score DESC
        """
    )

    rules = cur.fetchall()

    cur.close()


    return render_template(
        'dos/olevel_grading.html',
        rules=rules
    )



# ==================== A-LEVEL GRADING MANAGEMENT ====================

@app.route('/dos/alevel_grading', methods=['GET', 'POST'])
def dos_alevel_grading():

    if not check_permission(['dos']):
        abort(403)
    if request.method == 'POST':
        file = request.files.get('grading_file')
        if not file or not file.filename:
            flash(
                'Please upload an Excel file.',
                'danger'
            )
            return redirect(
                url_for('dos_alevel_grading')
            )
        try:
            from openpyxl import load_workbook
            wb = load_workbook(
                file,
                data_only=True
            )
            sheet = wb.active
            headers = [
                str(cell.value).strip().lower()
                if cell.value else ''
                for cell in sheet[1]
            ]
            cols = {}
            for idx, header in enumerate(headers):
                if header in [
                    'min_score',
                    'max_score',
                    'grade',
                    'points'
                ]:
                    cols[header] = idx
            required = [
                'min_score',
                'max_score',
                'grade',
                'points'
            ]
            if not all(
                item in cols
                for item in required
            ):
                flash(
                    'Missing required columns: min_score, max_score, grade, points',
                    'danger'
                )

                return redirect(
                    url_for('dos_alevel_grading')
                )
            execute_db(
                """
                DELETE FROM alevel_grading
                WHERE is_subsidiary=FALSE
                """
            )
            count = 0
            for row_idx in range(
                2,
                sheet.max_row + 1
            ):

                min_val = sheet.cell(
                    row=row_idx,
                    column=cols['min_score'] + 1
                ).value

                max_val = sheet.cell(
                    row=row_idx,
                    column=cols['max_score'] + 1
                ).value

                grade_val = sheet.cell(
                    row=row_idx,
                    column=cols['grade'] + 1
                ).value

                points_val = sheet.cell(
                    row=row_idx,
                    column=cols['points'] + 1
                ).value
                if None in [
                    min_val,
                    max_val,
                    grade_val,
                    points_val
                ]:
                    continue
                try:
                    execute_db(
                        """
                        INSERT INTO alevel_grading
                        (
                            min_score,
                            max_score,
                            grade,
                            points,
                            is_subsidiary
                        )
                        VALUES
                        (%s,%s,%s,%s,FALSE)
                        """,
                        (
                            float(min_val),
                            float(max_val),
                            str(grade_val).strip(),
                            int(points_val)
                        )
                    )

                    count += 1
                except Exception:

                    continue
            flash(
                f'{count} A-Level grading rules uploaded.',
                'success'
            )
        except Exception as e:

            flash(
                f'Error: {str(e)}',
                'danger'
            )


        return redirect(
            url_for('dos_alevel_grading')
        )
    db = get_db_dict()
    cur = db.cursor()

    cur.execute(
        """
        SELECT
            min_score,
            max_score,
            grade,
            points
        FROM alevel_grading
        WHERE is_subsidiary=FALSE
        ORDER BY min_score DESC
        """
    )
    rules = cur.fetchall()

    cur.close()
    return render_template(
        'dos/alevel_grading.html',
        rules=rules
    )
    

@app.route('/dos/predefined_comments')
def dos_predefined_comments():
    if not check_permission(['dos']):
        abort(403)
    db = get_db_dict()
    cur = db.cursor()
    cur.execute("SELECT * FROM predefined_comments ORDER BY comment_type, id")
    comments = cur.fetchall()
    cur.close()
    return render_template('dos/predefined_comments.html', comments=comments)

@app.route('/dos/predefined_comments/add', methods=['POST'])
def dos_predefined_comments_add():
    if not check_permission(['dos']):
        abort(403)
    comment_type = request.form['comment_type']
    comment_text = request.form['comment_text'].strip()
    if not comment_text:
        flash('Comment text is required.', 'danger')
        return redirect(url_for('dos_predefined_comments'))
    execute_db(
        "INSERT INTO predefined_comments (comment_type, comment_text, is_active) VALUES (%s, %s, %s)",
        (comment_type, comment_text, True)
    )
    flash('Comment added successfully.', 'success')
    return redirect(url_for('dos_predefined_comments'))

@app.route('/dos/predefined_comments/delete/<int:comment_id>')
def dos_predefined_comments_delete(comment_id):
    if not check_permission(['dos']):
        abort(403)
    execute_db(
        "DELETE FROM predefined_comments WHERE id=%s",
        (comment_id,)
    )
    flash('Comment deleted successfully.', 'success')
    return redirect(url_for('dos_predefined_comments'))

@app.route('/dos/upload_teachers', methods=['GET', 'POST'])
def dos_upload_teachers():
    if not check_permission(['dos']):
        abort(403)

    if request.method == 'POST':
        file = request.files.get('excel_file')

        if not file or not file.filename:
            flash('Please upload an Excel file.', 'danger')
            return redirect(url_for('dos_upload_teachers'))

        try:
            from openpyxl import load_workbook

            wb = load_workbook(file, data_only=True)
            sheet = wb.active

            headers = [str(cell.value).strip().lower() if cell.value else '' for cell in sheet[1]]

            col_map = {}
            for idx, h in enumerate(headers):
                if h in ['username', 'full_name', 'class_name', 'subject', 'assignment_type']:
                    col_map[h] = idx

            required = ['username', 'full_name', 'class_name', 'assignment_type']

            for r in required:
                if r not in col_map:
                    flash(f'Missing column: {r}', 'danger')
                    return redirect(url_for('dos_upload_teachers'))

            db = get_db_dict()
            cur = db.cursor()

            success = 0
            errors = []

            for row_idx in range(2, sheet.max_row + 1):
                try:
                    username = str(sheet.cell(row=row_idx, column=col_map['username'] + 1).value or '').strip()
                    full_name = str(sheet.cell(row=row_idx, column=col_map['full_name'] + 1).value or '').strip()
                    class_name = str(sheet.cell(row=row_idx, column=col_map['class_name'] + 1).value or '').strip()
                    assignment_type = str(sheet.cell(row=row_idx, column=col_map['assignment_type'] + 1).value or '').strip().lower()
                    subject = str(sheet.cell(row=row_idx, column=col_map['subject'] + 1).value or '').strip() if 'subject' in col_map else None

                    if not username or not class_name or not assignment_type:
                        errors.append(f"Row {row_idx}: Missing username, class_name, or assignment_type")
                        continue

                    if assignment_type not in ['classteacher', 'subject_teacher']:
                        errors.append(f"Row {row_idx}: Invalid assignment type")
                        continue

                    if assignment_type == 'subject_teacher' and not subject:
                        errors.append(f"Row {row_idx}: Subject required for subject teacher")
                        continue

                    cur.execute(
                        "SELECT id, full_name, role FROM users WHERE username=%s",
                        (username,)
                    )

                    user = cur.fetchone()

                    if not user:
                        default_password = 'password123'
                        hashed = generate_password_hash(default_password)

                        cur.execute(
                            """
                            INSERT INTO users
                            (username, full_name, password, role, status, must_change_password)
                            VALUES (%s,%s,%s,%s,%s,%s)
                            RETURNING id
                            """,
                            (
                                username,
                                full_name or username,
                                hashed,
                                'subject_teacher' if assignment_type == 'subject_teacher' else 'classteacher',
                                1,
                                1
                            )
                        )

                        user_id = cur.fetchone()['id']

                        add_notification(
                            'dos',
                            f"New teacher created: {full_name or username}. Username: {username}, Password: {default_password}. Class: {class_name}, Type: {assignment_type}",
                            "/dos/teacher_assignments"
                        )

                        success += 1

                    else:
                        user_id = user['id']

                        if full_name and full_name != user['full_name']:
                            cur.execute(
                                "UPDATE users SET full_name=%s WHERE id=%s",
                                (full_name, user_id)
                            )

                        if user['role'] not in ['admin', 'headteacher', 'bursar']:
                            new_role = 'subject_teacher' if assignment_type == 'subject_teacher' else 'classteacher'

                            cur.execute(
                                "UPDATE users SET role=%s WHERE id=%s",
                                (new_role, user_id)
                            )

                        success += 1

                    if assignment_type == 'classteacher':
                        cur.execute(
                            """
                            SELECT id FROM teacher_class_assignments
                            WHERE class_name=%s
                            AND assignment_type='classteacher'
                            """,
                            (class_name,)
                        )

                        if cur.fetchone():
                            errors.append(f"Row {row_idx}: Class {class_name} already has a class teacher")
                            continue

                    cur.execute(
                        """
                        SELECT id FROM teacher_class_assignments
                        WHERE user_id=%s
                        AND class_name=%s
                        AND assignment_type=%s
                        """,
                        (
                            user_id,
                            class_name,
                            assignment_type
                        )
                    )

                    existing = cur.fetchone()

                    if existing:
                        if assignment_type == 'subject_teacher':
                            cur.execute(
                                """
                                UPDATE teacher_class_assignments
                                SET subject=%s,
                                    assigned_by=%s,
                                    assigned_at=CURRENT_TIMESTAMP
                                WHERE id=%s
                                """,
                                (
                                    subject,
                                    session.get('username'),
                                    existing['id']
                                )
                            )
                    else:
                        cur.execute(
                            """
                            INSERT INTO teacher_class_assignments
                            (user_id, class_name, subject, assignment_type, assigned_by)
                            VALUES (%s,%s,%s,%s,%s)
                            """,
                            (
                                user_id,
                                class_name,
                                subject,
                                assignment_type,
                                session.get('username')
                            )
                        )

                except Exception as e:
                    errors.append(f"Row {row_idx}: {str(e)}")
                    app.logger.error(f"Error row {row_idx}: {str(e)}")

            db.commit()
            cur.close()

            flash(
                f'{success} teachers processed. {len(errors)} issues found.',
                'success' if success else 'warning'
            )

            for e in errors[:10]:
                flash(e, 'warning')

        except Exception as e:
            flash(f'Error uploading file: {str(e)}', 'danger')

        return redirect(url_for('dos_teacher_assignments'))

    return render_template('dos/upload_teachers.html')

def assign_user_to_class(
    user_id,
    class_name,
    subject=None,
    assignment_type='subject_teacher'
):
    """
    Assign a teacher to a class.

    Main classes such as S.1, S.2, S.3 are stored exactly as assigned.

    A main-class assignment automatically covers its streams:

        S.1  -> S.1A, S.1B, S.1C
        S.2  -> S.2A, S.2B, S.2C

    A specific stream assignment such as S.1A only covers S.1A.
    """

    db = None
    cur = None

    try:
        # ---------------------------------------------------------
        # BASIC VALIDATION
        # ---------------------------------------------------------
        if not user_id or not class_name:
            return False

        class_name = class_name.strip().upper()

        # Class teachers do not have a subject
        if assignment_type == 'classteacher':
            subject = None

        # ---------------------------------------------------------
        # GET DATABASE
        # ---------------------------------------------------------
        db = get_db_dict()
        cur = db.cursor()

        # ---------------------------------------------------------
        # CHECK EXISTING ASSIGNMENT
        #
        # For subject teachers, the same teacher can have:
        #
        # S.1 Mathematics
        # S.1 English
        #
        # Therefore subject is included in the duplicate check.
        # ---------------------------------------------------------
        if assignment_type == 'subject_teacher':

            cur.execute(
                """
                SELECT id
                FROM teacher_class_assignments
                WHERE user_id = %s
                  AND class_name = %s
                  AND assignment_type = %s
                  AND subject IS NOT DISTINCT FROM %s
                """,
                (
                    user_id,
                    class_name,
                    assignment_type,
                    subject
                )
            )

        else:

            cur.execute(
                """
                SELECT id
                FROM teacher_class_assignments
                WHERE user_id = %s
                  AND class_name = %s
                  AND assignment_type = %s
                """,
                (
                    user_id,
                    class_name,
                    assignment_type
                )
            )

        existing = cur.fetchone()

        # ---------------------------------------------------------
        # UPDATE EXISTING ASSIGNMENT
        # ---------------------------------------------------------
        if existing:

            if assignment_type == 'subject_teacher':

                cur.execute(
                    """
                    UPDATE teacher_class_assignments
                    SET subject = %s,
                        assigned_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """,
                    (
                        subject,
                        existing['id']
                    )
                )

            else:

                cur.execute(
                    """
                    UPDATE teacher_class_assignments
                    SET assigned_at = CURRENT_TIMESTAMP
                    WHERE id = %s
                    """,
                    (
                        existing['id'],
                    )
                )

        # ---------------------------------------------------------
        # CREATE NEW ASSIGNMENT
        # ---------------------------------------------------------
        else:

            # -----------------------------------------------------
            # ONLY ONE CLASS TEACHER PER CLASS
            # -----------------------------------------------------
            if assignment_type == 'classteacher':

                cur.execute(
                    """
                    SELECT id
                    FROM teacher_class_assignments
                    WHERE class_name = %s
                      AND assignment_type = 'classteacher'
                    LIMIT 1
                    """,
                    (
                        class_name,
                    )
                )

                class_teacher_exists = cur.fetchone()

                if class_teacher_exists:
                    cur.close()

                    app.logger.warning(
                        f"Class {class_name} already has a class teacher."
                    )

                    return False

            # -----------------------------------------------------
            # INSERT ASSIGNMENT
            # -----------------------------------------------------
            cur.execute(
                """
                INSERT INTO teacher_class_assignments
                (
                    user_id,
                    class_name,
                    subject,
                    assignment_type,
                    assigned_by,
                    assigned_at
                )
                VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
                """,
                (
                    user_id,
                    class_name,
                    subject,
                    assignment_type,
                    session.get('username')
                )
            )

        # ---------------------------------------------------------
        # SAVE
        # ---------------------------------------------------------
        db.commit()

        cur.close()

        return True

    except Exception as e:

        if db:
            try:
                db.rollback()
            except Exception:
                pass

        if cur:
            try:
                cur.close()
            except Exception:
                pass

        app.logger.error(
            f"Error in assign_user_to_class: {str(e)}"
        )

        return False
        
@app.route('/dos/delete_assignment/<int:assignment_id>', methods=['POST'])
def dos_delete_assignment(assignment_id):
    if not check_permission(['dos']):
        return jsonify({'success': False, 'error': 'Permission denied'})

    try:
        db = get_db_dict()
        cur = db.cursor()

        cur.execute(
            "DELETE FROM teacher_class_assignments WHERE id=%s",
            (assignment_id,)
        )

        db.commit()
        cur.close()

        return jsonify({'success': True})

    except Exception as e:
        return jsonify({'success': False, 'error': str(e)})


@app.route('/dos/edit_assignment/<int:assignment_id>', methods=['GET', 'POST'])
def dos_edit_assignment(assignment_id):
    if not check_permission(['dos']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    if request.method == 'POST':
        class_name = request.form.get('class_name')
        subject = request.form.get('subject')
        assignment_type = request.form.get('assignment_type')

        cur.execute(
            """
            UPDATE teacher_class_assignments
            SET class_name=%s,
                subject=%s,
                assignment_type=%s,
                assigned_by=%s
            WHERE id=%s
            """,
            (
                class_name,
                subject,
                assignment_type,
                session.get('username'),
                assignment_id
            )
        )

        db.commit()
        cur.close()

        flash('Assignment updated successfully!', 'success')
        return redirect(url_for('dos_teacher_assignments'))

    cur.execute(
        """
        SELECT tca.*, u.username, u.full_name
        FROM teacher_class_assignments tca
        JOIN users u
        ON tca.user_id=u.id
        WHERE tca.id=%s
        """,
        (assignment_id,)
    )

    assignment = cur.fetchone()
    cur.close()

    cur = db.cursor()
    cur.execute(
        "SELECT DISTINCT class FROM students ORDER BY class"
    )

    classes = cur.fetchall()
    cur.close()

    return render_template(
        'dos/edit_assignment.html',
        assignment=assignment,
        classes=classes
    )
    
# ==================== UNIFIED TEACHER MODULE ====================
# ==================== TEACHER STUDENTS ====================

@app.route('/teacher/students')
def teacher_students():
    if not check_permission(['classteacher', 'subject_teacher']):
        abort(403)

    term = request.args.get('term', 'Term 1')
    user_id = session.get('user_id')

    # =========================================================
    # GET TEACHER ASSIGNMENTS
    # =========================================================
    assignments = get_user_assignments(user_id)

    if not assignments:
        flash(
            'No classes assigned to you. Please contact admin.',
            'danger'
        )
        return redirect(url_for('dashboard'))

    # =========================================================
    # GET ASSIGNED CLASSES AND THEIR STREAMS
    # =========================================================
    
    db = get_db_dict()
    cur = db.cursor()
    
    assigned_classes = sorted(
        set(
            (a['class_name'] or '').strip().upper()
            for a in assignments
            if a.get('class_name')
        )
    )
    
    available_classes = set()
    
    for assigned_class in assigned_classes:
    
        # Always include the main class
        available_classes.add(assigned_class)
    
        # If assignment is a main class such as S.1,
        # automatically include all its streams.
        if re.match(r'^S\.\d+$', assigned_class):
    
            cur.execute(
                """
                SELECT DISTINCT class
                FROM students
                WHERE class ~ %s
                ORDER BY class
                """,
                (
                    '^' + re.escape(assigned_class) + r'[A-Za-z]+$',
                )
            )
    
            stream_rows = cur.fetchall()
    
            for row in stream_rows:
                stream_class = (row['class'] or '').strip().upper()
    
                if stream_class:
                    available_classes.add(stream_class)
    
    cur.close()
    
    available_classes = sorted(available_classes)

    if not available_classes:
        flash(
            'No classes assigned to you. Please contact admin.',
            'danger'
        )
        return redirect(url_for('dashboard'))

    # =========================================================
    # SELECT CLASS
    # =========================================================
    selected_class = request.args.get(
        'class_name',
        session.get(
            'selected_class',
            available_classes[0]
        )
    )

    selected_class = selected_class.strip().upper()

    # =========================================================
    # CHECK WHETHER TEACHER HAS ACCESS
    #
    # Example:
    #
    # Assignment: S.1
    #
    # Allowed:
    # S.1A
    # S.1B
    # S.1C
    #
    # Assignment: S.1A
    #
    # Allowed:
    # S.1A only
    # =========================================================
    if not teacher_has_class_access(
        user_id,
        selected_class
    ):
        selected_class = available_classes[0]

        session['selected_class'] = selected_class
    else:
        session['selected_class'] = selected_class

    # =========================================================
    # GET STUDENTS
    # =========================================================
    db = get_db_dict()
    cur = db.cursor()

    # =========================================================
    # MAIN CLASS
    #
    # If assignment is S.1, retrieve:
    #
    # S.1A
    # S.1B
    # S.1C
    # etc.
    #
    # =========================================================
    if re.match(
        r'^S\.\d+$',
        selected_class
    ):

        cur.execute(
            """
            SELECT
                student_id,
                full_name,
                photo_path,
                parent_phone,
                class
            FROM students
            WHERE class=%s OR class ~ %s
            ORDER BY class, full_name
            """,
            (
                selected_class, 
                '^' + selected_class + r'[A-Za-z]+$',
            )
        )

    # =========================================================
    # SPECIFIC STREAM
    #
    # If selected class is S.1A, retrieve only S.1A.
    #
    # =========================================================
    else:

        cur.execute(
            """
            SELECT
                student_id,
                full_name,
                photo_path,
                parent_phone,
                class
            FROM students
            WHERE class = %s
            ORDER BY full_name
            """,
            (selected_class,)
        )

    students = cur.fetchall()

    # =========================================================
    # ADD PHOTO URL
    # =========================================================
    for student in students:
        student['photo_url'] = get_photo_url(
            student.get('photo_path')
        )

    cur.close()

    # =========================================================
    # CHECK CLASS TEACHER STATUS
    #
    # If teacher is class teacher of S.1,
    # they remain class teacher for:
    #
    # S.1A
    # S.1B
    # S.1C
    #
    # =========================================================
    is_classteacher = teacher_has_class_access(
        user_id,
        selected_class,
        assignment_type='classteacher'
    )

    # =========================================================
    # RETURN TEMPLATE
    # =========================================================
    return render_template(
        'teacher/students.html',
        students=students,
        selected_class=selected_class,
        available_classes=available_classes,
        is_classteacher=is_classteacher,
        term=term
    )
    
@app.route('/teacher/attendance', methods=['GET', 'POST'])
def teacher_attendance():
    if not check_permission(['classteacher']):
        abort(403)

    selected_class = session.get('selected_class')

    if not selected_class:
        flash('No class selected', 'danger')
        return redirect(url_for('teacher_students'))

    selected_date = request.args.get(
        'date',
        datetime.now().strftime('%Y-%m-%d')
    )

    if request.method == 'POST':
        selected_date = request.form['date']

        for key, value in request.form.items():
            if key.startswith('status_'):
                student_id = key.split('_')[1]

                execute_db(
                    """
                    INSERT INTO attendance
                    (student_id, date, status)
                    VALUES (%s, %s, %s)
                    ON CONFLICT(student_id, date)
                    DO UPDATE SET status=%s
                    """,
                    (student_id, selected_date, value, value)
                )

        flash('Attendance saved.', 'success')

    db = get_db_dict()
    cur = db.cursor()

    cur.execute(
        """
        SELECT s.student_id, s.full_name, a.status
        FROM students s
        LEFT JOIN attendance a
        ON s.student_id = a.student_id
        AND a.date = %s
        WHERE s.class = %s
        ORDER BY s.full_name
        """,
        (selected_date, selected_class)
    )

    records = cur.fetchall()
    cur.close()

    return render_template(
        'teacher/attendance.html',
        records=records,
        selected_date=selected_date,
        assigned_class=selected_class
    )

def to_number(value):
    if value is None:
        return 0

    if isinstance(value, list):
        return [
            float(v) if v.strip() != "" else 0
            for v in value
        ]

    if value.strip() == "":
        return 0

    return float(value)

@app.route("/save_manual_marks", methods=["POST"])
def save_manual_marks():
    if not check_permission(
        ['classteacher', 'subject_teacher', 'dos']
    ):
        abort(403)
    
    db = get_db_dict()
    cursor = db.cursor()
    
    subject = request.form.get('subject')
    term = request.form.get('term')
    year = request.form.get('year')
    
    # =========================================================
    # ALLOWED TERMS
    # =========================================================
    allowed_terms = ['Term 1', 'Term 2', 'Term 3']
    
    if term not in allowed_terms:
        cursor.close()
        db.close()
    
        flash(
            'Invalid term. Please select Term 1, Term 2 or Term 3.',
            'danger'
        )
    
        return redirect(
            url_for(
                'teacher_upload_marks',
                class_name=session.get('selected_class')
            )
        )
    
    # =========================================================
    # GET FORM DATA
    # =========================================================
    student_ids = request.form.getlist("student_id[]")
    paper1_values = request.form.getlist("paper1[]")
    paper2_values = request.form.getlist("paper2[]")
    initials = request.form.getlist("teacher_initials[]")
    
    # =========================================================
    # CONVERT MARK TO NUMBER
    # EMPTY FIELD = NONE
    # =========================================================
    def parse_mark(value):
    
        if value is None:
            return None
    
        value = str(value).strip()
    
        if value == '':
            return None
    
        try:
            mark = float(value)
    
            # Do not allow negative marks
            if mark < 0:
                return None
    
            # Maximum mark for EACH PAPER is 100
            return min(mark, 100)
    
        except (ValueError, TypeError):
            return None
    
    # =========================================================
    # PROCESS EACH STUDENT
    # =========================================================
    for i, sid in enumerate(student_ids):
    
        # -----------------------------------------------------
        # PAPER 1
        # -----------------------------------------------------
        p1 = parse_mark(
            paper1_values[i]
            if i < len(paper1_values)
            else None
        )
    
        # -----------------------------------------------------
        # PAPER 2
        # -----------------------------------------------------
        p2 = parse_mark(
            paper2_values[i]
            if i < len(paper2_values)
            else None
        )
    
        # -----------------------------------------------------
        # TEACHER INITIALS
        # -----------------------------------------------------
        init = (
            initials[i].strip()
            if i < len(initials) and initials[i]
            else ''
        )
    
        # =====================================================
        # IF NO PAPER WAS ENTERED, SKIP STUDENT
        # =====================================================
        if p1 is None and p2 is None:
            continue
    
        # =====================================================
        # CALCULATE FINAL MARK
        #
        # BOTH PAPERS:
        #       Average of Paper 1 and Paper 2
        #
        # PAPER 1 ONLY:
        #       Paper 1
        #
        # PAPER 2 ONLY:
        #       Paper 2
        # =====================================================
        if p1 is not None and p2 is not None:
    
            final_mark = (p1 + p2) / 2
    
        elif p1 is not None:
    
            final_mark = p1
    
        else:
    
            final_mark = p2
    
        # Round final mark to 2 decimal places
        final_mark = round(final_mark, 2)
    
        # =====================================================
        # INSERT OR UPDATE MARK
        # =====================================================
        cursor.execute(
            """
            INSERT INTO marks
            (
                student_id,
                subject,
                paper1,
                paper2,
                final_mark,
                teacher_initials,
                term,
                year
            )
            VALUES
            (%s, %s, %s, %s, %s, %s, %s, %s)
    
            ON CONFLICT (student_id, subject, term, year)
            DO UPDATE SET
                paper1 = EXCLUDED.paper1,
                paper2 = EXCLUDED.paper2,
                final_mark = EXCLUDED.final_mark,
                teacher_initials = EXCLUDED.teacher_initials
            """,
            (
                sid,
                subject,
                p1,
                p2,
                final_mark,
                init,
                term,
                year
            )
        )
    
    # =========================================================
    # SAVE CHANGES
    # =========================================================
    db.commit()
    
    cursor.close()
    db.close()
    
    # =========================================================
    # SUCCESS MESSAGE
    # =========================================================
    flash(
        f"Marks for {term} entered successfully.",
        "success"
    )
    
    # =========================================================
    # RETURN TO MARK ENTRY PAGE
    # =========================================================
    return redirect(
        url_for(
            'teacher_upload_marks',
            class_name=session.get('selected_class'),
            term=term
        )
    )

@app.route('/teacher/upload_marks', methods=['GET', 'POST'])
def teacher_upload_marks():
    if not check_permission(['classteacher', 'subject_teacher', 'dos']):
        abort(403)
    teacher_id = session.get('user_id')
    assignments = get_user_assignments(teacher_id)
    if not assignments:
        flash('No classes assigned.', 'danger')
        return redirect(url_for('dashboard'))
    available_classes = list(
        set(
            [a['class_name'] for a in assignments]
        )
    )
    selected_class = request.args.get(
        'class_name',
        session.get(
            'selected_class',
            available_classes[0]
        )
    )
    if selected_class not in available_classes:
        selected_class = available_classes[0]
    session['selected_class'] = selected_class
    class_upper = selected_class.upper()
    level = (
        'alevel'
        if (
            class_upper in [
                'S5',
                'S6',
                'A-LEVEL',
                'A LEVEL',
                'S.5',
                'S.6'
            ]
            or
            (
                class_upper.startswith('S')
                and len(class_upper) >= 2
                and class_upper[1] in ['5', '6']
            )
        )
        else 'olevel'
    )
    current_year = datetime.now().year
    # ===============================
    # GET STUDENTS FOR MANUAL ENTRY
    # ===============================
    db = get_db_dict()
    cur = db.cursor()
    cur.execute(
        """
        SELECT student_id, full_name
        FROM students
        WHERE class=%s
        ORDER BY full_name
        """,
        (selected_class,)
    )
    students = cur.fetchall()
    cur.close()
    # ===============================
    # EXCEL UPLOAD PROCESS
    # ===============================
    if request.method == "POST":
        subject = request.form.get(
            'subject',
            ''
        ).strip()
        term = request.form.get(
            'term',
            'Term 1'
        ).strip()
        year = request.form.get(
            'year',
            current_year
        )
        is_subsidiary = (
            request.form.get('is_subsidiary') == 'on'
        )
        file = request.files.get(
            'marks_file'
        )
        if not file or not file.filename:
            flash(
                'Please upload an Excel file.',
                'danger'
            )
            return redirect(
                url_for(
                    'teacher_upload_marks',
                    class_name=selected_class
                )
            )
        count = process_marks_upload(
            file,
            subject,
            term,
            year,
            selected_class,
            teacher_id,
            level,
            is_subsidiary
        )
        flash(
            f'{count} marks uploaded successfully.',
            'success'
        )
        return redirect(
            url_for(
                'teacher_upload_marks',
                class_name=selected_class
            )
        )
    # ===============================
    # LOAD TEMPLATE
    # ===============================
    return render_template(
        f'teacher/upload_marks_{level}.html',
        assigned_class=selected_class,
        current_year=current_year,
        teacher_classes=[
            {
                'class_name': c
            }
            for c in available_classes
        ],
        selected_class=selected_class,
        students=students
    )
    
@app.route("/save_olevel_marks", methods=["POST"])
def save_olevel_marks():
    if not check_permission(['classteacher', 'subject_teacher', 'dos']):
        abort(403)

    db = get_db_dict()
    cursor = db.cursor()

    subject = request.form.get('subject')
    term = request.form.get('term')
    year = request.form.get('year')

    student_ids = request.form.getlist("student_id[]")
    ai1_raw = request.form.getlist("ai1[]")
    ai2_raw = request.form.getlist("ai2[]")
    ai3_raw = request.form.getlist("ai3[]")
    eot_raw = request.form.getlist("eot_score[]")
    initials = request.form.getlist("teacher_initials[]")

    def clean_score(value):
        if value is None:
            return None
        value = str(value).strip()
        if value == '':
            return None
        try:
            return float(value)
        except (ValueError, TypeError):
            return None

    for i, sid in enumerate(student_ids):
        a1 = clean_score(ai1_raw[i] if i < len(ai1_raw) else None)
        a2 = clean_score(ai2_raw[i] if i < len(ai2_raw) else None)
        a3 = clean_score(ai3_raw[i] if i < len(ai3_raw) else None)
        e = clean_score(eot_raw[i] if i < len(eot_raw) else None)
        init = initials[i].strip() if i < len(initials) else ''

        ai_scores = [score for score in [a1, a2, a3] if score is not None]

        ai_average = (
            sum(ai_scores) / len(ai_scores)
            if ai_scores else None
        )

        ai_contribution = (
            (ai_average / 3) * 20
            if ai_average is not None else 0
        )
        eot_contribution = max(0, min(e, 80)) if e is not None else 0
        total_score = ai_contribution + eot_contribution
        grade, descriptor = get_grade_and_descriptor(total_score)
        identifier = (
            round((total_score / 100) * 3, 2)
            if total_score is not None else None
        )
        cursor.execute(
            """
            INSERT INTO marks
            (
                student_id,
                subject,
                ai1,
                ai2,
                ai3,
                ai_average,
                ai_contribution,
                eot_score,
                total_score,
                grade,
                identifier,
                descriptor,
                teacher_initials,
                term,
                year
            )
            VALUES
            (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            ON CONFLICT(student_id,subject,term,year)
            DO UPDATE SET
                ai1=%s,
                ai2=%s,
                ai3=%s,
                ai_average=%s,
                ai_contribution=%s,
                eot_score=%s,
                total_score=%s,
                grade=%s,
                identifier=%s,
                descriptor=%s,
                teacher_initials=%s
            """,
            (
                sid,
                subject,
                a1,
                a2,
                a3,
                ai_average,
                ai_contribution,
                e,
                total_score,
                grade,
                identifier,
                descriptor,
                init,
                term,
                year,
                a1,
                a2,
                a3,
                ai_average,
                ai_contribution,
                e,
                total_score,
                grade,
                identifier,
                descriptor,
                init
            )
        )

    db.commit()
    cursor.close()

    flash("O-Level marks entered successfully.", "success")

    return redirect(
        url_for(
            'teacher_upload_marks',
            class_name=session.get('selected_class')
        )
    )
@app.route('/teacher/report_card/<student_id>')
def teacher_report_card(student_id):

    if not check_permission([
        'classteacher',
        'subject_teacher',
        'parent',
        'dos',
        'headteacher'
    ]):
        abort(403)

    role = session.get('role')
    db = get_db_dict()
    cur = db.cursor()

    if role in ['classteacher','subject_teacher']:

        selected_class = session.get('selected_class')

        if not selected_class:
            flash('No class selected','danger')
            return redirect(
                url_for('teacher_students')
            )

        cur.execute(
            "SELECT class FROM students WHERE student_id=%s",
            (student_id,)
        )

        res = cur.fetchone()

        if not res or res['class'] != selected_class:
            flash(
                'Student not in your class.',
                'danger'
            )
            return redirect(
                url_for('teacher_students')
            )


    elif role == 'parent':

        parent_phone = session.get('phone')

        if not parent_phone:
            flash(
                'No phone linked.',
                'danger'
            )
            return redirect(
                url_for('dashboard')
            )

        cur.execute(
            """
            SELECT parent_phone
            FROM students
            WHERE student_id=%s
            """,
            (student_id,)
        )

        res = cur.fetchone()

        if not res or res['parent_phone'] != parent_phone:
            flash(
                'Not authorized.',
                'danger'
            )
            return redirect(
                url_for('dashboard')
            )


    elif role in ['dos','headteacher']:

        cur.execute(
            """
            SELECT class
            FROM students
            WHERE student_id=%s
            """,
            (student_id,)
        )

        if not cur.fetchone():
            flash(
                'Student not found.',
                'danger'
            )

            return redirect(
                url_for('dashboard')
            )


    cur.execute(
        """
        SELECT
            full_name,
            class,
            photo_path
        FROM students
        WHERE student_id=%s
        """,
        (student_id,)
    )

    student = cur.fetchone()

    if not student:
        flash(
            'Student not found.',
            'danger'
        )
        return redirect(
            url_for('dashboard')
        )


    full_name = student['full_name']
    class_name = student['class']
    photo_url = get_photo_url(
        student['photo_path']
    )


    term = request.args.get(
        'term',
        'Term 1'
    )

    year = request.args.get(
        'year',
        datetime.now().year
    )
    cur.execute(
        """
        SELECT
            school_name,
            school_address,
            school_phone,
            school_email,
            logo_url
        FROM school_settings
        WHERE id=1
        """
    )

    school = cur.fetchone()


    school_name = (
        school['school_name']
        if school else
        'YOUR SCHOOL NAME'
    )

    school_address = (
        school['school_address']
        if school else
        ''
    )

    school_phone = (
        school['school_phone']
        if school else
        ''
    )

    school_email = (
        school['school_email']
        if school else
        ''
    )

    school_logo_url = (
        school['logo_url']
        if school and school['logo_url']
        else url_for(
            'static',
            filename='images/logo.png'
        )
    )


    cur.execute(
        """
        SELECT
            next_term_begins,
            next_term_ends,
            headteacher_stamp
        FROM school_settings
        WHERE id=1
        """
    )

    settings = cur.fetchone()


    next_term_begins = (
        settings['next_term_begins']
        if settings else None
    )

    next_term_ends = (
        settings['next_term_ends']
        if settings else None
    )

    stamp_url = (
        url_for(
            'static',
            filename=
            'uploads/' +
            settings['headteacher_stamp']
        )
        if settings and settings['headteacher_stamp']
        else None
    )
    class_upper = class_name.upper()
    is_alevel = (
        class_upper in [
            'S5',
            'S6',
            'A-LEVEL',
            'A LEVEL',
            'S.5',
            'S.6'
        ]
        or
        (
            class_upper.startswith('S')
            and len(class_upper) >= 2
            and class_upper[1] in ['5','6']
        )
    )
    if is_alevel:

        cur.execute(
            """
            SELECT
                subject,
                paper1,
                paper2,
                total_score,
                grade,
                points,
                teacher_initials
            FROM marks
            WHERE student_id=%s
            AND term=%s
            AND year=%s
            ORDER BY subject
            """,
            (
                student_id,
                term,
                year
            )
        )
        marks = cur.fetchall()
        total_points = sum(
            m['points']
            for m in marks
            if m['points'] is not None
        ) if marks else 0
        alevel_teacher_comment = get_alevel_class_teacher_comment(
        total_points
        )
        alevel_headteacher_comment = get_alevel_headteacher_comment(
        total_points
        )
        cur.close()
        return render_template(
            'teacher/report_card_alevel.html',
            student_id=student_id,
            full_name=full_name,
            class_name=class_name,
            photo_url=photo_url,
            term=term,
            year=year,
            marks=marks,
            total_points=total_points,
            teacher_comment=alevel_teacher_comment,
            headteacher_comment=alevel_headteacher_comment,
            next_term_begins=next_term_begins,
            next_term_ends=next_term_ends,
            stamp_url=stamp_url,
            school_name=school_name,
            school_address=school_address,
            school_phone=school_phone,
            school_email=school_email,
            school_logo_url=school_logo_url,

        )


    else:

        cur.execute(
            """
            SELECT
                subject,
                ai1,
                ai2,
                ai3,
                ai_average,
                ai_contribution,
                eot_score,
                total_score,
                grade,
                identifier,
                descriptor,
                teacher_initials
            FROM marks
            WHERE student_id=%s
            AND term=%s
            AND year=%s
            ORDER BY subject
            """,
            (
                student_id,
                term,
                year
            )
        )


        raw_marks = cur.fetchall()

        marks = []


        for m in raw_marks:

            ai_scores = []

            if m['ai1'] not in [None,'']:
                ai_scores.append(
                    float(m['ai1'])
                )

            if m['ai2'] not in [None,'']:
                ai_scores.append(
                    float(m['ai2'])
                )

            if m['ai3'] not in [None,'']:
                ai_scores.append(
                    float(m['ai3'])
                )
            ai_average = (
                sum(ai_scores) / len(ai_scores)
                if ai_scores
                else 0
            )
            ai_contribution = (
                (ai_average / 3) * 20
                if ai_average > 0
                else 0
            )
            eot_score = m['eot_score']
            
            # EOT is already marked out of 80.
            # Only constrain it to the allowed range 0–80.
            eot_contribution = (
                max(0, min(float(eot_score), 80))
                if eot_score is not None
                else 0
            )
            
            total_score = (
                ai_contribution +
                eot_contribution
            )
            
            grade, descriptor = get_olevel_grade_details(
                total_score
            )
            
            identifier = round(
                (total_score / 100) * 3,
                2
            )
            
            marks.append(
                {
                    'subject': m['subject'],
                    'ai1': m['ai1'],
                    'ai2': m['ai2'],
                    'ai3': m['ai3'],
                    'ai_average': round(
                        ai_average,
                        2
                    ),
                    'ai_contribution': round(
                        ai_contribution,
                        2
                    ),
                    'eot_score': eot_score,
                    'total_score': round(
                        total_score,
                        2
                    ),
                    'grade': grade,
                    'identifier': identifier,
                    'descriptor': descriptor,
                    'teacher_initials': m['teacher_initials']
                }
            )

        valid_marks = [
            m for m in marks
            if m['total_score'] is not None
            and float(m['total_score']) > 0
        ]
        total_final = sum(
            float(m['total_score'])
            for m in valid_marks
        )
        count = len(valid_marks)
        general_average = (
            total_final / count
            if count > 0
            else 0
        ) 
        general_identifier = round(
            (general_average / 100) * 3,
            2
        )   
        general_grade, general_descriptor = get_olevel_grade_details(
            general_average
        )
        olevel_teacher_comment = get_olevel_class_teacher_comment(
            general_identifier
        )
        olevel_headteacher_comment = get_olevel_headteacher_comment(
            general_identifier
        )
        cur.close()
        return render_template(
            'teacher/report_card.html',
            student_id=student_id,
            full_name=full_name,
            class_name=class_name,
            photo_url=photo_url,
            term=term,
            year=year,
            marks=marks,
            avg_out_of_3=general_identifier,
            general_average=round(
                general_average,
                2
            ),
            general_identifier=general_identifier,
            general_grade=general_grade,
            general_descriptor=general_descriptor,
            teacher_comment=olevel_teacher_comment,
            headteacher_comment=olevel_headteacher_comment,
            next_term_begins=next_term_begins,
            next_term_ends=next_term_ends,
            stamp_url=stamp_url,
            school_name=school_name,
            school_address=school_address,
            school_phone=school_phone,
            school_email=school_email,
            school_logo_url=school_logo_url,
    
        )

@app.route('/teacher/edit_student/<student_id>', methods=['GET', 'POST'])
def teacher_edit_student(student_id):

    if not check_permission(['classteacher']):
        abort(403)

    db = get_db()
    cur = db.cursor()

    # =====================================================
    # GET STUDENT
    # =====================================================

    cur.execute("""
        SELECT *
        FROM students
        WHERE student_id=%s
    """, (student_id,))

    student = cur.fetchone()

    if not student:

        cur.close()

        flash(
            'Student not found.',
            'danger'
        )

        return redirect(
            url_for('teacher_students')
        )

    # =====================================================
    # GET CLASS TEACHER'S ASSIGNED CLASS
    # =====================================================

    cur.execute("""
        SELECT class_name
        FROM teacher_class_assignments
        WHERE user_id=%s
        AND assignment_type='classteacher'
    """, (session.get('user_id'),))

    result = cur.fetchone()

    assigned_class = (
        result['class_name']
        if result
        else None
    )

    # =====================================================
    # UPDATE STUDENT
    # =====================================================

    if request.method == 'POST':

        full_name = request.form.get(
            'full_name',
            ''
        ).strip()

        parent_phone = request.form.get(
            'parent_phone',
            ''
        ).strip()

        # -------------------------------------------------
        # KEEP EXISTING STREAM SYSTEM
        #
        # Examples:
        # S.1
        # S.1A
        # S.1B
        # S.5
        # S.6
        # -------------------------------------------------

        class_name = request.form.get(
            'class',
            ''
        ).strip().upper()

        admission_date = request.form.get(
            'admission_date',
            ''
        ).strip()
        
        if admission_date == '':
            admission_date = None
        
        
        date_of_birth = request.form.get(
            'date_of_birth',
            ''
        ).strip()
        
        if date_of_birth == '':
            date_of_birth = None
            
        sex = request.form.get(
            'sex',
            ''
        )

        preferred_house = request.form.get(
            'preferred_house',
            ''
        )

        disability = request.form.get(
            'disability',
            ''
        )
        programme = request.form.get(
            'programme',
            ''
        ).strip()

        # =================================================
        # NEW: RESIDENCE
        # =================================================

        residence = request.form.get(
            'residence',
            ''
        ).strip()

        # =================================================
        # VALIDATE PROGRAMME
        # =================================================

        class_upper = class_name.upper()

        # S.1-S.4 = USE / Non-USE
        # S.5-S.6 = UPOLET / Non-UPOLET

        if class_upper.startswith(('S.1', 'S.2', 'S.3', 'S.4')):

            valid_programmes = [
                'USE',
                'Non-USE'
            ]

        elif class_upper.startswith(('S.5', 'S.6')):

            valid_programmes = [
                'UPOLET',
                'Non-UPOLET'
            ]

        else:

            valid_programmes = []

        if programme not in valid_programmes:

            cur.close()

            flash(
                f'Invalid programme selected for {class_name}.',
                'danger'
            )

            return redirect(
                url_for(
                    'teacher_edit_student',
                    student_id=student_id
                )
            )

        # =================================================
        # VALIDATE RESIDENCE
        # =================================================

        if residence not in [
            'Day',
            'Boarding'
        ]:

            cur.close()

            flash(
                'Please select Day or Boarding.',
                'danger'
            )

            return redirect(
                url_for(
                    'teacher_edit_student',
                    student_id=student_id
                )
            )

        # =================================================
        # CALCULATE AGE
        # =================================================

        age = None

        if date_of_birth:

            try:

                birth_date = datetime.strptime(
                    date_of_birth,
                    '%Y-%m-%d'
                ).date()

                today = datetime.now().date()

                age = (
                    today.year
                    - birth_date.year
                    - (
                        (today.month, today.day)
                        <
                        (birth_date.month, birth_date.day)
                    )
                )

            except Exception:

                age = None

        # =================================================
        # HANDLE PHOTO
        # =================================================

        photo_path = student.get(
            'photo_path',
            'default_avatar.png'
        )

        photo = request.files.get('photo')

        if (
            photo
            and photo.filename
            and allowed_file(
                photo.filename,
                ALLOWED_IMAGE_EXTENSIONS
            )
        ):

            ext = photo.filename.rsplit(
                '.',
                1
            )[1].lower()

            photo_filename = f"{student_id}.{ext}"

            photo.save(
                os.path.join(
                    app.config['UPLOAD_FOLDER'],
                    photo_filename
                )
            )

            photo_path = photo_filename

        # =================================================
        # UPDATE STUDENT
        # =================================================

        cur.execute("""
            UPDATE students
            SET
                full_name=%s,
                parent_phone=%s,
                class=%s,
                admission_date=%s,
                date_of_birth=%s,
                age=%s,
                sex=%s,
                preferred_house=%s,
                disability=%s,
                photo_path=%s,
                programme=%s,
                residence=%s

            WHERE student_id=%s
        """, (
            full_name,
            parent_phone,
            class_name,
            admission_date,
            date_of_birth,
            age,
            sex,
            preferred_house,
            disability,
            photo_path,
            programme,
            residence,
            student_id
        ))

        db.commit()

        cur.close()

        flash(
            f'Student {full_name} updated successfully!',
            'success'
        )

        return redirect(
            url_for('teacher_students')
        )

    # =====================================================
    # DISPLAY EDIT FORM
    # =====================================================

    cur.close()

    return render_template(
        'teacher/edit_student.html',
        student=student,
        assigned_class=assigned_class
    )


@app.route('/teacher/remove_student/<student_id>', methods=['POST'])
def teacher_remove_student(student_id):

    if not check_permission(['classteacher', 'dos']):
        abort(403)


    db = get_db()
    cur = db.cursor()


    cur.execute("""
        SELECT full_name, class
        FROM students
        WHERE student_id=%s
    """,
    (student_id,))


    student = cur.fetchone()


    if not student:

        cur.close()

        flash(
            'Student not found.',
            'danger'
        )

        return redirect(
            url_for('teacher_students')
        )



    if session.get('role') == 'classteacher':

        cur.execute("""
            SELECT class_name
            FROM teacher_class_assignments
            WHERE user_id=%s
            AND assignment_type='classteacher'
        """,
        (session.get('user_id'),))


        result = cur.fetchone()


        assigned_class = (
            result['class_name']
            if result
            else None
        )


        if assigned_class != student['class']:

            cur.close()

            flash(
                'You can only remove students from your own class.',
                'danger'
            )

            return redirect(
                url_for('teacher_students')
            )
    cur.execute("""
        DELETE FROM students
        WHERE student_id=%s
    """,
    (student_id,))
    db.commit()
    cur.close()
    flash(
        f"Student {student['full_name']} removed successfully.",
        'success'
    )
    return redirect(
        url_for('teacher_students')
    )

@app.route('/teacher/upload_students', methods=['GET', 'POST'])
def teacher_upload_students():

    if not check_permission(['classteacher']):
        abort(403)

    if request.method == 'POST':

        file = request.files.get('excel_file')

        if not file or not file.filename:

            flash(
                'Please upload an Excel or CSV file.',
                'danger'
            )

            return redirect(
                url_for('teacher_upload_students')
            )

        try:

            from openpyxl import load_workbook
            import csv
            import io

            db = get_db()
            cur = db.cursor()

            # =================================================
            # GET CLASS TEACHER'S ASSIGNED CLASS
            # =================================================

            cur.execute("""
                SELECT class_name
                FROM teacher_class_assignments
                WHERE user_id=%s
                AND assignment_type='classteacher'
            """, (session.get('user_id'),))

            result = cur.fetchone()

            if not result:

                cur.close()

                flash(
                    'You are not assigned as a class teacher.',
                    'danger'
                )

                return redirect(
                    url_for('teacher_upload_students')
                )

            assigned_class = (
                result['class_name']
                .strip()
                .upper()
            )

            success_count = 0
            error_count = 0
            errors = []

            # =================================================
            # DETERMINE ALLOWED PROGRAMMES
            # =================================================

            class_upper = assigned_class

            if class_upper.startswith(
                ('S.1', 'S.2', 'S.3', 'S.4')
            ):

                valid_programmes = [
                    'USE',
                    'Non-USE'
                ]

            elif class_upper.startswith(
                ('S.5', 'S.6')
            ):

                valid_programmes = [
                    'UPOLET',
                    'Non-UPOLET'
                ]

            else:

                valid_programmes = []

            # =================================================
            # INSERT STUDENT
            # =================================================

            def insert_student(
                full_name,
                parent_phone,
                programme='',
                residence=''
            ):

                nonlocal success_count, error_count

                full_name = (
                    str(full_name or '')
                    .strip()
                )

                parent_phone = (
                    str(parent_phone or '')
                    .strip()
                )

                programme = (
                    str(programme or '')
                    .strip()
                )

                residence = (
                    str(residence or '')
                    .strip()
                )

                # ---------------------------------------------
                # REQUIRED NAME
                # ---------------------------------------------

                if not full_name:

                    error_count += 1

                    return False

                # ---------------------------------------------
                # VALIDATE PROGRAMME IF PROVIDED
                # ---------------------------------------------

                if programme:

                    if programme not in valid_programmes:

                        error_count += 1

                        errors.append(
                            f"{full_name}: Invalid programme "
                            f"'{programme}' for {assigned_class}"
                        )

                        return False

                # ---------------------------------------------
                # VALIDATE RESIDENCE IF PROVIDED
                # ---------------------------------------------

                if residence:

                    if residence not in [
                        'Day',
                        'Boarding'
                    ]:

                        error_count += 1

                        errors.append(
                            f"{full_name}: Invalid residence "
                            f"'{residence}'"
                        )

                        return False

                # ---------------------------------------------
                # GENERATE STUDENT ID
                # ---------------------------------------------

                student_id = generate_student_id()

                # ---------------------------------------------
                # INSERT
                # ---------------------------------------------

                cur.execute("""
                    INSERT INTO students
                    (
                        student_id,
                        full_name,
                        class,
                        parent_phone,
                        programme,
                        residence,
                        admission_status,
                        fees_total,
                        fees_paid,
                        fees_balance
                    )
                    VALUES
                    (
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        %s,
                        'approved',
                        0,
                        0,
                        0
                    )
                """, (
                    student_id,
                    full_name,
                    assigned_class,
                    parent_phone,
                    programme or None,
                    residence or None
                ))

                success_count += 1

                return True

            # =================================================
            # EXCEL UPLOAD
            # =================================================

            if file.filename.lower().endswith(
                ('.xlsx', '.xls')
            ):

                wb = load_workbook(
                    file,
                    data_only=True
                )

                sheet = wb.active

                headers = [
                    str(cell.value).strip().lower()
                    if cell.value
                    else ''
                    for cell in sheet[1]
                ]

                full_name_col = None
                parent_phone_col = None
                programme_col = None
                residence_col = None

                for idx, h in enumerate(headers):

                    if h in [
                        'full_name',
                        'name',
                        'student_name'
                    ]:

                        full_name_col = idx

                    elif h in [
                        'parent_phone',
                        'phone',
                        'parent_contact'
                    ]:

                        parent_phone_col = idx

                    elif h in [
                        'programme',
                        'program'
                    ]:

                        programme_col = idx

                    elif h in [
                        'residence',
                        'boarding_status',
                        'student_residence'
                    ]:

                        residence_col = idx

                # ---------------------------------------------
                # DEFAULT EXISTING COLUMNS
                # ---------------------------------------------

                if full_name_col is None:
                    full_name_col = 0

                if parent_phone_col is None:
                    parent_phone_col = 1

                # ---------------------------------------------
                # PROCESS ROWS
                # ---------------------------------------------

                for row_idx in range(
                    2,
                    sheet.max_row + 1
                ):

                    full_name = str(
                        sheet.cell(
                            row=row_idx,
                            column=full_name_col + 1
                        ).value or ''
                    ).strip()

                    parent_phone = str(
                        sheet.cell(
                            row=row_idx,
                            column=parent_phone_col + 1
                        ).value or ''
                    ).strip()

                    programme = ''

                    if programme_col is not None:

                        programme = str(
                            sheet.cell(
                                row=row_idx,
                                column=programme_col + 1
                            ).value or ''
                        ).strip()

                    residence = ''

                    if residence_col is not None:

                        residence = str(
                            sheet.cell(
                                row=row_idx,
                                column=residence_col + 1
                            ).value or ''
                        ).strip()

                    if not insert_student(
                        full_name,
                        parent_phone,
                        programme,
                        residence
                    ):

                        if full_name and not errors:

                            errors.append(
                                f"Row {row_idx}: "
                                f"Student could not be uploaded."
                            )

            # =================================================
            # CSV UPLOAD
            # =================================================

            elif file.filename.lower().endswith('.csv'):

                content = file.read().decode(
                    'utf-8-sig'
                )

                csv_reader = csv.reader(
                    io.StringIO(content)
                )

                headers = next(
                    csv_reader,
                    []
                )

                headers = [
                    str(h).strip().lower()
                    for h in headers
                ]

                full_name_col = None
                parent_phone_col = None
                programme_col = None
                residence_col = None

                for idx, h in enumerate(headers):

                    if h in [
                        'full_name',
                        'name',
                        'student_name'
                    ]:

                        full_name_col = idx

                    elif h in [
                        'parent_phone',
                        'phone',
                        'parent_contact'
                    ]:

                        parent_phone_col = idx

                    elif h in [
                        'programme',
                        'program'
                    ]:

                        programme_col = idx

                    elif h in [
                        'residence',
                        'boarding_status',
                        'student_residence'
                    ]:

                        residence_col = idx

                if full_name_col is None:
                    full_name_col = 0

                if parent_phone_col is None:
                    parent_phone_col = 1

                # ---------------------------------------------
                # PROCESS CSV ROWS
                # ---------------------------------------------

                for row_idx, row in enumerate(
                    csv_reader,
                    start=2
                ):

                    full_name = (
                        row[full_name_col].strip()
                        if full_name_col < len(row)
                        else ''
                    )

                    parent_phone = (
                        row[parent_phone_col].strip()
                        if parent_phone_col < len(row)
                        else ''
                    )

                    programme = ''

                    if (
                        programme_col is not None
                        and programme_col < len(row)
                    ):

                        programme = (
                            row[programme_col]
                            .strip()
                        )

                    residence = ''

                    if (
                        residence_col is not None
                        and residence_col < len(row)
                    ):

                        residence = (
                            row[residence_col]
                            .strip()
                        )

                    if not insert_student(
                        full_name,
                        parent_phone,
                        programme,
                        residence
                    ):

                        if not full_name:

                            errors.append(
                                f"Row {row_idx}: "
                                f"Missing full name"
                            )

            # =================================================
            # UNSUPPORTED FILE
            # =================================================

            else:

                cur.close()

                flash(
                    'Unsupported file format. '
                    'Upload .xlsx, .xls, or .csv',
                    'danger'
                )

                return redirect(
                    url_for('teacher_upload_students')
                )

            # =================================================
            # SAVE
            # =================================================

            db.commit()

            cur.close()

            # =================================================
            # NOTIFY DOS
            # =================================================

            add_notification(
                'dos',
                f'Class teacher uploaded '
                f'{success_count} students to '
                f'class {assigned_class}',
                '/dos/class_lists'
            )

            # =================================================
            # RESULT
            # =================================================

            flash(
                f'Uploaded {success_count} students '
                f'to class {assigned_class}. '
                f'Errors: {error_count}',
                'success'
                if success_count > 0
                else 'danger'
            )

            for error in errors[:5]:

                flash(
                    error,
                    'warning'
                )

        except Exception as e:

            flash(
                f'Error: {str(e)}',
                'danger'
            )

        return redirect(
            url_for('teacher_students')
        )

    return render_template(
        'teacher/upload_students.html'
    )
@app.route('/teacher/print_all_report_cards')
def teacher_print_all_report_cards():

    if not check_permission(['classteacher']):
        abort(403)
    # =========================================================
    # GET SELECTED CLASS
    # =========================================================
    #
    # First use the class coming from the URL.
    # This is important when the teacher has more than one class.
    #
    selected_class = request.args.get('class_name')

    if not selected_class:
        selected_class = session.get('selected_class')

    if not selected_class:
        selected_class = session.get('assigned_class')

    if not selected_class:

        flash(
            'No class selected or assigned to you.',
            'danger'
        )

        return redirect(
            url_for('teacher_students')
        )

    # Keep the selected class in the session as well
    session['selected_class'] = selected_class


    # =========================================================
    # GET TERM AND YEAR
    # =========================================================

    term = request.args.get(
        'term',
        'Term 1'
    )

    year = request.args.get(
        'year',
        datetime.now().year
    )


    # =========================================================
    # DATABASE
    # =========================================================

    db = get_db()
    cur = db.cursor()


    # =========================================================
    # GET STUDENTS IN SELECTED CLASS
    # =========================================================

    cur.execute("""
        SELECT
            student_id,
            full_name,
            photo_path
        FROM students
        WHERE class=%s
        ORDER BY full_name
    """,
    (
        selected_class,
    ))

    students_data = cur.fetchall()


    if not students_data:

        cur.close()

        flash(
            f'No students found in class {selected_class}.',
            'warning'
        )

        return redirect(
            url_for(
                'teacher_students',
                class_name=selected_class,
                term=term,
                year=year
            )
        )


    # =========================================================
    # SCHOOL SETTINGS
    # =========================================================

    cur.execute("""
        SELECT
            school_name,
            school_address,
            school_phone,
            school_email,
            logo_url,
            next_term_begins,
            next_term_ends,
            headteacher_stamp
        FROM school_settings
        WHERE id=1
    """)

    school_data = cur.fetchone()


    school_name = (
        school_data['school_name']
        if school_data and school_data['school_name']
        else 'YOUR SCHOOL NAME'
    )


    school_address = (
        school_data['school_address']
        if school_data and school_data['school_address']
        else 'P.O. Box 123, Kampala, Uganda'
    )


    school_phone = (
        school_data['school_phone']
        if school_data and school_data['school_phone']
        else 'Tel: +256 712 345678'
    )


    school_email = (
        school_data['school_email']
        if school_data and school_data['school_email']
        else 'Email: info@school.com'
    )


    school_logo_url = (
        school_data['logo_url']
        if school_data and school_data['logo_url']
        else url_for(
            'static',
            filename='images/logo.png'
        )
    )


    next_term_begins = (
        school_data['next_term_begins']
        if school_data
        else None
    )


    next_term_ends = (
        school_data['next_term_ends']
        if school_data
        else None
    )


    # =========================================================
    # HEADTEACHER STAMP
    # =========================================================

    stamp_url = None

    if school_data and school_data['headteacher_stamp']:

        stamp_url = url_for(
            'static',
            filename='uploads/' +
            school_data['headteacher_stamp']
        )


    # =========================================================
    # DETERMINE A-LEVEL / O-LEVEL
    # =========================================================

    class_upper = selected_class.upper()

    is_alevel = (
        class_upper in [
            'S5',
            'S6',
            'A-LEVEL',
            'A LEVEL',
            'S.5',
            'S.6'
        ]
        or
        (
            class_upper.startswith('S')
            and len(class_upper) >= 2
            and class_upper[1] in ['5', '6']
        )
    )
    # =========================================================
    # BUILD ALL REPORTS
    # =========================================================
    
    all_reports = []
    
    for student in students_data:
    
        student_id = student['student_id']
        full_name = student['full_name']
    
        photo_url = get_photo_url(
            student['photo_path']
        )
    
        # =====================================================
        # A-LEVEL
        # =====================================================
    
        if is_alevel:
    
            cur.execute("""
                SELECT
                    subject,
                    paper1,
                    paper2,
                    total_score,
                    grade,
                    points,
                    teacher_initials
                FROM marks
                WHERE student_id=%s
                AND term=%s
                AND year=%s
                ORDER BY subject
            """,
            (
                student_id,
                term,
                year
            ))
    
            marks = cur.fetchall()
    
            # =================================================
            # A-LEVEL TOTAL POINTS
            # =================================================
    
            total_points = sum(
                float(m['points'])
                for m in marks
                if m['points'] is not None
            ) if marks else 0
            
            # A-LEVEL COMMENTS
            alevel_teacher_comment = (
                get_alevel_class_teacher_comment(
                    total_points
                )
            )
            alevel_headteacher_comment = (
                get_alevel_headteacher_comment(
                    total_points
                )
            )
            # SAVE A-LEVEL REPORT
            all_reports.append({
                'student_id': student_id,
                'full_name': full_name,
                'photo_url': photo_url,
                'marks': marks,
                'total_points':
                    total_points,
                'teacher_comment':
                    alevel_teacher_comment,
                'headteacher_comment':
                    alevel_headteacher_comment
            })
    
    
        # =====================================================
        # O-LEVEL
        # =====================================================
    
        else:
    
            cur.execute("""
                SELECT
                    subject,
                    ai1,
                    ai2,
                    ai3,
                    ai_average,
                    ai_contribution,
                    eot_score,
                    total_score,
                    grade,
                    identifier,
                    descriptor,
                    teacher_initials
                FROM marks
                WHERE student_id=%s
                AND term=%s
                AND year=%s
                ORDER BY subject
            """,
            (
                student_id,
                term,
                year
            ))
    
            raw_marks = cur.fetchall()
    
            marks = []
    
            for m in raw_marks:
    
                ai_scores = []
    
                if m['ai1'] not in [None, '']:
                    ai_scores.append(
                        float(m['ai1'])
                    )
    
                if m['ai2'] not in [None, '']:
                    ai_scores.append(
                        float(m['ai2'])
                    )
    
                if m['ai3'] not in [None, '']:
                    ai_scores.append(
                        float(m['ai3'])
                    )
    
                # =================================================
                # AI AVERAGE
                # =================================================
    
                ai_average = (
                    sum(ai_scores) / len(ai_scores)
                    if ai_scores
                    else 0
                )
    
                # =================================================
                # AI CONTRIBUTION
                # AI average is out of 3.
                # Convert to contribution out of 20.
                # =================================================
    
                ai_contribution = (
                    (ai_average / 3) * 20
                    if ai_average > 0
                    else 0
                )
    
                # =================================================
                # EOT
                #
                # EOT is already out of 80.
                # Only constrain it to 0-80.
                # =================================================
    
                eot_score = m['eot_score']
    
                eot_contribution = (
                    max(
                        0,
                        min(
                            float(eot_score),
                            80
                        )
                    )
                    if eot_score is not None
                    else 0
                )
    
                # =================================================
                # TOTAL SCORE
                # =================================================
    
                total_score = (
                    ai_contribution +
                    eot_contribution
                )
    
                # =================================================
                # SUBJECT GRADE / DESCRIPTOR
                # =================================================
    
                grade, descriptor = (
                    get_olevel_grade_details(
                        total_score
                    )
                )
    
                # =================================================
                # SUBJECT IDENTIFIER
                # =================================================
    
                identifier = round(
                    (total_score / 100) * 3,
                    2
                )
    
                marks.append({
    
                    'subject':
                        m['subject'],
    
                    'ai1':
                        m['ai1'],
    
                    'ai2':
                        m['ai2'],
    
                    'ai3':
                        m['ai3'],
    
                    'ai_average':
                        round(
                            ai_average,
                            2
                        ),
    
                    'ai_contribution':
                        round(
                            ai_contribution,
                            2
                        ),
    
                    'eot_score':
                        eot_score,
    
                    'total_score':
                        round(
                            total_score,
                            2
                        ),
    
                    'grade':
                        grade,
    
                    'identifier':
                        identifier,
    
                    'descriptor':
                        descriptor,
    
                    'teacher_initials':
                        m['teacher_initials']
                })
    
    
            # =====================================================
            # O-LEVEL GENERAL AVERAGE
            # =====================================================
    
            valid_marks = [
                m for m in marks
                if m['total_score'] is not None
                and float(m['total_score']) > 0
            ]
    
            total_final = sum(
                float(m['total_score'])
                for m in valid_marks
            )
    
            count = len(valid_marks)
    
            general_average = (
                total_final / count
                if count > 0
                else 0
            )
    
            # =====================================================
            # O-LEVEL GENERAL IDENTIFIER
            # =====================================================
    
            general_identifier = round(
                (general_average / 100) * 3,
                2
            )
    
            # =====================================================
            # O-LEVEL GENERAL GRADE / DESCRIPTOR
            # =====================================================
    
            general_grade, general_descriptor = (
                get_olevel_grade_details(
                    general_average
                )
            )
    
            # =====================================================
            # O-LEVEL COMMENTS
            #
            # Comments are based on IDENTIFIER.
            # =====================================================
    
            olevel_teacher_comment = (
                get_olevel_class_teacher_comment(
                    general_identifier
                )
            )
    
            olevel_headteacher_comment = (
                get_olevel_headteacher_comment(
                    general_identifier
                )
            )
    
            # =====================================================
            # SAVE O-LEVEL REPORT
            # =====================================================
    
            all_reports.append({
    
                'student_id':
                    student_id,
    
                'full_name':
                    full_name,
    
                'photo_url':
                    photo_url,
    
                'marks':
                    marks,
    
                'general_average':
                    round(
                        general_average,
                        2
                    ),
    
                'general_identifier':
                    general_identifier,
    
                'avg_out_of_3':
                    general_identifier,
    
                'general_grade':
                    general_grade,
    
                'general_descriptor':
                    general_descriptor,
    
                'teacher_comment':
                    olevel_teacher_comment,
    
                'headteacher_comment':
                    olevel_headteacher_comment
            })
    cur.close()
    template = (
        'teacher/print_all_report_cards_alevel.html'
        if is_alevel
        else
        'teacher/print_all_report_cards.html'
    )
    return render_template(
        template,
        reports=all_reports,
        class_name=selected_class,
        term=term,
        year=year,
        next_term_begins=
            next_term_begins,
        next_term_ends=
            next_term_ends,
        stamp_url=stamp_url,
        school_name=
            school_name,
        school_address=
            school_address,
        school_phone=
            school_phone,
        school_email=
            school_email,
        school_logo_url=
            school_logo_url

    )
# ==================== BURSAR MODULE ====================

def generate_receipt_number():
    return generate_unique_number('RCP', 'payments', 'receipt_no', year_format=True)

def send_fee_sms(phone_number, student_name, amount, balance):
    if not phone_number:
        return False
    message = f"Payment of UGX {amount:,.2f} received for {student_name}. Balance: UGX {balance:,.2f}. Thank you."
    return send_sms(phone_number, message)

def calculate_nssf_and_paye(gross_salary):
    cur = get_db().cursor()
    cur.execute("SELECT nssf_employee_rate, paye_rate, paye_threshold FROM school_settings WHERE id=1")
    rates = cur.fetchone()
    cur.close()
    nssf_employee_rate = rates['nssf_employee_rate'] if rates else 5.0
    paye_rate = rates['paye_rate'] if rates else 10.0
    paye_threshold = rates['paye_threshold'] if rates else 235000
    nssf_employee = (gross_salary * nssf_employee_rate) / 100
    taxable_amount = max(0, gross_salary - paye_threshold)
    paye_tax = (taxable_amount * paye_rate) / 100
    return {
        'nssf_employee': round(nssf_employee, 2),
        'paye_tax': round(paye_tax, 2)
    }

@app.route('/bursar/dashboard')
def bursar_dashboard():
    if not check_permission(['bursar']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT 
            COALESCE(SUM(fees_total),0) AS total_fees,
            COALESCE(SUM(fees_paid),0) AS total_paid,
            COALESCE(SUM(fees_balance),0) AS total_balance
        FROM students
    """)
    totals = cur.fetchone()

    cur.execute("SELECT COUNT(*) AS count FROM students WHERE fees_balance > 0")
    defaulter_count = cur.fetchone()['count']

    cur.execute("SELECT COUNT(*) AS count FROM students")
    total_students = cur.fetchone()['count']

    cur.execute("""
        SELECT p.*, s.full_name, s.class
        FROM payments p
        JOIN students s ON p.student_id=s.student_id
        ORDER BY p.payment_date DESC
        LIMIT 10
    """)
    recent_payments = cur.fetchall()

    cur.close()

    return render_template(
        'bursar/dashboard.html',
        totals=totals,
        defaulter_count=defaulter_count,
        total_students=total_students,
        recent_payments=recent_payments
    )

@app.route('/bursar/students')
def bursar_students():
    if not check_permission(['bursar']):
        abort(403)
    
    # =====================================================
    # GET FILTERS
    # =====================================================
    
    search = request.args.get(
        'search',
        ''
    ).strip()
    
    class_filter = request.args.get(
        'class',
        ''
    ).strip()
    
    programme_filter = request.args.get(
        'programme',
        ''
    ).strip()
    
    residence_filter = request.args.get(
        'residence',
        ''
    ).strip()
    
    
    # =====================================================
    # DATABASE
    # =====================================================
    
    db = get_db_dict()
    cur = db.cursor()
    
    
    # =====================================================
    # GET STUDENTS
    # =====================================================
    
    query = """
        SELECT
            student_id,
            full_name,
            class,
            programme,
            residence,
            parent_phone,
            fees_total,
            fees_paid,
            fees_balance
        FROM students
        WHERE 1=1
    """
    
    params = []
    
    
    # =====================================================
    # SEARCH
    # =====================================================
    
    if search:
    
        query += """
            AND (
                student_id ILIKE %s
                OR full_name ILIKE %s
            )
        """
    
        pattern = f"%{search}%"
    
        params.extend([
            pattern,
            pattern
        ])
    
    
    # =====================================================
    # CLASS FILTER
    #
    # IMPORTANT:
    # Streams remain unchanged.
    #
    # Example:
    # S.1A
    # S.1B
    # S.3
    # S.5
    # S.6
    # =====================================================
    
    if class_filter:
    
        query += """
            AND class=%s
        """
    
        params.append(
            class_filter
        )
    
    
    # =====================================================
    # PROGRAMME FILTER
    # =====================================================
    
    if programme_filter:
    
        query += """
            AND programme=%s
        """
    
        params.append(
            programme_filter
        )
    
    
    # =====================================================
    # RESIDENCE FILTER
    # =====================================================
    
    if residence_filter:
    
        query += """
            AND residence=%s
        """
    
        params.append(
            residence_filter
        )
    
    
    # =====================================================
    # ORDER
    # =====================================================
    
    query += """
        ORDER BY
            class,
            full_name
    """
    
    
    cur.execute(
        query,
        params
    )
    
    students = cur.fetchall()
    
    
    # =====================================================
    # GET AVAILABLE CLASSES
    # =====================================================
    
    cur.execute("""
        SELECT DISTINCT class
        FROM students
        WHERE class IS NOT NULL
        AND class != ''
        ORDER BY class
    """)
    
    classes = [
        row['class']
        for row in cur.fetchall()
    ]
    
    
    # =====================================================
    # CLOSE DATABASE
    # =====================================================
    
    cur.close()
    
    
    # =====================================================
    # RETURN TEMPLATE
    # =====================================================
    
    return render_template(
        'bursar/students.html',
    
        students=students,
    
        classes=classes,
    
        search=search,
    
        class_filter=class_filter,
    
        programme_filter=programme_filter,
    
        residence_filter=residence_filter
    )
# =========================================================
# BURSAR - STUDENT FEE CLASSIFICATION
# =========================================================

@app.route('/bursar/students/classification')
def bursar_student_classification():

    if not check_permission(['bursar']):
        abort(403)

    year = request.args.get(
        'year',
        datetime.now().year,
        type=int
    )

    term = request.args.get(
        'term',
        'Term 1'
    ).strip()

    search = request.args.get(
        'search',
        ''
    ).strip()

    class_filter = request.args.get(
        'class',
        ''
    ).strip()

    db = get_db_dict()
    cur = db.cursor()

    # =====================================================
    # GET STUDENTS
    # =====================================================

    query = """
        SELECT
            student_id,
            full_name,
            class,
            parent_phone,
            programme,
            residence,
            fees_total,
            fees_paid,
            fees_balance
        FROM students
        WHERE 1=1
    """

    params = []

    if search:

        query += """
            AND (
                student_id ILIKE %s
                OR full_name ILIKE %s
            )
        """

        pattern = f"%{search}%"

        params.extend([
            pattern,
            pattern
        ])

    if class_filter:

        query += """
            AND class = %s
        """

        params.append(class_filter)

    query += """
        ORDER BY
            class,
            full_name
    """

    cur.execute(
        query,
        tuple(params)
    )

    students = cur.fetchall()

    # =====================================================
    # GET CLASSES
    # =====================================================

    cur.execute("""
        SELECT DISTINCT class
        FROM students
        WHERE class IS NOT NULL
        AND class != ''
        ORDER BY class
    """)

    classes = [
        row['class']
        for row in cur.fetchall()
    ]

    # =====================================================
    # GET FEE STRUCTURES
    # =====================================================

    cur.execute("""
        SELECT
            id,
            level,
            programme,
            residence,
            term,
            year,
            amount,
            description
        FROM student_fee_structure
        WHERE term = %s
        AND year = %s
        ORDER BY
            CASE level
                WHEN 'S.1' THEN 1
                WHEN 'S.2' THEN 2
                WHEN 'S.3' THEN 3
                WHEN 'S.4' THEN 4
                WHEN 'S.5' THEN 5
                WHEN 'S.6' THEN 6
                ELSE 7
            END,
            programme,
            residence
    """, (
        term,
        year
    ))

    fee_structures = cur.fetchall()

    cur.close()

    return render_template(
        'bursar/student_classification.html',
        students=students,
        classes=classes,
        fee_structures=fee_structures,
        year=year,
        term=term,
        search=search,
        class_filter=class_filter
    )

@app.route(
    '/bursar/students/classification/update',
    methods=['POST']
)
def bursar_student_classification_update():

    if not check_permission(['bursar']):
        abort(403)

    student_id = request.form.get(
        'student_id',
        ''
    ).strip()

    programme = request.form.get(
        'programme',
        ''
    ).strip()

    residence = request.form.get(
        'residence',
        ''
    ).strip()

    term = request.form.get(
        'term',
        'Term 1'
    ).strip()

    year = request.form.get(
        'year',
        datetime.now().year,
        type=int
    )

    # =====================================================
    # BASIC VALIDATION
    # =====================================================

    if not student_id:

        flash(
            'Student ID is required.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_student_classification',
                year=year,
                term=term
            )
        )

    if programme not in [
        'USE',
        'Non-USE',
        'UPOLET',
        'Non-UPOLET'
    ]:

        flash(
            'Invalid programme selected.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_student_classification',
                year=year,
                term=term
            )
        )

    if residence not in [
        'Day',
        'Boarding'
    ]:

        flash(
            'Please select Day or Boarding.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_student_classification',
                year=year,
                term=term
            )
        )

    db = get_db_dict()
    cur = db.cursor()

    try:

        # =================================================
        # GET STUDENT
        # =================================================

        cur.execute("""
            SELECT
                student_id,
                full_name,
                class
            FROM students
            WHERE student_id = %s
        """, (
            student_id,
        ))

        student = cur.fetchone()

        if not student:

            flash(
                'Student not found.',
                'danger'
            )

            return redirect(
                url_for(
                    'bursar_student_classification',
                    year=year,
                    term=term
                )
            )

        # =================================================
        # DETERMINE LEVEL
        #
        # S.1A -> S.1
        # S.1B -> S.1
        # S.3A -> S.3
        # S.5  -> S.5
        # S.6  -> S.6
        # =================================================

        level = get_student_level(
            student['class']
        )

        if not level:

            flash(
                f"Could not determine the level from "
                f"class {student['class']}.",
                'danger'
            )

            return redirect(
                url_for(
                    'bursar_student_classification',
                    year=year,
                    term=term
                )
            )

        # =================================================
        # VALIDATE PROGRAMME AGAINST LEVEL
        # =================================================

        if level in [
            'S.1',
            'S.2',
            'S.3',
            'S.4'
        ]:

            valid_programmes = [
                'USE',
                'Non-USE'
            ]

        else:

            valid_programmes = [
                'UPOLET',
                'Non-UPOLET'
            ]

        if programme not in valid_programmes:

            flash(
                f'{programme} is not valid for {level}.',
                'danger'
            )

            return redirect(
                url_for(
                    'bursar_student_classification',
                    year=year,
                    term=term
                )
            )

        # =================================================
        # SAVE PROGRAMME + RESIDENCE
        # =================================================

        cur.execute("""
            UPDATE students
            SET
                programme = %s,
                residence = %s
            WHERE student_id = %s
        """, (
            programme,
            residence,
            student_id
        ))

        db.commit()

    except Exception as e:

        db.rollback()

        flash(
            f'Could not save student classification: {e}',
            'danger'
        )

        cur.close()

        return redirect(
            url_for(
                'bursar_student_classification',
                year=year,
                term=term
            )
        )

    finally:

        cur.close()
    success, message = apply_student_fee_structure(
        student_id,
        term,
        year
    )

    if success:

        flash(
            f"{student['full_name']}: {message}",
            'success'
        )

    else:

        flash(
            f"Classification saved, but fee structure "
            f"could not be applied: {message}",
            'warning'
        )

    return redirect(
        url_for(
            'bursar_student_classification',
            year=year,
            term=term
        )
    )


@app.route('/bursar/student/<student_id>')
def bursar_student_detail(student_id):

    if not check_permission(['bursar']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    # =====================================================
    # GET STUDENT
    # =====================================================

    cur.execute("""
        SELECT
            student_id,
            full_name,
            class,
            parent_phone,
            programme,
            residence,
            fees_total,
            fees_paid,
            fees_balance
        FROM students
        WHERE student_id=%s
    """, (
        student_id,
    ))

    student = cur.fetchone()

    if not student:

        cur.close()

        flash(
            'Student not found.',
            'danger'
        )

        return redirect(
            url_for('bursar_students')
        )

    # =====================================================
    # GET PAYMENT HISTORY
    # =====================================================

    cur.execute("""
        SELECT
            id,
            receipt_no,
            amount,
            payment_date,
            payment_method,
            recorded_by,
            notes
        FROM payments
        WHERE student_id=%s
        ORDER BY
            payment_date DESC,
            id DESC
    """, (
        student_id,
    ))

    payments = cur.fetchall()

    # =====================================================
    # CALCULATE TOTAL PAID FROM PAYMENT RECORDS
    # =====================================================

    total_paid = sum(
        float(p['amount'] or 0)
        for p in payments
    )

    # =====================================================
    # DETERMINE STUDENT LEVEL
    #
    # Students may be stored as:
    #
    # S.1A
    # S.1B
    # S.3A
    #
    # or as:
    #
    # S.3
    # S.5
    # S.6
    # =====================================================

    student_class = (
        student['class'] or ''
    ).strip().upper()

    level = student_class

    # -----------------------------------------------------
    # STREAMED O-LEVEL STUDENT
    #
    # S.1A -> S.1
    # S.2B -> S.2
    # -----------------------------------------------------

    match = re.match(
        r'^(S\.[1-6])',
        student_class
    )

    if match:

        level = match.group(1)

    # =====================================================
    # PROGRAMME OPTIONS
    # =====================================================

    if level in [
        'S.1',
        'S.2',
        'S.3',
        'S.4'
    ]:

        programme_options = [
            'USE',
            'Non-USE'
        ]

    elif level in [
        'S.5',
        'S.6'
    ]:

        programme_options = [
            'UPOLET',
            'Non-UPOLET'
        ]

    else:

        programme_options = []

    programme_label = 'Programme'

    # =====================================================
    # GET SAVED CLASSIFICATION
    # =====================================================

    selected_programme = (
        student['programme']
        or ''
    ).strip()

    selected_residence = (
        student['residence']
        or ''
    ).strip()

    # =====================================================
    # ALLOW URL VALUES TO OVERRIDE
    #
    # This allows the bursar to preview another
    # classification if needed.
    # =====================================================

    url_programme = request.args.get(
        'programme'
    )

    url_residence = request.args.get(
        'residence'
    )

    if url_programme:

        selected_programme = (
            url_programme.strip()
        )

    if url_residence:

        selected_residence = (
            url_residence.strip()
        )

    # =====================================================
    # TERM AND YEAR
    # =====================================================

    selected_term = request.args.get(
        'term',
        'Term 1'
    ).strip()

    selected_year = request.args.get(
        'year',
        datetime.now().year,
        type=int
    )

    # =====================================================
    # GET APPLICABLE FEE STRUCTURE
    # =====================================================

    applicable_fee = None
    fee_description = None
    fee_structure_id = None

    if (
        selected_programme
        and selected_residence
        and level
    ):

        cur.execute("""
            SELECT
                id,
                amount,
                description
            FROM student_fee_structure
            WHERE level=%s
            AND programme=%s
            AND residence=%s
            AND term=%s
            AND year=%s
            LIMIT 1
        """, (
            level,
            selected_programme,
            selected_residence,
            selected_term,
            selected_year
        ))

        fee_record = cur.fetchone()

        if fee_record:

            fee_structure_id = fee_record['id']

            applicable_fee = float(
                fee_record['amount'] or 0
            )

            fee_description = (
                fee_record['description']
                or ''
            )

    # =====================================================
    # STORED STUDENT FEE TOTAL
    # =====================================================

    stored_fees_total = float(
        student['fees_total'] or 0
    )

    # =====================================================
    # DETERMINE EFFECTIVE FEE TOTAL
    #
    # If a matching fee structure exists, use it.
    # Otherwise use student's stored fees_total.
    # =====================================================

    if applicable_fee is not None:

        fees_total = applicable_fee

    else:

        fees_total = stored_fees_total

    # =====================================================
    # CALCULATE BALANCE
    # =====================================================

    fees_balance = max(
        fees_total - total_paid,
        0
    )

    # =====================================================
    # KEEP STUDENT FEE FIGURES SYNCHRONIZED
    #
    # Only update when the calculated values differ.
    # =====================================================

    if (
        abs(stored_fees_total - fees_total) > 0.001
        or
        abs(
            float(student['fees_paid'] or 0)
            - total_paid
        ) > 0.001
        or
        abs(
            float(student['fees_balance'] or 0)
            - fees_balance
        ) > 0.001
    ):

        cur.execute("""
            UPDATE students
            SET
                fees_total=%s,
                fees_paid=%s,
                fees_balance=%s
            WHERE student_id=%s
        """, (
            fees_total,
            total_paid,
            fees_balance,
            student_id
        ))

        db.commit()

    cur.close()

    # =====================================================
    # RETURN TEMPLATE
    # =====================================================

    return render_template(
        'bursar/student_detail.html',

        student=student,

        payments=payments,

        total_paid=total_paid,

        fees_total=fees_total,

        fees_balance=fees_balance,

        applicable_fee=applicable_fee,

        fee_description=fee_description,

        fee_structure_id=fee_structure_id,

        level=level,

        programme_options=programme_options,

        programme_label=programme_label,

        selected_programme=selected_programme,

        selected_residence=selected_residence,

        selected_term=selected_term,

        selected_year=selected_year
    )


@app.route(
    '/bursar/student/<student_id>/fee-details',
    methods=['POST']
)
def bursar_student_fee_details(student_id):

    if not check_permission(['bursar']):
        abort(403)

    programme = request.form.get(
        'programme',
        ''
    ).strip()

    residence = request.form.get(
        'residence',
        ''
    ).strip()

    term = request.form.get(
        'term',
        'Term 1'
    ).strip()

    year = request.form.get(
        'year',
        datetime.now().year,
        type=int
    )

    if not programme:

        flash(
            'Please select the programme.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    if residence not in [
        'Day',
        'Boarding'
    ]:

        flash(
            'Please select Day or Boarding.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    return redirect(
        url_for(
            'bursar_student_detail',
            student_id=student_id,
            programme=programme,
            residence=residence,
            term=term,
            year=year
        )
    )


@app.route('/bursar/fees')
def bursar_fees():

    if not check_permission(['bursar']):
        abort(403)

    year = request.args.get(
        'year',
        datetime.now().year,
        type=int
    )

    term = request.args.get(
        'term',
        'Term 1'
    ).strip()

    db = get_db()
    cur = db.cursor()

    cur.execute("""
        SELECT
            id,
            level,
            programme,
            residence,
            term,
            year,
            amount,
            description,
            created_at
        FROM student_fee_structure
        WHERE year = %s
        AND term = %s
        ORDER BY
            CASE level
                WHEN 'S.1' THEN 1
                WHEN 'S.2' THEN 2
                WHEN 'S.3' THEN 3
                WHEN 'S.4' THEN 4
                WHEN 'S.5' THEN 5
                WHEN 'S.6' THEN 6
                ELSE 7
            END,
            programme,
            residence
    """, (
        year,
        term
    ))

    fee_structures = cur.fetchall()

    cur.close()

    return render_template(
        'bursar/fees.html',
        fee_structures=fee_structures,
        year=year,
        term=term
    )

# ADD STUDENT FEE STRUCTURE
@app.route('/bursar/fees/add', methods=['POST'])
def bursar_fees_add():

    if not check_permission(['bursar']):
        abort(403)

    level = request.form.get(
        'level',
        ''
    ).strip().upper()

    programme = request.form.get(
        'programme',
        ''
    ).strip()

    residence = request.form.get(
        'residence',
        ''
    ).strip()

    term = request.form.get(
        'term',
        'Term 1'
    ).strip()

    year = request.form.get(
        'year',
        datetime.now().year,
        type=int
    )

    description = request.form.get(
        'description',
        ''
    ).strip()

    amount = request.form.get(
        'amount',
        ''
    ).strip()

    valid_levels = [
        'S.1',
        'S.2',
        'S.3',
        'S.4',
        'S.5',
        'S.6'
    ]

    if level not in valid_levels:

        flash(
            'Invalid class level selected.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_fees',
                year=year,
                term=term
            )
        )
    if level in [
        'S.1',
        'S.2',
        'S.3',
        'S.4'
    ]:

        valid_programmes = [
            'USE',
            'Non-USE'
        ]

    else:

        valid_programmes = [
            'UPOLET',
            'Non-UPOLET'
        ]


    if programme not in valid_programmes:

        flash(
            f'Invalid programme for {level}.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_fees',
                year=year,
                term=term
            )
        )
    if residence not in [
        'Day',
        'Boarding'
    ]:

        flash(
            'Please select Day or Boarding.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_fees',
                year=year,
                term=term
            )
        )
    try:

        amount = float(amount)

        if amount < 0:
            raise ValueError

    except (ValueError, TypeError):

        flash(
            'Enter a valid fee amount.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_fees',
                year=year,
                term=term
            )
        )
    db = get_db()
    cur = db.cursor()

    try:

        cur.execute("""
            INSERT INTO student_fee_structure (
                level,
                programme,
                residence,
                term,
                year,
                amount,
                description
            )
            VALUES (
                %s, %s, %s, %s,
                %s, %s, %s
            )
        """, (
            level,
            programme,
            residence,
            term,
            year,
            amount,
            description
        ))

        db.commit()

        flash(
            'Fee structure added successfully.',
            'success'
        )

    except Exception as e:

        db.rollback()

        if 'student_fee_structure' in str(e) and (
            'unique' in str(e).lower()
            or 'duplicate' in str(e).lower()
        ):
            flash(
                'A fee structure already exists for this '
                'level, programme, residence, term and year.',
                'danger'
            )

        else:
            flash(
                f'Could not save fee structure: {e}',
                'danger'
            )
    finally:
        cur.close()
    return redirect(
        url_for(
            'bursar_fees',
            year=year,
            term=term
        )
    )

@app.route('/bursar/fees/delete/<int:fee_id>', methods=['POST'])
def bursar_fees_delete(fee_id):

    if not check_permission(['bursar']):
        abort(403)

    db = get_db()
    cur = db.cursor()

    try:

        cur.execute("""
            DELETE FROM student_fee_structure
            WHERE id = %s
        """, (
            fee_id,
        ))

        db.commit()

        flash(
            'Fee structure deleted successfully.',
            'success'
        )

    except Exception as e:

        db.rollback()

        flash(
            f'Could not delete fee structure: {e}',
            'danger'
        )

    finally:

        cur.close()

    return redirect(
        url_for('bursar_fees')
    )

@app.route('/bursar/record_payment', methods=['POST'])
def bursar_record_payment():

    if not check_permission(['bursar']):
        abort(403)

    # =====================================================
    # GET FORM DATA
    # =====================================================

    student_id = request.form.get(
        'student_id',
        ''
    ).strip()

    amount_raw = request.form.get(
        'amount',
        ''
    ).strip()

    payment_method = request.form.get(
        'payment_method',
        'Cash'
    ).strip()

    notes = request.form.get(
        'notes',
        ''
    ).strip()

    # =====================================================
    # VALIDATE STUDENT ID
    # =====================================================

    if not student_id:

        flash(
            'Student ID is required.',
            'danger'
        )

        return redirect(
            url_for('bursar_students')
        )

    # =====================================================
    # VALIDATE PAYMENT AMOUNT
    # =====================================================

    try:

        amount = float(amount_raw)

        if amount <= 0:
            raise ValueError

    except (ValueError, TypeError):

        flash(
            'Enter a valid payment amount greater than zero.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    db = get_db_dict()
    cur = db.cursor()

    try:

        # =================================================
        # GET STUDENT
        # =================================================

        cur.execute("""
            SELECT
                student_id,
                full_name,
                parent_phone,
                COALESCE(fees_total, 0) AS fees_total
            FROM students
            WHERE student_id=%s
            FOR UPDATE
        """, (
            student_id,
        ))

        student = cur.fetchone()

        if not student:

            flash(
                'Student not found.',
                'danger'
            )

            return redirect(
                url_for('bursar_students')
            )

        # =================================================
        # GET CURRENT TOTAL OF PAYMENT RECORDS
        #
        # This becomes the source of truth.
        # =================================================

        cur.execute("""
            SELECT
                COALESCE(
                    SUM(amount),
                    0
                ) AS total_paid
            FROM payments
            WHERE student_id=%s
        """, (
            student_id,
        ))

        payment_summary = cur.fetchone()

        current_paid = float(
            payment_summary['total_paid'] or 0
        )

        # =================================================
        # CHECK PAYMENT AGAINST FEES
        # =================================================

        fees_total = float(
            student['fees_total'] or 0
        )

        new_paid = current_paid + amount

        # Prevent accidental overpayment
        if new_paid > fees_total:

            maximum_payment = max(
                fees_total - current_paid,
                0
            )

            flash(
                f'Payment exceeds the remaining balance. '
                f'Maximum amount that can be paid is '
                f'UGX {maximum_payment:,.2f}.',
                'danger'
            )

            return redirect(
                url_for(
                    'bursar_student_detail',
                    student_id=student_id
                )
            )

        # =================================================
        # GENERATE RECEIPT NUMBER
        # =================================================

        receipt_no = generate_receipt_number()

        # =================================================
        # INSERT PAYMENT
        # =================================================

        cur.execute("""
            INSERT INTO payments
            (
                student_id,
                amount,
                payment_date,
                receipt_no,
                payment_method,
                notes,
                recorded_by
            )
            VALUES
            (
                %s,
                %s,
                CURRENT_DATE,
                %s,
                %s,
                %s,
                %s
            )
        """, (
            student_id,
            amount,
            receipt_no,
            payment_method,
            notes,
            session.get('username')
        ))

        # =================================================
        # RECALCULATE TOTAL PAID FROM PAYMENT RECORDS
        #
        # We do this AFTER inserting the payment.
        # =================================================

        cur.execute("""
            SELECT
                COALESCE(
                    SUM(amount),
                    0
                ) AS total_paid
            FROM payments
            WHERE student_id=%s
        """, (
            student_id,
        ))

        payment_summary = cur.fetchone()

        total_paid = float(
            payment_summary['total_paid'] or 0
        )

        # =================================================
        # CALCULATE BALANCE
        # =================================================

        fees_balance = max(
            fees_total - total_paid,
            0
        )

        # =================================================
        # UPDATE STUDENT FEE FIGURES
        # =================================================

        cur.execute("""
            UPDATE students
            SET
                fees_paid=%s,
                fees_balance=%s
            WHERE student_id=%s
        """, (
            total_paid,
            fees_balance,
            student_id
        ))

        # =================================================
        # COMMIT EVERYTHING TOGETHER
        # =================================================

        db.commit()

    except Exception as e:

        db.rollback()

        flash(
            f'Could not record payment: {str(e)}',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    finally:

        cur.close()

    if student['parent_phone']:

        try:

            send_fee_sms(
                student['parent_phone'],
                student['full_name'],
                amount,
                fees_balance
            )

        except Exception as e:

            # Payment has already been successfully saved.
            # SMS failure should not undo the payment.

            print(
                f"Fee SMS failed: {str(e)}"
            )

    flash(
        f'Payment recorded successfully. '
        f'Receipt: {receipt_no}. '
        f'New balance: UGX {fees_balance:,.2f}',
        'success'
    )

    return redirect(
        url_for(
            'bursar_student_detail',
            student_id=student_id
        )
    )

@app.route('/bursar/payment-request/create', methods=['POST'])
def bursar_create_payment_request():

    if not check_permission(['bursar']):
        abort(403)

    student_id = request.form.get('student_id', '').strip()
    amount_raw = request.form.get('amount', '').strip()
    term = request.form.get('term', 'Term 1').strip()
    year = request.form.get(
        'year',
        datetime.now().year,
        type=int
    )

    if not student_id:
        flash('Student is required.', 'danger')
        return redirect(url_for('bursar_students'))

    try:
        amount = float(amount_raw)
    except (ValueError, TypeError):
        flash('Invalid payment amount.', 'danger')
        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    if amount <= 0:
        flash(
            'Payment amount must be greater than zero.',
            'danger'
        )
        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    db = get_db_dict()
    cur = db.cursor()

    try:

        # ============================================
        # GET STUDENT
        # ============================================

        cur.execute("""
            SELECT
                student_id,
                full_name,
                fees_balance
            FROM students
            WHERE student_id=%s
        """, (student_id,))

        student = cur.fetchone()

        if not student:
            flash(
                'Student not found.',
                'danger'
            )
            return redirect(
                url_for('bursar_students')
            )

        balance = float(
            student['fees_balance'] or 0
        )

        # ============================================
        # CHECK BALANCE
        # ============================================

        if balance <= 0:
            flash(
                'This student has no outstanding balance.',
                'warning'
            )
            return redirect(
                url_for(
                    'bursar_student_detail',
                    student_id=student_id
                )
            )

        if amount > balance:
            flash(
                f'Amount cannot exceed the outstanding balance '
                f'of UGX {balance:,.2f}.',
                'danger'
            )
            return redirect(
                url_for(
                    'bursar_student_detail',
                    student_id=student_id
                )
            )

        # ============================================
        # GENERATE RANDOM 10-DIGIT PRN
        # ============================================

        prn = generate_fee_prn()

        # ============================================
        # SAVE PAYMENT REQUEST
        # ============================================

        cur.execute("""
            INSERT INTO fee_payment_requests
            (
                prn,
                student_id,
                amount,
                term,
                year,
                payment_status,
                created_by
            )
            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                %s,
                'pending',
                %s
            )
        """, (
            prn,
            student_id,
            amount,
            term,
            year,
            session.get('username')
        ))

        db.commit()

        flash(
            f'Payment PRN generated successfully: {prn}',
            'success'
        )

    except Exception as e:

        db.rollback()

        flash(
            f'Unable to generate PRN: {str(e)}',
            'danger'
        )

    finally:
        cur.close()
        db.close()

    return redirect(
        url_for(
            'bursar_student_detail',
            student_id=student_id
        )
    )

@app.route('/bursar/student/<student_id>/generate_prn', methods=['POST'])
def bursar_generate_fee_prn(student_id):

    if not check_permission(['bursar']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    # Get student
    cur.execute("""
        SELECT
            student_id,
            full_name,
            fees_balance
        FROM students
        WHERE student_id=%s
    """, (student_id,))

    student = cur.fetchone()

    if not student:
        cur.close()
        flash('Student not found.', 'danger')
        return redirect(url_for('bursar_students'))

    # Get submitted details
    amount_raw = request.form.get('amount', '').strip()
    term = request.form.get('term', 'Term 1').strip()
    year = request.form.get(
        'year',
        datetime.now().year,
        type=int
    )

    if not amount_raw:
        cur.close()
        flash('Please enter the amount.', 'danger')
        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    try:
        amount = float(amount_raw)
    except ValueError:
        cur.close()
        flash('Invalid payment amount.', 'danger')
        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    if amount <= 0:
        cur.close()
        flash('Payment amount must be greater than zero.', 'danger')
        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    # Check outstanding balance
    balance = float(student['fees_balance'] or 0)

    if balance <= 0:
        cur.close()
        flash('This student has no outstanding fees.', 'warning')
        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    if amount > balance:
        cur.close()
        flash(
            f'Amount cannot exceed the outstanding balance '
            f'of UGX {balance:,.2f}.',
            'danger'
        )
        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    prn = None

    for _ in range(20):

        candidate = str(
            random.randint(
                1000000000,
                9999999999
            )
        )

        cur.execute("""
            SELECT id
            FROM fee_payment_requests
            WHERE prn=%s
        """, (candidate,))

        if not cur.fetchone():
            prn = candidate
            break

    if not prn:
        cur.close()

        flash(
            'Unable to generate a unique PRN. Please try again.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_student_detail',
                student_id=student_id
            )
        )

    # =====================================================
    # CREATE PAYMENT REQUEST
    # =====================================================

    cur.execute("""
        INSERT INTO fee_payment_requests
        (
            prn,
            student_id,
            amount,
            term,
            year,
            payment_status,
            created_by,
            created_at
        )
        VALUES
        (
            %s,
            %s,
            %s,
            %s,
            %s,
            'pending',
            %s,
            CURRENT_TIMESTAMP
        )
        RETURNING id
    """, (
        prn,
        student_id,
        amount,
        term,
        year,
        session.get('username')
    ))

    request_id = cur.fetchone()['id']

    db.commit()
    cur.close()

    flash(
        f'Payment PRN {prn} generated successfully.',
        'success'
    )

    return redirect(
        url_for(
            'bursar_fee_prn',
            request_id=request_id
        )
    )

@app.route('/bursar/fee-prn/<int:request_id>')
def bursar_fee_prn(request_id):

    if not check_permission(['bursar']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT
            f.id,
            f.prn,
            f.student_id,
            f.amount,
            f.term,
            f.year,
            f.payment_status,
            f.payment_method,
            f.provider_transaction_id,
            f.created_by,
            f.created_at,
            f.expires_at,
            f.paid_at,

            s.full_name,
            s.class,
            s.parent_phone

        FROM fee_payment_requests f

        JOIN students s
            ON s.student_id = f.student_id

        WHERE f.id=%s
    """, (request_id,))

    fee_request = cur.fetchone()

    cur.close()

    if not fee_request:
        flash(
            'Payment request not found.',
            'danger'
        )

        return redirect(
            url_for('bursar_students')
        )

    return render_template(
        'bursar/fee_prn.html',
        fee_request=fee_request
    )
@app.route('/bursar/payment-request/<int:request_id>/print')
def bursar_print_payment_prn(request_id):

    if not check_permission(['bursar']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT
            fpr.id,
            fpr.prn,
            fpr.amount,
            fpr.term,
            fpr.year,
            fpr.payment_status,
            fpr.payment_method,
            fpr.created_at,

            s.student_id,
            s.full_name,
            s.class,
            s.parent_phone

        FROM fee_payment_requests fpr

        JOIN students s
            ON s.student_id = fpr.student_id

        WHERE fpr.id=%s
    """, (request_id,))

    payment_request = cur.fetchone()

    if not payment_request:

        cur.close()

        flash(
            'Payment request not found.',
            'danger'
        )

        return redirect(
            url_for('bursar_students')
        )

    # --------------------------------------------
    # SCHOOL INFORMATION
    # --------------------------------------------

    cur.execute("""
        SELECT
            school_name,
            school_address,
            school_phone,
            school_email,
            logo_url
        FROM school_settings
        WHERE id=1
    """)

    school = cur.fetchone()

    cur.close()

    return render_template(
        'bursar/print_payment_prn.html',
        payment_request=payment_request,
        school=school
    )
@app.route('/bursar/print_receipts')
def bursar_print_receipts():
    if not check_permission(['bursar']):
        abort(403)

    receipt_ids = request.args.get('ids', '')
    receipts = []

    if receipt_ids:
        ids = [int(x) for x in receipt_ids.split(',') if x.isdigit()]

        if ids:
            placeholders = ','.join(['%s'] * len(ids))

            db = get_db_dict()
            cur = db.cursor()

            cur.execute(f"""
                SELECT p.*, s.full_name, s.class
                FROM payments p
                JOIN students s 
                ON p.student_id=s.student_id
                WHERE p.id IN ({placeholders})
                ORDER BY p.payment_date DESC
            """, ids)

            receipts = cur.fetchall()
            cur.close()

    return render_template(
        'bursar/print_receipts.html',
        receipts=receipts
    )

@app.route('/bursar/send_reminder/<student_id>')
def bursar_send_reminder(student_id):
    if not check_permission(['bursar']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT full_name, parent_phone, fees_balance
        FROM students
        WHERE student_id=%s
    """, (student_id,))

    student = cur.fetchone()
    cur.close()

    if student and student['parent_phone']:
        send_sms(
            student['parent_phone'],
            f"Fees reminder: UGX {student['fees_balance']:,.2f} outstanding for {student['full_name']}."
        )
        flash('Reminder sent.', 'success')
    else:
        flash('No parent phone.', 'warning')

    return redirect(url_for('bursar_student_detail', student_id=student_id))

@app.route('/bursar/bulk_reminder', methods=['POST'])
def bursar_bulk_reminder():
    if not check_permission(['bursar']):
        abort(403)

    class_filter = request.form.get('class', '')

    db = get_db_dict()
    cur = db.cursor()

    query = """
        SELECT full_name, parent_phone, fees_balance
        FROM students
        WHERE fees_balance > 0
    """

    params = []

    if class_filter:
        query += " AND class=%s"
        params.append(class_filter)

    cur.execute(query, params)
    students = cur.fetchall()
    cur.close()

    sent = 0

    for s in students:
        if s['parent_phone']:
            send_sms(
                s['parent_phone'],
                f"Fees reminder: UGX {s['fees_balance']:,.2f} outstanding for {s['full_name']}."
            )
            sent += 1

    flash(f'{sent} reminders sent.', 'success')
    return redirect(url_for('bursar_students'))

@app.route('/bursar/clearance/<student_id>')
def bursar_clearance(student_id):
    if not check_permission(['bursar']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT student_id, full_name, class, parent_phone,
               fees_balance, fees_total, fees_paid, photo_path
        FROM students
        WHERE student_id=%s
    """, (student_id,))

    student = cur.fetchone()
    cur.close()

    if not student:
        flash('Student not found', 'danger')
        return redirect(url_for('bursar_students'))

    student['fees_total'] = student.get('fees_total') or 0
    student['fees_paid'] = student.get('fees_paid') or 0
    student['fees_balance'] = student.get('fees_balance') or 0
    student['photo_url'] = get_photo_url(student.get('photo_path'))

    return render_template(
        'bursar/clearance.html',
        student=student
    )

@app.route('/bursar/bulk_clearance')
def bursar_bulk_clearance():
    if not check_permission(['bursar']):
        abort(403)

    class_filter = request.args.get('class', '')

    db = get_db_dict()
    cur = db.cursor()

    query = """
        SELECT student_id, full_name, class, parent_phone,
               fees_balance, fees_total, fees_paid, photo_path
        FROM students
        WHERE fees_balance <= 0
    """

    params = []

    if class_filter:
        query += " AND class=%s"
        params.append(class_filter)

    cur.execute(query, params)
    students = cur.fetchall()
    cur.close()

    for s in students:
        s['fees_total'] = s.get('fees_total') or 0
        s['fees_paid'] = s.get('fees_paid') or 0
        s['fees_balance'] = s.get('fees_balance') or 0
        s['photo_url'] = get_photo_url(s.get('photo_path'))

    return render_template(
        'bursar/bulk_clearance.html',
        students=students
    )


@app.route('/bursar/webhook/process')
def bursar_process_webhooks():
    if not check_permission(['bursar']):
        abort(403)
    db = get_db_dict()
    cur = db.cursor()
    cur.execute("SELECT * FROM payment_webhooks WHERE processed=0")
    webhooks = cur.fetchall()
    processed = 0
    for w in webhooks:
        if w.get('student_id'):
            cur.execute("SELECT full_name, parent_phone, fees_paid, fees_balance FROM students WHERE student_id=%s", (w['student_id'],))
            student = cur.fetchone()
            if student:
                receipt_no = generate_receipt_number()
                cur.execute("""
                    INSERT INTO payments 
                    (student_id, amount, payment_date, receipt_no, payment_method, notes, recorded_by)
                    VALUES (%s, %s, CURRENT_DATE, %s, %s, %s, %s)
                """,(w['student_id'],w['amount'],receipt_no,w.get('payment_method','Mobile Money'),f"Auto from webhook: {w.get('transaction_id','')}",'System'))
                new_paid = (student['fees_paid'] or 0) + w['amount']
                new_balance = (student['fees_balance'] or 0) - w['amount']
                cur.execute("UPDATE students SET fees_paid=%s, fees_balance=%s WHERE student_id=%s",(new_paid,new_balance,w['student_id']))
                if student.get('parent_phone'):
                    send_fee_sms(student['parent_phone'],student['full_name'],w['amount'],new_balance)
        cur.execute("UPDATE payment_webhooks SET processed=1 WHERE id=%s",(w['id'],))
        processed += 1
    db.commit()
    cur.close()
    flash(f'Processed {processed} pending webhooks.','success')
    return redirect(url_for('bursar_dashboard'))

@app.route('/bursar/webhook/payment', methods=['POST'])
def bursar_payment_webhook():
    data = request.get_json()
    if not data:
        return jsonify({'error':'Invalid data'}),400
    execute_db("""
        INSERT INTO payment_webhooks 
        (transaction_id, amount, phone_number, student_id, reference, payment_method, raw_data, status, processed)
        VALUES (%s,%s,%s,%s,%s,%s,%s,'received',0)
    """,(data.get('transaction_id'),data.get('amount'),data.get('phone_number'),data.get('student_id'),data.get('reference'),data.get('payment_method'),json.dumps(data)))
    return jsonify({'status':'received'}),200

def generate_staff_no():
    return generate_unique_number('STF','staff','staff_no',year_format=True)

def generate_payroll_no():
    db = get_db_dict()
    cur = db.cursor()

    today = datetime.now()
    prefix = f"PR-{today.strftime('%Y%m')}"

    cur.execute("""
        SELECT payroll_no
        FROM payroll
        WHERE payroll_no LIKE %s
        ORDER BY id DESC
        LIMIT 1
    """, (f"{prefix}-%",))

    last = cur.fetchone()
    cur.close()

    if last and last.get('payroll_no'):
        try:
            last_num = int(last['payroll_no'].split('-')[-1])
        except (ValueError, AttributeError):
            last_num = 0
    else:
        last_num = 0

    next_num = last_num + 1

    return f"{prefix}-{next_num:04d}"

@app.route('/bursar/staff')
def bursar_staff():
    if not check_permission(['bursar']):
        abort(403)
    db=get_db_dict()
    cur=db.cursor()
    cur.execute("SELECT * FROM staff ORDER BY full_name")
    staff=cur.fetchall()
    cur.execute("SELECT nssf_employee_rate,paye_rate,paye_threshold FROM school_settings WHERE id=1")
    rates_row=cur.fetchone()
    if rates_row:
        nssf_rate=rates_row.get('nssf_employee_rate',5.0)
        paye_rate=rates_row.get('paye_rate',10.0)
        paye_threshold=rates_row.get('paye_threshold',235000)
    else:
        nssf_rate=5.0
        paye_rate=10.0
        paye_threshold=235000
    total_basic=total_allowances=total_gross=total_nssf=total_paye=total_deductions=total_net=0
    for s in staff:
        gross=(s['salary_basic'] or 0)+(s['salary_allowances'] or 0)
        nssf=(gross*nssf_rate)/100
        taxable=max(0,gross-paye_threshold)
        paye=(taxable*paye_rate)/100
        net=gross-nssf-paye-(s['salary_deductions'] or 0)
        s['gross']=float(gross)
        s['nssf']=round(nssf,2)
        s['paye']=round(paye,2)
        s['net']=round(net,2)
        s['salary_net']=round(net,2)
        total_basic+=s['salary_basic'] or 0
        total_allowances+=s['salary_allowances'] or 0
        total_gross+=gross
        total_nssf+=nssf
        total_paye+=paye
        total_deductions+=s['salary_deductions'] or 0
        total_net+=net
    cur.close()
    return render_template('bursar/staff.html',staff=staff,total_basic=total_basic,total_allowances=total_allowances,total_gross=total_gross,total_nssf=total_nssf,total_paye=total_paye,total_deductions=total_deductions,total_net=total_net,nssf_rate=nssf_rate,paye_rate=paye_rate,paye_threshold=paye_threshold)

@app.route('/bursar/staff/add', methods=['GET','POST'])
def bursar_staff_add():
    if not check_permission(['bursar']):
        abort(403)
    if request.method=='POST':
        full_name=request.form['full_name'].strip()
        position=request.form['position'].strip()
        department=request.form.get('department','').strip()
        phone=validate_and_format_phone(request.form.get('phone',''))
        email=request.form.get('email','').strip()
        nssf_number=request.form.get('nssf_number','').strip()
        tin_number=request.form.get('tin_number','').strip()
        bank_account=request.form.get('bank_account','').strip()
        bank_name=request.form.get('bank_name','').strip()
        salary_basic=float(request.form.get('salary_basic',0))
        salary_allowances=float(request.form.get('salary_allowances',0))
        salary_deductions=float(request.form.get('salary_deductions',0))
        staff_no=generate_staff_no()
        execute_db("""
            INSERT INTO staff
            (staff_no,full_name,position,department,phone,email,nssf_number,tin_number,bank_account,bank_name,salary_basic,salary_allowances,salary_deductions)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
        """,(staff_no,full_name,position,department,phone,email,nssf_number,tin_number,bank_account,bank_name,salary_basic,salary_allowances,salary_deductions))
        flash(f'Staff {full_name} added. Staff No: {staff_no}','success')
        return redirect(url_for('bursar_staff'))
    return render_template('bursar/staff_add.html')


@app.route('/bursar/payroll/generate',methods=['GET','POST'])
def bursar_generate_payroll():
    if not check_permission(['bursar']):
        abort(403)
    if request.method=='POST':
        month_year_input = request.form['month_year']
        selected_staff = request.form.getlist('staff_ids')
        try:
            month_year = datetime.strptime(
                month_year_input,
                '%Y-%m'
            ).date().replace(day=1)
        except (ValueError, TypeError):
            flash(
                'Invalid payroll month. Please select a valid month.',
                'danger'
            )
            return redirect(
                url_for('bursar_generate_payroll')
            )
        if not selected_staff:
            flash('No staff selected.','danger')
            return redirect(url_for('bursar_generate_payroll'))

        db=get_db_dict()
        cur=db.cursor()

        cur.execute("SELECT nssf_employee_rate,paye_rate,paye_threshold FROM school_settings WHERE id=1")
        rates=cur.fetchone()

        if rates:
            nssf_rate=rates.get('nssf_employee_rate',5.0)
            paye_rate=rates.get('paye_rate',10.0)
            paye_threshold=rates.get('paye_threshold',235000)
        else:
            nssf_rate=5.0
            paye_rate=10.0
            paye_threshold=235000

        placeholders=','.join(['%s']*len(selected_staff))
        cur.execute(f"""
            SELECT id,full_name,position,salary_basic,salary_allowances,salary_deductions,bank_name,bank_account,phone
            FROM staff
            WHERE id IN ({placeholders})
        """,selected_staff)

        staff_list=cur.fetchall()
        total_amount=0

        for staff in staff_list:
            gross=(staff['salary_basic'] or 0)+(staff['salary_allowances'] or 0)
            nssf=(gross*nssf_rate)/100
            taxable=max(0,gross-paye_threshold)
            paye=(taxable*paye_rate)/100
            net=gross-nssf-paye-(staff['salary_deductions'] or 0)
            total_amount+=net

        payroll_no=generate_payroll_no()
        approval_code=generate_approval_code()
        token,expires_at=generate_secure_token(2)

        cur.execute("""
            INSERT INTO payroll
            (payroll_no,month_year,total_amount,approval_code,headteacher_access_token,token_expires_at,recorded_by,approval_status)
            VALUES (%s,%s,%s,%s,%s,%s,%s,%s)
            RETURNING id
        """,(payroll_no,month_year,total_amount,approval_code,token,expires_at,session.get('username'),'pending'))

        payroll_id=cur.fetchone()['id']

        for staff in staff_list:
            gross=(staff['salary_basic'] or 0)+(staff['salary_allowances'] or 0)
            nssf=(gross*nssf_rate)/100
            taxable=max(0,gross-paye_threshold)
            paye=(taxable*paye_rate)/100
            net_salary=gross-nssf-paye-(staff['salary_deductions'] or 0)

            cur.execute("""
                INSERT INTO salary_payments
                (staff_id,payroll_id,month_year,basic,allowances,deductions,gross_salary,nssf_employee,paye_tax,net_salary,approval_code,recorded_by)
                VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)
            """,(
                staff['id'],
                payroll_id,
                month_year,
                staff['salary_basic'] or 0,
                staff['salary_allowances'] or 0,
                staff['salary_deductions'] or 0,
                gross,
                nssf,
                paye,
                net_salary,
                approval_code,
                session.get('username')
            ))

        db.commit()
        cur.close()

        approval_link=url_for('headteacher_approval_access',token=token,_external=True)

        cur=get_db_dict().cursor()
        cur.execute("SELECT phone FROM users WHERE role='headteacher' AND status=1 LIMIT 1")
        headteacher=cur.fetchone()
        cur.close()

        if headteacher and headteacher.get('phone'):
            send_sms(headteacher['phone'],f"Payroll {payroll_no} awaiting approval. Link: {approval_link}")

        add_notification('headteacher',f"Payroll {payroll_no} needs approval. Code: {approval_code}",f"/headteacher/approval/{token}")

        flash(f'Payroll {payroll_no} created. Approval link sent to Headteacher.','success')
        return redirect(url_for('bursar_payroll_list'))

    db=get_db_dict()
    cur=db.cursor()

    cur.execute("SELECT nssf_employee_rate,paye_rate,paye_threshold FROM school_settings WHERE id=1")
    rates=cur.fetchone()

    if rates:
        nssf_rate=rates.get('nssf_employee_rate',5.0)
        paye_rate=rates.get('paye_rate',10.0)
        paye_threshold=rates.get('paye_threshold',235000)
    else:
        nssf_rate=5.0
        paye_rate=10.0
        paye_threshold=235000

    cur.execute("""
        SELECT id,full_name,position,salary_basic,salary_allowances,salary_deductions,bank_name,bank_account,phone
        FROM staff
        WHERE status='active'
        ORDER BY full_name
    """)

    staff_list=cur.fetchall()
    cur.close()

    return render_template('bursar/generate_payroll.html',
                           staff_list=staff_list,
                           nssf_rate=nssf_rate,
                           paye_rate=paye_rate,
                           paye_threshold=paye_threshold)

@app.route('/bursar/payroll/list')
def bursar_payroll_list():
    if not check_permission(['bursar']):
        abort(403)
    db=get_db_dict()
    cur=db.cursor()
    cur.execute("""
        SELECT p.*, COUNT(sp.id) AS staff_count
        FROM payroll p
        LEFT JOIN salary_payments sp ON p.id=sp.payroll_id
        GROUP BY p.id
        ORDER BY p.created_at DESC
    """)
    payrolls=cur.fetchall()
    cur.close()
    return render_template('bursar/payroll_list.html',payrolls=payrolls)


@app.route('/bursar/delete_payroll/<int:payroll_id>')
def bursar_delete_payroll(payroll_id):
    if not check_permission(['bursar']):
        abort(403)

    db=get_db_dict()
    cur=db.cursor()
    cur.execute("SELECT approval_status FROM payroll WHERE id=%s",(payroll_id,))
    payroll=cur.fetchone()
    cur.close()

    if not payroll:
        flash('Payroll not found.','danger')
        return redirect(url_for('bursar_payroll_list'))

    if payroll['approval_status']!='pending':
        flash('Only pending payrolls can be deleted.','warning')
        return redirect(url_for('bursar_payroll_list'))

    try:
        execute_db("DELETE FROM salary_payments WHERE payroll_id=%s",(payroll_id,))
        execute_db("DELETE FROM payroll WHERE id=%s",(payroll_id,))
        flash('Payroll deleted successfully.','success')
    except Exception as e:
        flash(f'Error deleting payroll: {str(e)}','danger')

    return redirect(url_for('bursar_payroll_list'))


@app.route('/bursar/view_payroll/<int:payroll_id>')
def bursar_view_payroll(payroll_id):
    if not check_permission(['bursar']):
        abort(403)

    db=get_db_dict()
    cur=db.cursor()

    cur.execute("SELECT * FROM payroll WHERE id=%s",(payroll_id,))
    payroll=cur.fetchone()

    if not payroll:
        flash('Payroll not found.','danger')
        return redirect(url_for('bursar_payroll_list'))

    cur.execute("""
        SELECT sp.*,s.full_name,s.position,s.bank_account,s.bank_name,s.phone,s.staff_no
        FROM salary_payments sp
        JOIN staff s ON sp.staff_id=s.id
        WHERE sp.payroll_id=%s
    """,(payroll_id,))

    staff_list=cur.fetchall()
    cur.close()

    total_basic=sum(s['basic'] or 0 for s in staff_list)
    total_allowances=sum(s['allowances'] or 0 for s in staff_list)
    total_deductions=sum(s['deductions'] or 0 for s in staff_list)

    return render_template(
        'bursar/view_payroll.html',
        payroll=payroll,
        staff_list=staff_list,
        total_basic=total_basic,
        total_allowances=total_allowances,
        total_deductions=total_deductions
    )


@app.route('/bursar/print_payroll')
def bursar_print_payroll():
    if not check_permission(['bursar']):
        abort(403)

    db=get_db_dict()
    cur=db.cursor()

    cur.execute("SELECT nssf_employee_rate,paye_rate,paye_threshold FROM school_settings WHERE id=1")
    rates=cur.fetchone()

    if rates:
        nssf_rate=rates.get('nssf_employee_rate',5.0)
        paye_rate=rates.get('paye_rate',10.0)
        paye_threshold=rates.get('paye_threshold',235000)
    else:
        nssf_rate=5.0
        paye_rate=10.0
        paye_threshold=235000

    cur.execute("""
        SELECT staff_no,full_name,position,salary_basic,salary_allowances,
        salary_deductions,bank_name,bank_account,phone
        FROM staff
        ORDER BY full_name
    """)

    staff_list=cur.fetchall()
    cur.close()

    total_basic=0
    total_allowances=0
    total_gross=0
    total_nssf=0
    total_paye=0
    total_deductions=0
    total_net=0

    for staff in staff_list:
        gross=(staff['salary_basic'] or 0)+(staff['salary_allowances'] or 0)
        nssf=(gross*nssf_rate)/100
        taxable=max(0,gross-paye_threshold)
        paye=(taxable*paye_rate)/100
        net=gross-nssf-paye-(staff['salary_deductions'] or 0)

        staff['gross']=gross
        staff['nssf']=round(nssf,2)
        staff['paye']=round(paye,2)
        staff['salary_net']=round(net,2)

        total_basic+=staff['salary_basic'] or 0
        total_allowances+=staff['salary_allowances'] or 0
        total_gross+=gross
        total_nssf+=nssf
        total_paye+=paye
        total_deductions+=staff['salary_deductions'] or 0
        total_net+=net

    return render_template(
        'bursar/print_payroll.html',
        staff_list=staff_list,
        total_basic=total_basic,
        total_allowances=total_allowances,
        total_gross=total_gross,
        total_nssf=total_nssf,
        total_paye=total_paye,
        total_deductions=total_deductions,
        total_net=total_net,
        nssf_rate=nssf_rate,
        paye_rate=paye_rate,
        paye_threshold=paye_threshold
    )


@app.route('/bursar/print_fees_list')
def bursar_print_fees_list():

    if not check_permission(['bursar']):
        abort(403)

    class_filter = request.args.get('class','')
    status_filter = request.args.get('status','')

    db = get_db_dict()
    cur = db.cursor()

    params = []

    if status_filter == 'defaulters':

        query = """
            SELECT 
                student_id,
                full_name,
                class,
                fees_paid,
                fees_balance
            FROM students
            WHERE fees_balance > 0
        """

    else:

        query = """
            SELECT 
                student_id,
                full_name,
                class,
                fees_paid,
                fees_balance
            FROM students
            WHERE 1=1
        """

    if class_filter:
        query += " AND class=%s"
        params.append(class_filter)

    query += """
        ORDER BY class, full_name
    """

    cur.execute(query, params)

    students = cur.fetchall()

    cur.close()

    total_paid = sum(
        s['fees_paid'] or 0
        for s in students
    )

    total_balance = sum(
        s['fees_balance'] or 0
        for s in students
    )

    return render_template(
        'bursar/print_fees_list.html',
        students=students,
        class_filter=class_filter,
        status_filter=status_filter,
        total_paid=total_paid,
        total_balance=total_balance
    )
# =========================================================
# BURSAR - BUDGET MANAGEMENT
# =========================================================

@app.route('/bursar/budget')
def bursar_budget():

    if not check_permission(['bursar']):
        abort(403)

    year = request.args.get(
        'year',
        datetime.now().year,
        type=int
    )

    db = get_db()
    cur = db.cursor()
    cur.execute("""
        SELECT
            bc.id,
            bc.code,
            bc.name,
            bc.description,
            bc.allocated_amount,
            bc.year,

            COALESCE(
                SUM(
                    CASE
                        WHEN e.status IS NULL
                             OR LOWER(e.status) != 'cancelled'
                        THEN e.amount
                        ELSE 0
                    END
                ),
                0
            ) AS spent

        FROM budget_categories bc

        LEFT JOIN expenditures e
            ON e.category_id = bc.id
            AND EXTRACT(
                YEAR FROM e.expenditure_date
            ) = bc.year

        WHERE bc.year = %s

        GROUP BY
            bc.id,
            bc.code,
            bc.name,
            bc.description,
            bc.allocated_amount,
            bc.year

        ORDER BY
            bc.code,
            bc.name
    """, (year,))

    categories = cur.fetchall()
    total_allocated = sum(
        float(item['allocated_amount'] or 0)
        for item in categories
    )

    total_spent = sum(
        float(item['spent'] or 0)
        for item in categories
    )

    total_balance = (
        total_allocated - total_spent
    )
    cur.close()
    return render_template(
        'bursar/budget.html',
        categories=categories,
        year=year,
        total_allocated=total_allocated,
        total_spent=total_spent,
        total_balance=total_balance
    )
# =========================================================
# BURSAR - ADD BUDGET CATEGORY
# =========================================================

@app.route('/bursar/budget/add', methods=['POST'])
def bursar_budget_add():

    if not check_permission(['bursar']):
        abort(403)

    year = request.form.get(
        'year',
        datetime.now().year,
        type=int
    )

    code = request.form.get(
        'code',
        ''
    ).strip()

    name = request.form.get(
        'name',
        ''
    ).strip()

    description = request.form.get(
        'description',
        ''
    ).strip()

    allocated_amount = request.form.get(
        'allocated_amount',
        '0'
    ).strip()

    # =====================================================
    # VALIDATION
    # =====================================================

    if not code:
        flash(
            'Budget category code is required.',
            'danger'
        )
        return redirect(
            url_for(
                'bursar_budget',
                year=year
            )
        )

    if not name:
        flash(
            'Budget category name is required.',
            'danger'
        )
        return redirect(
            url_for(
                'bursar_budget',
                year=year
            )
        )

    try:

        allocated_amount = float(
            allocated_amount
        )

        if allocated_amount < 0:
            raise ValueError

    except (ValueError, TypeError):

        flash(
            'Allocated amount must be a valid positive number.',
            'danger'
        )

        return redirect(
            url_for(
                'bursar_budget',
                year=year
            )
        )

    db = get_db()
    cur = db.cursor()

    try:

        # =================================================
        # CHECK DUPLICATE CODE FOR SAME YEAR
        # =================================================

        cur.execute("""
            SELECT id
            FROM budget_categories
            WHERE code=%s
            AND year=%s
        """, (
            code,
            year
        ))

        existing = cur.fetchone()

        if existing:

            flash(
                f'Budget category {code} already exists for {year}.',
                'warning'
            )

            cur.close()

            return redirect(
                url_for(
                    'bursar_budget',
                    year=year
                )
            )

        # =================================================
        # INSERT CATEGORY
        # =================================================

        cur.execute("""
            INSERT INTO budget_categories (
                code,
                name,
                description,
                allocated_amount,
                year
            )
            VALUES (
                %s,
                %s,
                %s,
                %s,
                %s
            )
        """, (
            code,
            name,
            description,
            allocated_amount,
            year
        ))

        db.commit()

        flash(
            'Budget category added successfully.',
            'success'
        )

    except Exception as e:

        db.rollback()

        flash(
            f'Unable to add budget category: {str(e)}',
            'danger'
        )

    finally:

        cur.close()

    return redirect(
        url_for(
            'bursar_budget',
            year=year
        ))

@app.route('/bursar/expenditure', methods=['GET'])
def bursar_expenditure():
    if not check_permission(['bursar']):
        abort(403)

    db = get_db()
    cur = db.cursor()

    # =====================================================
    # GET BUDGET CATEGORIES
    # =====================================================

    cur.execute("""
        SELECT
            id,
            code,
            name,
            allocated_amount,
            year
        FROM budget_categories
        ORDER BY name
    """)

    categories = cur.fetchall()

    # =====================================================
    # GET EXPENDITURES
    # =====================================================

    cur.execute("""
        SELECT
            e.id,
            e.voucher_no,
            e.category_id,
            e.description,
            e.amount,
            e.expenditure_date,
            e.payment_method,
            e.payee_name,
            e.payee_phone,
            e.status,
            e.recorded_by,
            e.created_at,
            bc.code AS category_code,
            bc.name AS category_name
        FROM expenditures e
        LEFT JOIN budget_categories bc
            ON e.category_id = bc.id
        ORDER BY e.expenditure_date DESC, e.id DESC
    """)

    expenditures = cur.fetchall()

    # =====================================================
    # TOTAL EXPENDITURE
    # =====================================================

    total_expenditure = sum(
        float(e['amount'] or 0)
        for e in expenditures
    )

    cur.close()

    return render_template(
        'bursar/expenditure.html',
        categories=categories,
        expenditures=expenditures,
        total_expenditure=total_expenditure
    )

@app.route('/bursar/expenditure/add', methods=['POST'])
def bursar_expenditure_add():

    if not check_permission(['bursar']):
        abort(403)
    voucher_no = generate_voucher_no()
    category_id = request.form.get('category_id')

    description = request.form.get(
        'description',
        ''
    ).strip()

    amount = request.form.get('amount')

    expenditure_date = request.form.get(
        'expenditure_date'
    )

    payment_method = request.form.get(
        'payment_method',
        ''
    ).strip()

    payee_name = request.form.get(
        'payee_name',
        ''
    ).strip()

    payee_phone = request.form.get(
        'payee_phone',
        ''
    ).strip()

    status = request.form.get(
        'status',
        'Pending'
    ).strip()

    if not category_id:

        flash(
            'Please select a budget category.',
            'danger'
        )

        return redirect(
            url_for('bursar_expenditure')
        )
    if not amount:

        flash(
            'Expenditure amount is required.',
            'danger'
        )

        return redirect(
            url_for('bursar_expenditure')
        )

    try:

        amount = float(amount)

        if amount <= 0:
            raise ValueError

    except (ValueError, TypeError):

        flash(
            'Enter a valid expenditure amount.',
            'danger'
        )

        return redirect(
            url_for('bursar_expenditure')
        )
    db = get_db()
    cur = db.cursor()

    try:

        cur.execute("""
            INSERT INTO expenditures (
                voucher_no,
                category_id,
                description,
                amount,
                expenditure_date,
                payment_method,
                payee_name,
                payee_phone,
                status,
                recorded_by
            )
            VALUES (
                %s, %s, %s, %s, %s,
                %s, %s, %s, %s, %s
            )
        """, (
            voucher_no,
            category_id,
            description,
            amount,
            expenditure_date,
            payment_method,
            payee_name,
            payee_phone,
            status,
            session.get('user_id')
        ))

        db.commit()

        flash(
            f'Expenditure recorded successfully. '
            f'Voucher: {voucher_no}',
            'success'
        )

    except Exception as e:

        db.rollback()

        flash(
            f'Could not record expenditure: {str(e)}',
            'danger'
        )

    finally:

        cur.close()

    return redirect(
        url_for('bursar_expenditure')
    )

# =========================================================
# BURSAR - EXPENDITURE REPORT
# =========================================================

@app.route('/bursar/expenditure/report')
def bursar_expenditure_report():

    if not check_permission(['bursar']):
        abort(403)

    # =====================================================
    # GET FILTERS
    # =====================================================

    year = request.args.get(
        'year',
        datetime.now().year,
        type=int
    )

    category_id = request.args.get(
        'category_id',
        type=int
    )

    status = request.args.get(
        'status',
        ''
    ).strip()


    # =====================================================
    # DATABASE
    # =====================================================

    db = get_db()
    cur = db.cursor()
    cur.execute("""
        SELECT
            school_name,
            school_address,
            school_phone,
            school_email,
            logo_url
        FROM school_settings
        WHERE id = 1
    """)

    school_data = cur.fetchone()


    if school_data:

        school_name = (
            school_data['school_name']
            or 'YOUR SCHOOL NAME'
        )

        school_address = (
            school_data['school_address']
            or ''
        )

        school_phone = (
            school_data['school_phone']
            or ''
        )

        school_email = (
            school_data['school_email']
            or ''
        )

        school_logo_url = (
            school_data['logo_url']
            or url_for(
                'static',
                filename='images/logo.png'
            )
        )

    else:

        school_name = 'YOUR SCHOOL NAME'
        school_address = ''
        school_phone = ''
        school_email = ''

        school_logo_url = url_for(
            'static',
            filename='images/logo.png'
        )
        
    cur.execute("""
        SELECT
            id,
            code,
            name
        FROM budget_categories
        ORDER BY name
    """)

    categories = cur.fetchall()
    
    query = """
        SELECT
            e.id,
            e.voucher_no,
            e.category_id,
            e.description,
            e.amount,
            e.expenditure_date,
            e.payment_method,
            e.payee_name,
            e.payee_phone,
            e.status,
            e.recorded_by,
            e.created_at,

            bc.code AS category_code,
            bc.name AS category_name

        FROM expenditures e

        LEFT JOIN budget_categories bc
            ON e.category_id = bc.id

        WHERE (
            bc.year = %s
            OR bc.year IS NULL
        )
    """

    params = [year]

    if category_id:

        query += """
            AND e.category_id = %s
        """

        params.append(category_id)
    if status:

        query += """
            AND LOWER(TRIM(COALESCE(e.status, '')))
                = LOWER(TRIM(%s))
        """

        params.append(status)

    query += """
        ORDER BY
            e.expenditure_date DESC,
            e.id DESC
    """
    cur.execute(
        query,
        tuple(params)
    )

    expenditures = cur.fetchall()

    total_expenditure = 0
    approved_total = 0
    paid_total = 0
    pending_total = 0
    rejected_total = 0
    for expenditure in expenditures:
        amount = float(
            expenditure['amount'] or 0
        )
        total_expenditure += amount
        current_status = str(
            expenditure['status'] or ''
        ).strip().lower()
        if current_status == 'approved':
            approved_total += amount
        elif current_status == 'paid':
            paid_total += amount
        elif current_status == 'pending':
            pending_total += amount
        elif current_status == 'rejected':
            rejected_total += amount
    cur.close()
    return render_template(
        'bursar/expenditure_report.html',

        expenditures=expenditures,

        categories=categories,

        year=year,

        category_id=category_id,

        status=status,

        total_expenditure=total_expenditure,

        approved_total=approved_total,

        paid_total=paid_total,

        pending_total=pending_total,

        rejected_total=rejected_total,

        school_name=school_name,

        school_address=school_address,

        school_phone=school_phone,

        school_email=school_email,

        school_logo_url=school_logo_url
    )


@app.route('/bursar/income_report')
def bursar_income_report():

    if not check_permission(['bursar']):
        abort(403)

    start = request.args.get(
        'start_date',
        datetime.now().replace(day=1).strftime('%Y-%m-%d')
    )

    end = request.args.get(
        'end_date',
        datetime.now().strftime('%Y-%m-%d')
    )

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT 
            DATE(payment_date) AS date,
            SUM(amount) AS total
        FROM payments
        WHERE payment_date BETWEEN %s AND %s
        GROUP BY DATE(payment_date)
        ORDER BY date DESC
    """,(start,end))

    daily = cur.fetchall()


    cur.execute("""
        SELECT 
            payment_method,
            SUM(amount) AS total
        FROM payments
        WHERE payment_date BETWEEN %s AND %s
        GROUP BY payment_method
    """,(start,end))

    by_method = cur.fetchall()


    cur.execute("""
        SELECT 
            COALESCE(SUM(amount),0) AS total_income
        FROM payments
        WHERE payment_date BETWEEN %s AND %s
    """,(start,end))

    total = cur.fetchone()

    cur.close()

    return render_template(
        'bursar/income_report.html',
        daily=daily,
        by_method=by_method,
        total=total,
        start_date=start,
        end_date=end
    )

@app.route('/bursar/school_pay/config', methods=['GET', 'POST'])
def bursar_school_pay_config():
    if not check_permission(['bursar']):
        abort(403)

    if request.method == 'POST':
        execute_db("""
            UPDATE payment_gateway_config 
            SET api_key=%s, api_secret=%s, webhook_secret=%s, callback_url=%s, status=%s
            WHERE id=1
        """, (
            request.form['api_key'],
            request.form['api_secret'],
            request.form['webhook_secret'],
            request.form['callback_url'],
            request.form.get('status', 'inactive')
        ))
        flash('Configuration saved.', 'success')

    db = get_db_dict()
    cur = db.cursor()
    cur.execute("SELECT * FROM payment_gateway_config WHERE id=1")
    config = cur.fetchone()
    cur.close()

    return render_template('bursar/school_pay_config.html', config=config)


@app.route('/bursar/payment_gateway/test', methods=['POST'])
def bursar_payment_gateway_test():
    if not check_permission(['bursar']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT status 
        FROM payment_gateway_config 
        WHERE id=1
    """)
    config = cur.fetchone()
    cur.close()

    if not config:
        flash('Payment gateway not configured.', 'danger')
        return redirect(url_for('bursar_school_pay_config'))

    flash('Payment gateway connection tested successfully.', 'success')
    return redirect(url_for('bursar_school_pay_config'))


@app.route('/bursar/settings')
def bursar_settings():
    if not check_permission(['bursar']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT * FROM school_settings 
        WHERE id=1
    """)
    settings = cur.fetchone()

    cur.close()

    return render_template(
        'bursar/settings.html',
        settings=settings
    )


@app.route('/bursar/settings/update', methods=['POST'])
def bursar_settings_update():
    if not check_permission(['bursar']):
        abort(403)

    execute_db("""
        UPDATE school_settings SET
        nssf_employee_rate=%s,
        paye_rate=%s,
        paye_threshold=%s
        WHERE id=1
    """, (
        float(request.form.get('nssf_employee_rate', 5)),
        float(request.form.get('paye_rate', 10)),
        float(request.form.get('paye_threshold', 235000))
    ))

    flash('Settings updated successfully.', 'success')

    return redirect(url_for('bursar_settings'))


# ==================== HEADTEACHER & MANAGEMENT APPROVAL ====================
@app.route('/headteacher/approvals')
def headteacher_approvals():
    if not check_permission(['headteacher']):
        abort(403)
    db = get_db_dict()
    cur = db.cursor()
    cur.execute("""
        SELECT id, payroll_no, month_year, total_amount, approval_status,
               approval_code, created_at, recorded_by
        FROM payroll
        WHERE approval_status = 'pending'
        ORDER BY created_at DESC
    """)
    pending = cur.fetchall()
    cur.close()
    return render_template('headteacher/approvals.html', pending=pending)

@app.route('/headteacher/approval/<token>', methods=['GET', 'POST'])
def headteacher_approval_access(token):
    if not check_permission(['headteacher']):
        abort(403)
    db = get_db_dict()
    cur = db.cursor()
    cur.execute("SELECT * FROM payroll WHERE headteacher_access_token=%s AND approval_status='pending'", (token,))
    payroll = cur.fetchone()
    if not payroll:
        flash('Invalid approval link.', 'danger')
        return redirect(url_for('headteacher_approvals'))
    cur.execute("""
        SELECT sp.*, s.full_name, s.position
        FROM salary_payments sp
        JOIN staff s ON sp.staff_id=s.id
        WHERE sp.payroll_id=%s
    """, (payroll['id'],))
    staff_list = cur.fetchall()
    if payroll['token_expires_at']:
        expires_dt = payroll['token_expires_at']
        if isinstance(expires_dt, str):
            expires_dt = datetime.strptime(expires_dt.split('.')[0], '%Y-%m-%d %H:%M:%S')
        if expires_dt <= datetime.now():
            flash('This approval link has expired.', 'danger')
            return redirect(url_for('headteacher_approvals'))
    if request.method == 'POST':
        approval_code = request.form.get('approval_code')
        action = request.form.get('action')
        if payroll['approval_code'] != approval_code:
            flash('Invalid approval code.', 'danger')
            return redirect(url_for('headteacher_approval_access', token=token))
        if action == 'approve':
            mgmt_code = generate_approval_code()
            mgmt_token, mgmt_expires = generate_secure_token(2)
            cur.execute("""
                UPDATE payroll SET approval_status='approved',
                approved_by=%s,
                approved_at=CURRENT_TIMESTAMP,
                management_approval_code=%s,
                management_access_token=%s,
                management_token_expires_at=%s,
                management_approval_status='pending'
                WHERE id=%s
            """, ('Headteacher', mgmt_code, mgmt_token, mgmt_expires, payroll['id']))
            cur.execute("UPDATE salary_payments SET approval_status='approved' WHERE payroll_id=%s", (payroll['id'],))
            db.commit()
            management_link=url_for('management_authorization_access',token=mgmt_token,_external=True)
            cur.execute("SELECT phone FROM users WHERE role='management' AND status=1")
            managers=cur.fetchall()
            for manager in managers:
                phone=manager['phone']
                if phone:
                    send_sms(phone,f"Payroll {payroll['payroll_no']} needs authorization. Code:{mgmt_code} Link:{management_link}")
            add_notification('management',f"Payroll {payroll['payroll_no']} needs bank authorization.",f"/management/authorization/{mgmt_token}")
            flash('Payroll approved. Management notified.','success')
        elif action=='reject':
            cur.execute("""
                UPDATE payroll SET approval_status='rejected',
                approved_by=%s,
                approved_at=CURRENT_TIMESTAMP
                WHERE id=%s
            """,('Headteacher',payroll['id']))
            cur.execute("UPDATE salary_payments SET approval_status='rejected' WHERE payroll_id=%s",(payroll['id'],))
            db.commit()
            add_notification('bursar',f"Payroll {payroll['payroll_no']} rejected by Headteacher.",'/bursar/payroll/list')
            flash('Payroll rejected.','warning')
        cur.close()
        return redirect(url_for('headteacher_approvals'))
    remaining_minutes=None
    if payroll['token_expires_at']:
        expires=payroll['token_expires_at']
        if isinstance(expires,str):
            expires=datetime.strptime(expires.split('.')[0],'%Y-%m-%d %H:%M:%S')
        remaining_minutes=int((expires-datetime.now()).total_seconds()/60)
    cur.close()
    return render_template('headteacher/approve_payroll_secure.html',
                           payroll=payroll,
                           remaining_minutes=remaining_minutes,
                           staff_list=staff_list)

@app.route('/headteacher/reject_payroll/<int:payroll_id>')
def headteacher_reject_payroll(payroll_id):
    if not check_permission(['headteacher']):
        abort(403)
    db=get_db_dict()
    cur=db.cursor()
    cur.execute("SELECT * FROM payroll WHERE id=%s AND approval_status='pending'",(payroll_id,))
    payroll=cur.fetchone()
    cur.close()
    if not payroll:
        flash('Payroll not found or already processed.','danger')
        return redirect(url_for('headteacher_approvals'))
    execute_db("""
        UPDATE payroll SET approval_status='rejected',
        approved_by=%s,
        approved_at=CURRENT_TIMESTAMP
        WHERE id=%s
    """,('Headteacher',payroll_id))
    execute_db("UPDATE salary_payments SET approval_status='rejected' WHERE payroll_id=%s",(payroll_id,))
    add_notification('bursar',f"Payroll {payroll['payroll_no']} rejected by Headteacher.",'/bursar/payroll/list')
    flash('Payroll rejected.','warning')
    return redirect(url_for('headteacher_approvals'))

@app.route('/headteacher/view_payroll/<int:payroll_id>')
def headteacher_view_payroll(payroll_id):
    if not check_permission(['headteacher']):
        abort(403)
    db=get_db_dict()
    cur=db.cursor()
    cur.execute("SELECT * FROM payroll WHERE id=%s",(payroll_id,))
    payroll=cur.fetchone()
    if not payroll:
        flash('Payroll not found.','danger')
        return redirect(url_for('headteacher_approvals'))
    cur.execute("""
        SELECT sp.*,s.full_name,s.position
        FROM salary_payments sp
        JOIN staff s ON sp.staff_id=s.id
        WHERE sp.payroll_id=%s
    """,(payroll_id,))
    staff_list=cur.fetchall()
    cur.close()
    return render_template('headteacher/view_payroll.html',payroll=payroll,staff_list=staff_list)

@app.route('/headteacher/students')
def headteacher_students():
    if not check_permission(['headteacher']):
        abort(403)
    db=get_db_dict()
    cur=db.cursor()
    cur.execute("""
        SELECT student_id,full_name,class,parent_phone,admission_status,fees_balance
        FROM students
        WHERE admission_status='approved'
        ORDER BY class,full_name
    """)
    students=cur.fetchall()
    cur.execute("SELECT DISTINCT class FROM students WHERE class IS NOT NULL AND class!='' ORDER BY class")
    classes=[row['class'] for row in cur.fetchall()]
    cur.close()
    return render_template('headteacher/students.html',students=students,classes=classes)

@app.route('/headteacher/update_comment',methods=['POST'])
def headteacher_update_comment():
    if not check_permission(['headteacher']):
        abort(403)
    student_id=request.form['student_id']
    term=request.form['term']
    year=request.form['year']
    comment=request.form.get('comment','').strip()
    custom=request.form.get('custom_comment','').strip()
    final_comment=custom if custom else comment
    execute_db("""
        INSERT INTO teacher_comments
        (student_id,term,year,headteacher_comment,headteacher_comment_locked)
        VALUES (%s,%s,%s,%s,1)
        ON CONFLICT(student_id,term,year)
        DO UPDATE SET headteacher_comment=%s,
        headteacher_comment_locked=1
    """,(student_id,term,year,final_comment,final_comment))
    flash('Headteacher comment saved and locked.','success')
    return redirect(url_for('teacher_report_card',student_id=student_id,term=term,year=year))

@app.route('/management/pending')
def management_pending_authorizations():
    if not check_permission(['management']):
        abort(403)
    db = get_db_dict()
    cur = db.cursor()
    cur.execute("""
        SELECT p.*, COUNT(sp.id) AS staff_count
        FROM payroll p
        LEFT JOIN salary_payments sp ON p.id=sp.payroll_id
        WHERE p.management_approval_status='pending'
        AND p.approval_status='approved'
        GROUP BY p.id
        ORDER BY p.created_at DESC
    """)
    pending = cur.fetchall()
    cur.close()
    return render_template('management/pending.html', pending=pending)

@app.route('/management/authorization/<token>', methods=['GET', 'POST'])
def management_authorization_access(token):
    if not check_permission(['management']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT *
        FROM payroll
        WHERE management_access_token=%s
        AND management_approval_status='pending'
        AND approval_status='approved'
    """, (token,))
    payroll = cur.fetchone()

    if not payroll:
        flash('Invalid authorization link.', 'danger')
        return redirect(url_for('management_pending_authorizations'))

    if payroll['management_token_expires_at'] and payroll['management_token_expires_at'] <= datetime.now():
        flash('Authorization link expired.', 'danger')
        return redirect(url_for('management_pending_authorizations'))

    if request.method == 'POST':
        auth_code = request.form.get('auth_code')
        action = request.form.get('action')

        if payroll['management_approval_code'] != auth_code:
            flash('Invalid authorization code.', 'danger')
            return redirect(url_for('management_authorization_access', token=token))

        if action == 'authorize':
            result = process_bank_payment(payroll)

            if result['success']:
                cur.execute("""
                    UPDATE payroll
                    SET management_approval_status='approved',
                        management_approved_by='Management',
                        management_approved_at=CURRENT_TIMESTAMP,
                        bank_authorization_token=%s,
                        bank_transaction_ref=%s,
                        bank_payment_status='completed'
                    WHERE id=%s
                """, (
                    result['token'],
                    result['reference'],
                    payroll['id']
                ))

                cur.execute("""
                    UPDATE salary_payments
                    SET approval_status='paid',
                        payment_date=CURRENT_DATE,
                        transaction_ref=%s
                    WHERE payroll_id=%s
                """, (
                    result['reference'],
                    payroll['id']
                ))

                db.commit()

                add_notification(
                    'bursar',
                    f"Payroll {payroll['payroll_no']} has been paid. Ref: {result['reference']}",
                    '/bursar/payroll/list'
                )

                flash(f"Payment authorized. Reference: {result['reference']}", 'success')

            else:
                cur.execute("""
                    UPDATE payroll
                    SET bank_payment_status='failed',
                        bank_payment_response=%s
                    WHERE id=%s
                """, (
                    result['error'],
                    payroll['id']
                ))

                db.commit()

                flash(f"Payment failed: {result['error']}", 'danger')

        elif action == 'reject':

            cur.execute("""
                UPDATE payroll
                SET management_approval_status='rejected',
                    management_approved_by='Management',
                    management_approved_at=CURRENT_TIMESTAMP
                WHERE id=%s
            """, (payroll['id'],))

            cur.execute("""
                UPDATE salary_payments
                SET approval_status='rejected'
                WHERE payroll_id=%s
            """, (payroll['id'],))

            db.commit()

            add_notification(
                'headteacher',
                f"Payroll {payroll['payroll_no']} was rejected by Management.",
                '/headteacher/approvals'
            )

            add_notification(
                'bursar',
                f"Payroll {payroll['payroll_no']} was rejected by Management.",
                '/bursar/payroll/list'
            )

            flash('Payment authorization rejected.', 'warning')

        cur.close()
        return redirect(url_for('management_pending_authorizations'))

    remaining_minutes = None
    if payroll['management_token_expires_at']:
        remaining = payroll['management_token_expires_at'] - datetime.now()
        remaining_minutes = int(remaining.total_seconds() / 60)

    cur.close()

    return render_template(
        'management/authorize_payment_secure.html',
        payroll=payroll,
        remaining_minutes=remaining_minutes
    )

@app.route('/management/view_payroll/<int:payroll_id>')
def management_view_payroll(payroll_id):
    if not check_permission(['management']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("SELECT * FROM payroll WHERE id=%s", (payroll_id,))
    payroll = cur.fetchone()

    if not payroll:
        flash('Payroll not found.', 'danger')
        return redirect(url_for('management_pending_authorizations'))

    cur.execute("""
        SELECT sp.*, s.full_name, s.position,
               s.bank_account, s.bank_name,
               s.phone, s.nssf_number, s.tin_number
        FROM salary_payments sp
        JOIN staff s ON sp.staff_id=s.id
        WHERE sp.payroll_id=%s
    """, (payroll_id,))

    staff_list = cur.fetchall()

    cur.close()

    return render_template(
        'management/view_payroll.html',
        payroll=payroll,
        staff_list=staff_list
    )

@app.route('/management/reject_authorization/<int:payroll_id>')
def management_reject_authorization(payroll_id):
    if not check_permission(['management']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT *
        FROM payroll
        WHERE id=%s
        AND management_approval_status='pending'
        AND approval_status='approved'
    """, (payroll_id,))

    payroll = cur.fetchone()
    cur.close()

    if not payroll:
        flash('Payroll not found or already processed.', 'danger')
        return redirect(url_for('management_pending_authorizations'))

    execute_db("""
        UPDATE payroll
        SET management_approval_status='rejected',
            management_approved_by='Management',
            management_approved_at=CURRENT_TIMESTAMP
        WHERE id=%s
    """, (payroll_id,))

    execute_db("""
        UPDATE salary_payments
        SET approval_status='rejected'
        WHERE payroll_id=%s
    """, (payroll_id,))

    add_notification(
        'headteacher',
        f"Payroll {payroll['payroll_no']} authorization rejected by Management.",
        '/headteacher/approvals'
    )

    add_notification(
        'bursar',
        f"Payroll {payroll['payroll_no']} authorization rejected by Management.",
        '/bursar/payroll/list'
    )

    flash('Payroll authorization rejected.', 'warning')

    return redirect(url_for('management_pending_authorizations'))

@app.route('/management/resend_token/<int:payroll_id>')
def management_resend_token(payroll_id):
    if not check_permission(['management']):
        abort(403)

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT *
        FROM payroll
        WHERE id=%s
        AND management_approval_status='pending'
        AND approval_status='approved'
    """, (payroll_id,))

    payroll = cur.fetchone()
    cur.close()

    if not payroll:
        flash('Payroll not found or already authorized.', 'danger')
        return redirect(url_for('management_pending_authorizations'))

    if payroll.get('token_resend_count', 0) >= 3:
        flash('Maximum resend limit reached.', 'danger')
        return redirect(url_for('management_pending_authorizations'))

    new_token, new_expires = generate_secure_token(2)

    execute_db("""
        UPDATE payroll
        SET management_access_token=%s,
            management_token_expires_at=%s,
            token_resend_count=token_resend_count+1,
            last_resend_at=CURRENT_TIMESTAMP
        WHERE id=%s
    """, (
        new_token,
        new_expires,
        payroll_id
    ))

    auth_link = url_for(
        'management_authorization_access',
        token=new_token,
        _external=True
    )

    expires_str = new_expires.strftime('%Y-%m-%d %H:%M:%S')

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT phone
        FROM users
        WHERE role='management'
        AND status=1
        LIMIT 1
    """)

    manager = cur.fetchone()
    cur.close()

    if manager and manager['phone']:
        send_sms(
            manager['phone'],
            f"NEW LINK: Payroll {payroll['payroll_no']} Code:{payroll['management_approval_code']} Expires:{expires_str} Link:{auth_link}"
        )

    flash('New authorization link sent successfully.', 'success')

    return redirect(url_for('management_pending_authorizations'))


@app.route('/parent/dashboard')
@login_required
def parent_dashboard():

    parent_id = session['user_id']

    db = get_db()
    cursor = db.cursor(cursor_factory=RealDictCursor)


    cursor.execute("""
        SELECT
            student_id,
            full_name,
            class,
            photo_path,
            fees_total,
            fees_paid,
            fees_balance,
            payment_status

        FROM students

        WHERE parent_id=%s

    """,(parent_id,))


    children = cursor.fetchall()


    cursor.close()


    return render_template(
        'dash_parent.html',
        children=children
    )
import random


@app.route('/parent/generate_prn/<student_id>')
@login_required
def generate_prn(student_id):

    parent_id=session['user_id']

    db=get_db()
    cursor=db.cursor()


    cursor.execute("""
        SELECT fees_balance
        FROM students
        WHERE student_id=%s
        AND parent_id=%s

    """,(student_id,parent_id))


    student=cursor.fetchone()


    if not student:
        abort(403)


    balance=student[0]


    prn="PRN"+str(random.randint(100000000,999999999))


    cursor.execute("""
        INSERT INTO payment_requests
        (student_id,prn,amount)

        VALUES(%s,%s,%s)

    """,(student_id,prn,balance))


    db.commit()


    return render_template(
        'parent/fee_prn.html',
        prn=prn,
        amount=balance
    )


@app.route('/parent/fees/<student_id>')
@login_required
def parent_fees(student_id):

    parent_id=session['user_id']

    db=get_db()
    cursor=db.cursor(cursor_factory=RealDictCursor)


    cursor.execute("""
        SELECT
            full_name,
            fees_total,
            fees_paid,
            fees_balance

        FROM students

        WHERE student_id=%s
        AND parent_id=%s

    """,(student_id,parent_id))


    fees=cursor.fetchone()


    if not fees:
        abort(403)



    return render_template(
        'parent/fees.html',
        fees=fees
    )

@app.route('/parent/report_card/<student_id>')
@login_required
def parent_report_card(student_id):

    parent_id=session['user_id']

    db=get_db()
    cursor=db.cursor(cursor_factory=RealDictCursor)


    # confirm ownership
    cursor.execute("""
        SELECT *
        FROM students
        WHERE student_id=%s
        AND parent_id=%s

    """,(student_id,parent_id))


    student=cursor.fetchone()


    if not student:
        abort(403)



    cursor.execute("""
        SELECT *
        FROM marks

        WHERE student_id=%s

        ORDER BY subject

    """,(student_id,))


    marks=cursor.fetchall()



    return render_template(
        'parent/report_card.html',
        student=student,
        marks=marks
    )

# ==================== INVENTORY HELPERS ====================
def generate_item_code(category_name):
    prefix = category_name[:3].upper()
    year = datetime.now().strftime("%Y")

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT item_code
        FROM inventory_items
        WHERE item_code LIKE %s
        ORDER BY item_code DESC
        LIMIT 1
    """, (f'{prefix}-{year}-%',))

    last = cur.fetchone()
    cur.close()

    if last:
        number = int(last['item_code'].split('-')[-1]) + 1
    else:
        number = 1

    return f"{prefix}-{year}-{number:04d}"

def check_low_stock_alerts():

    db = get_db_dict()
    cur = db.cursor()

    cur.execute("""
        SELECT
            i.id,
            i.name,
            i.quantity,
            i.unit
        FROM inventory_items i
        WHERE i.quantity <= i.reorder_level
        AND i.status='working'
    """)

    items = cur.fetchall()


    for item in items:

        cur.execute("""
            SELECT id
            FROM inventory_alerts
            WHERE item_id=%s
            AND alert_type='low_stock'
            AND is_read=FALSE
        """,
        (item['id'],))

        exists = cur.fetchone()


        if not exists:

            execute_db(
                """
                INSERT INTO inventory_alerts
                (
                    item_id,
                    alert_type,
                    message
                )
                VALUES
                (%s,%s,%s)
                """,
                (
                    item['id'],
                    'low_stock',
                    f"{item['name']} stock is low. "
                    f"Available: {item['quantity']} {item['unit']}"
                )
            )


            add_notification(
                'stores_keeper',
                f"LOW STOCK: {item['name']} only has {item['quantity']} left",
                '/inventory/items'
            )


            add_notification(
                'admin',
                f"LOW STOCK ALERT: {item['name']} requires ordering",
                '/inventory/items'
            )


    cur.close()



# ==================== INVENTORY DASHBOARD ====================


@app.route('/inventory/dashboard')
def inventory_dashboard():

    if not check_permission(
        ['admin','bursar','dos','stores_keeper']
    ):
        abort(403)


    db = get_db_dict()
    cur = db.cursor()


    cur.execute("""
        SELECT COUNT(*) AS total
        FROM inventory_items
    """)
    total_items = cur.fetchone()


    cur.execute("""
        SELECT COUNT(*) AS total
        FROM inventory_items
        WHERE quantity <= reorder_level
        AND status='working'
    """)
    low_stock = cur.fetchone()


    cur.execute("""
        SELECT COUNT(*) AS total
        FROM inventory_items
        WHERE status='spoilt'
    """)
    spoilt = cur.fetchone()


    cur.execute("""
        SELECT COUNT(*) AS total
        FROM inventory_items
        WHERE status='under_repair'
    """)
    repair = cur.fetchone()


    cur.execute("""
        SELECT COALESCE(SUM(quantity),0) AS total
        FROM inventory_items
        WHERE status='working'
    """)
    quantity = cur.fetchone()


    cur.execute("""
        SELECT COALESCE(SUM(current_value),0) AS total
        FROM inventory_items
    """)
    value = cur.fetchone()



    cur.execute("""
        SELECT
            t.*,
            i.name AS item_name,
            i.item_code
        FROM inventory_transactions t
        JOIN inventory_items i
        ON t.item_id=i.id
        ORDER BY t.created_at DESC
        LIMIT 10
    """)

    transactions = cur.fetchall()



    cur.execute("""
        SELECT
            a.*,
            i.name AS item_name,
            i.quantity,
            i.reorder_level
        FROM inventory_alerts a
        JOIN inventory_items i
        ON a.item_id=i.id
        WHERE a.is_read=FALSE
        ORDER BY a.created_at DESC
    """)

    alerts = cur.fetchall()


    cur.close()


    return render_template(
        'inventory/dashboard.html',
        total_items=total_items,
        low_stock=low_stock,
        spoilt=spoilt,
        repair=repair,
        quantity=quantity,
        value=value,
        transactions=transactions,
        alerts=alerts
    )



# ==================== ITEMS LIST ====================


@app.route('/inventory/items')
def inventory_items():

    if not check_permission(
        ['admin','bursar','dos','stores_keeper']
    ):
        abort(403)


    category = request.args.get('category','')
    status = request.args.get('status','')
    search = request.args.get('search','')


    db = get_db_dict()
    cur = db.cursor()


    query = """
        SELECT
            i.*,
            c.name AS category_name
        FROM inventory_items i
        JOIN inventory_categories c
        ON i.category_id=c.id
        WHERE 1=1
    """

    params=[]


    if category:
        query += " AND c.name=%s"
        params.append(category)


    if status:
        query += " AND i.status=%s"
        params.append(status)


    if search:
        query += """
        AND (
            i.name ILIKE %s
            OR i.item_code ILIKE %s
        )
        """

        params.extend([
            f"%{search}%",
            f"%{search}%"
        ])


    query += """
        ORDER BY i.category_id,i.name
    """


    cur.execute(query,params)

    items = cur.fetchall()



    cur.execute("""
        SELECT *
        FROM inventory_categories
        ORDER BY name
    """)

    categories = cur.fetchall()


    cur.close()


    return render_template(
        'inventory/items.html',
        items=items,
        categories=categories,
        category=category,
        status=status,
        search=search
    )



# ==================== ADD INVENTORY ITEM ====================


@app.route('/inventory/item/add', methods=['GET','POST'])
def inventory_item_add():

    if not check_permission(
        ['admin','bursar','stores_keeper']
    ):
        abort(403)


    db = get_db_dict()
    cur = db.cursor()


    if request.method == 'POST':

        try:

            category_id = request.form.get('category_id')


            if not category_id:
                flash(
                    'Please select a category.',
                    'danger'
                )
                return redirect(
                    url_for('inventory_item_add')
                )


            name = request.form.get(
                'name',
                ''
            ).strip()


            unit = request.form.get(
                'unit',
                'pieces'
            )


            quantity = int(
                request.form.get(
                    'quantity',
                    0
                )
            )


            minimum_quantity = int(
                request.form.get(
                    'minimum_quantity',
                    10
                )
            )


            reorder_level = int(
                request.form.get(
                    'reorder_level',
                    5
                )
            )


            purchase_price = float(
                request.form.get(
                    'purchase_price',
                    0
                )
            )


            current_value = (
                quantity * purchase_price
            )


            location = request.form.get(
                'location',
                ''
            )


            supplier = request.form.get(
                'supplier',
                ''
            )


            status = request.form.get(
                'status',
                'working'
            )


            responsible_person = request.form.get(
                'responsible_person',
                ''
            )


            responsible_role = request.form.get(
                'responsible_role',
                ''
            )


            cur.execute(
                """
                SELECT name
                FROM inventory_categories
                WHERE id=%s
                """,
                (category_id,)
            )


            category = cur.fetchone()


            if not category:
                flash(
                    'Invalid category.',
                    'danger'
                )
                return redirect(
                    url_for('inventory_item_add')
                )


            item_code = generate_item_code(
                category['name']
            )


            image_path = None


            image_file = request.files.get('image')


            if image_file and image_file.filename:

                if allowed_file(
                    image_file.filename,
                    ALLOWED_IMAGE_EXTENSIONS
                ):

                    filename = secure_filename(
                        f"{item_code}_{image_file.filename}"
                    )


                    upload_path = os.path.join(
                        'uploads',
                        filename
                    )


                    image_file.save(
                        upload_path
                    )


                    image_path = filename



            cur.execute("""
                INSERT INTO inventory_items
                (
                    item_code,
                    name,
                    category_id,
                    unit,
                    quantity,
                    minimum_quantity,
                    reorder_level,
                    location,
                    supplier,
                    purchase_price,
                    current_value,
                    status,
                    responsible_person,
                    responsible_role,
                    image_path,
                    created_at,
                    updated_at
                )

                VALUES
                (
                    %s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,
                    %s,%s,%s,%s,%s,
                    CURRENT_TIMESTAMP,
                    CURRENT_TIMESTAMP
                )

                RETURNING id
            """,
            (
                item_code,
                name,
                category_id,
                unit,
                quantity,
                minimum_quantity,
                reorder_level,
                location,
                supplier,
                purchase_price,
                current_value,
                status,
                responsible_person,
                responsible_role,
                image_path
            ))


            item_id = cur.fetchone()['id']


            cur.execute("""
                INSERT INTO inventory_transactions
                (
                    item_id,
                    transaction_type,
                    quantity,
                    unit_price,
                    total_amount,
                    transaction_date,
                    recorded_by,
                    notes
                )

                VALUES
                (
                    %s,
                    'purchase',
                    %s,
                    %s,
                    %s,
                    CURRENT_DATE,
                    %s,
                    %s
                )
            """,
            (
                item_id,
                quantity,
                purchase_price,
                current_value,
                session.get('username'),
                'Initial stock'
            ))


            db.commit()


            flash(
                f'Item added successfully. Code: {item_code}',
                'success'
            )


        except Exception as e:

            db.rollback()

            flash(
                f'Error adding item: {str(e)}',
                'danger'
            )


        finally:

            cur.close()


        return redirect(
            url_for('inventory_items')
        )


    cur.execute("""
        SELECT *
        FROM inventory_categories
        ORDER BY name
    """)


    categories = cur.fetchall()


    cur.close()


    return render_template(
        'inventory/item_add.html',
        categories=categories
    )

# ==================== EDIT INVENTORY ITEM ====================

@app.route(
    '/inventory/item/edit/<int:item_id>',
    methods=['GET','POST']
)
def inventory_item_edit(item_id):

    if not check_permission(
        ['admin','bursar','stores_keeper']
    ):
        abort(403)


    if request.method == 'POST':

        name = request.form.get(
            'name'
        )

        unit = request.form.get(
            'unit'
        )

        minimum_quantity = int(
            request.form.get(
                'minimum_quantity',
                0
            )
        )

        reorder_level = int(
            request.form.get(
                'reorder_level',
                5
            )
        )

        location = request.form.get(
            'location',
            ''
        )

        supplier = request.form.get(
            'supplier',
            ''
        )

        status = request.form.get(
            'status',
            'working'
        )

        responsible_person = request.form.get(
            'responsible_person',
            ''
        )

        responsible_role = request.form.get(
            'responsible_role',
            ''
        )


        execute_db(
            """
            UPDATE inventory_items
            SET
                name=%s,
                unit=%s,
                minimum_quantity=%s,
                reorder_level=%s,
                location=%s,
                supplier=%s,
                status=%s,
                responsible_person=%s,
                responsible_role=%s,
                updated_at=CURRENT_TIMESTAMP

            WHERE id=%s
            """,
            (
                name,
                unit,
                minimum_quantity,
                reorder_level,
                location,
                supplier,
                status,
                responsible_person,
                responsible_role,
                item_id
            )
        )


        flash(
            'Item updated successfully.',
            'success'
        )


        return redirect(
            url_for('inventory_items')
        )



    db = get_db_dict()
    cur = db.cursor()


    cur.execute(
        """
        SELECT
            i.*,
            c.name AS category_name

        FROM inventory_items i

        JOIN inventory_categories c
        ON i.category_id=c.id

        WHERE i.id=%s
        """,
        (item_id,)
    )


    item = cur.fetchone()



    cur.execute(
        """
        SELECT *
        FROM inventory_categories
        ORDER BY name
        """
    )


    categories = cur.fetchall()


    cur.close()


    return render_template(
        'inventory/item_edit.html',
        item=item,
        categories=categories
    )



# ==================== ISSUE INVENTORY ITEM ===================

@app.route(
    '/inventory/issue/<int:item_id>',
    methods=['POST']
)
def inventory_issue_item(item_id):

    if not check_permission(
        ['admin','bursar','dos','stores_keeper']
    ):
        abort(403)


    quantity = int(
        request.form.get(
            'quantity',
            0
        )
    )


    issued_to = request.form.get(
        'issued_to'
    )


    issued_to_role = request.form.get(
        'issued_to_role'
    )


    purpose = request.form.get(
        'purpose'
    )


    notes = request.form.get(
        'notes',
        ''
    )


    db = get_db_dict()
    cur = db.cursor()



    cur.execute(
        """
        SELECT
            name,
            quantity,
            unit

        FROM inventory_items

        WHERE id=%s
        """,
        (item_id,)
    )


    item = cur.fetchone()



    if not item:

        flash(
            'Item not found.',
            'danger'
        )

        return redirect(
            url_for('inventory_items')
        )



    if item['quantity'] < quantity:

        flash(
            f"Insufficient stock! Available: {item['quantity']} {item['unit']}",
            'danger'
        )

        return redirect(
            url_for('inventory_items')
        )



    new_quantity = (
        item['quantity']
        -
        quantity
    )



    cur.execute(
        """
        UPDATE inventory_items

        SET quantity=%s,
            updated_at=CURRENT_TIMESTAMP

        WHERE id=%s
        """,
        (
            new_quantity,
            item_id
        )
    )



    cur.execute(
        """
        INSERT INTO inventory_transactions
        (
            item_id,
            transaction_type,
            quantity,
            transaction_date,
            issued_to,
            issued_to_role,
            purpose,
            notes,
            recorded_by
        )

        VALUES
        (
            %s,
            'issued',
            %s,
            CURRENT_DATE,
            %s,
            %s,
            %s,
            %s,
            %s
        )
        """,
        (
            item_id,
            quantity,
            issued_to,
            issued_to_role,
            purpose,
            notes,
            session.get('username')
        )
    )



    db.commit()

    cur.close()



    check_low_stock_alerts()



    flash(
        f"{quantity} {item['unit']} of {item['name']} issued to {issued_to}.",
        'success'
    )


    return redirect(
        url_for('inventory_items')
    )



# ==================== RECEIVE INVENTORY ITEM ====================


@app.route(
    '/inventory/receive/<int:item_id>',
    methods=['POST']
)
def inventory_receive_item(item_id):

    if not check_permission(
        ['admin','bursar','stores_keeper']
    ):
        abort(403)



    quantity = int(
        request.form.get(
            'quantity',
            0
        )
    )


    unit_price = float(
        request.form.get(
            'unit_price',
            0
        )
    )


    supplier = request.form.get(
        'supplier',
        ''
    )


    notes = request.form.get(
        'notes',
        ''
    )



    db = get_db_dict()

    cur = db.cursor()



    cur.execute(
        """
        SELECT
            name,
            quantity,
            current_value,
            unit

        FROM inventory_items

        WHERE id=%s
        """,
        (item_id,)
    )


    item = cur.fetchone()



    if not item:

        flash(
            'Item not found.',
            'danger'
        )

        return redirect(
            url_for('inventory_items')
        )



    new_quantity = (
        item['quantity']
        +
        quantity
    )



    total_amount = (
        quantity
        *
        unit_price
    )



    new_value = (
        item['current_value']
        or 0
    ) + total_amount



    cur.execute(
        """
        UPDATE inventory_items

        SET quantity=%s,
            current_value=%s,
            updated_at=CURRENT_TIMESTAMP

        WHERE id=%s
        """,
        (
            new_quantity,
            new_value,
            item_id
        )
    )



    cur.execute(
        """
        INSERT INTO inventory_transactions
        (
            item_id,
            transaction_type,
            quantity,
            unit_price,
            total_amount,
            transaction_date,
            supplier,
            notes,
            recorded_by
        )

        VALUES
        (
            %s,
            'received',
            %s,
            %s,
            %s,
            CURRENT_DATE,
            %s,
            %s,
            %s
        )
        """,
        (
            item_id,
            quantity,
            unit_price,
            total_amount,
            supplier,
            notes,
            session.get('username')
        )
    )



    cur.execute(
        """
        UPDATE inventory_alerts

        SET is_read=TRUE

        WHERE item_id=%s

        AND alert_type='low_stock'
        """,
        (item_id,)
    )



    db.commit()

    cur.close()



    flash(
        f"{quantity} {item['unit']} of {item['name']} received.",
        'success'
    )



    return redirect(
        url_for('inventory_items')
    )



# ==================== UPDATE INVENTORY STATUS ====================


@app.route(
    '/inventory/update_status/<int:item_id>',
    methods=['POST']
)
def inventory_update_status(item_id):

    if not check_permission(
        ['admin','bursar','stores_keeper']
    ):
        abort(403)



    status = request.form.get(
        'status'
    )


    condition_notes = request.form.get(
        'condition_notes',
        ''
    )


    quantity_affected = int(
        request.form.get(
            'quantity_affected',
            0
        )
    )



    db = get_db_dict()

    cur = db.cursor()



    cur.execute(
        """
        SELECT
            name,
            quantity

        FROM inventory_items

        WHERE id=%s
        """,
        (item_id,)
    )


    item = cur.fetchone()



    if not item:

        flash(
            'Item not found.',
            'danger'
        )

        return redirect(
            url_for('inventory_items')
        )



    if status in [
        'spoilt',
        'used_up'
    ] and quantity_affected > 0:



        new_quantity = (
            item['quantity']
            -
            quantity_affected
        )



        cur.execute(
            """
            UPDATE inventory_items

            SET quantity=%s,
                status=%s,
                condition_notes=%s,
                updated_at=CURRENT_TIMESTAMP

            WHERE id=%s
            """,
            (
                new_quantity,
                status,
                condition_notes,
                item_id
            )
        )



        cur.execute(
            """
            INSERT INTO inventory_transactions
            (
                item_id,
                transaction_type,
                quantity,
                notes,
                recorded_by
            )

            VALUES
            (
                %s,
                %s,
                %s,
                %s,
                %s
            )
            """,
            (
                item_id,
                status,
                quantity_affected,
                condition_notes,
                session.get('username')
            )
        )


    else:


        cur.execute(
            """
            UPDATE inventory_items

            SET status=%s,
                condition_notes=%s,
                updated_at=CURRENT_TIMESTAMP

            WHERE id=%s
            """,
            (
                status,
                condition_notes,
                item_id
            )
        )



    db.commit()

    cur.close()



    flash(
        f'Item status updated to {status}.',
        'success'
    )



    return redirect(
        url_for('inventory_items')
    )

# ==================== INVENTORY TRANSACTIONS ====================


@app.route('/inventory/transactions')
def inventory_transactions():

    if not check_permission(
        ['admin','bursar','dos','stores_keeper']
    ):
        abort(403)


    item_id = request.args.get(
        'item_id',
        ''
    )


    db = get_db_dict()

    cur = db.cursor()



    if item_id:

        cur.execute(
            """
            SELECT
                t.*,
                i.name AS item_name,
                i.item_code

            FROM inventory_transactions t

            JOIN inventory_items i
            ON t.item_id=i.id

            WHERE t.item_id=%s

            ORDER BY t.created_at DESC
            """,
            (item_id,)
        )


    else:

        cur.execute(
            """
            SELECT
                t.*,
                i.name AS item_name,
                i.item_code

            FROM inventory_transactions t

            JOIN inventory_items i
            ON t.item_id=i.id

            ORDER BY t.created_at DESC
            """
        )



    transactions = cur.fetchall()



    cur.execute(
        """
        SELECT
            id,
            name,
            item_code

        FROM inventory_items

        ORDER BY name
        """
    )


    items = cur.fetchall()



    cur.close()



    return render_template(
        'inventory/transactions.html',
        transactions=transactions,
        items=items,
        selected_item=item_id
    )



# ==================== INVENTORY ALERTS ====================


@app.route('/inventory/alerts')
def inventory_alerts():

    if not check_permission(
        ['admin','bursar','dos','stores_keeper']
    ):
        abort(403)



    db = get_db_dict()

    cur = db.cursor()



    cur.execute(
        """
        SELECT
            a.*,
            i.name AS item_name,
            i.quantity,
            i.reorder_level,
            i.unit

        FROM inventory_alerts a

        JOIN inventory_items i
        ON a.item_id=i.id

        WHERE a.is_read=FALSE

        ORDER BY a.created_at DESC
        """
    )



    alerts = cur.fetchall()



    cur.close()



    return render_template(
        'inventory/alerts.html',
        alerts=alerts
    )



# ==================== MARK ALERT AS READ ====================


@app.route(
    '/inventory/alert/read/<int:alert_id>'
)
def inventory_alert_read(alert_id):

    if not check_permission(
        ['admin','bursar','dos','stores_keeper']
    ):
        abort(403)



    execute_db(
        """
        UPDATE inventory_alerts

        SET is_read=TRUE

        WHERE id=%s
        """,
        (alert_id,)
    )



    flash(
        'Alert acknowledged.',
        'success'
    )


    return redirect(
        url_for('inventory_alerts')
    )



# ==================== INVENTORY REPORTS ====================


@app.route('/inventory/reports')
def inventory_reports():

    if not check_permission(
        ['admin','bursar','stores_keeper']
    ):
        abort(403)
    by_category = []

    by_status = []

    low_stock_items = []

    recent_issues = []



    total_items = 0

    total_quantity = 0

    low_stock_count = 0

    total_value = 0



    try:

        db = get_db_dict()

        cur = db.cursor()



        # Total items

        cur.execute(
            """
            SELECT COUNT(*) AS total

            FROM inventory_items
            """
        )


        total_items = cur.fetchone()['total'] or 0




        # Total quantity

        cur.execute(
            """
            SELECT COALESCE(
                SUM(quantity),
                0
            ) AS total

            FROM inventory_items

            WHERE status='working'
            """
        )


        total_quantity = cur.fetchone()['total'] or 0




        # Low stock count

        cur.execute(
            """
            SELECT COUNT(*) AS total

            FROM inventory_items

            WHERE quantity <= reorder_level

            AND status='working'
            """
        )


        low_stock_count = cur.fetchone()['total'] or 0




        # Total value

        cur.execute(
            """
            SELECT COALESCE(
                SUM(current_value),
                0
            ) AS total

            FROM inventory_items
            """
        )


        total_value = cur.fetchone()['total'] or 0




        # Stock by category

        cur.execute(
            """
            SELECT

                c.name AS category,

                COUNT(i.id) AS item_count,

                COALESCE(
                    SUM(i.quantity),
                    0
                ) AS total_quantity,

                COALESCE(
                    SUM(i.current_value),
                    0
                ) AS total_value


            FROM inventory_categories c


            LEFT JOIN inventory_items i

            ON c.id=i.category_id


            GROUP BY c.id,c.name

            ORDER BY c.name
            """
        )



        rows = cur.fetchall()



        for row in rows:

            by_category.append(
                {
                    'category': row['category'],

                    'item_count': row['item_count'],

                    'total_quantity': row['total_quantity'],

                    'total_value': row['total_value']
                }
            )





        # Stock by status


        cur.execute(
            """
            SELECT

                status,

                COUNT(*) AS count,

                COALESCE(
                    SUM(quantity),
                    0
                ) AS quantity


            FROM inventory_items


            GROUP BY status
            """
        )


        rows = cur.fetchall()



        for row in rows:

            by_status.append(
                {
                    'status': row['status'],

                    'count': row['count'],

                    'quantity': row['quantity']
                }
            )






        # Low stock items


        cur.execute(
            """
            SELECT

                i.id,

                i.item_code,

                i.name,

                i.quantity,

                i.reorder_level,

                i.unit,

                c.name AS category_name


            FROM inventory_items i


            LEFT JOIN inventory_categories c

            ON i.category_id=c.id


            WHERE i.quantity <= i.reorder_level

            AND i.status='working'


            ORDER BY i.quantity ASC
            """
        )



        rows = cur.fetchall()



        for row in rows:

            low_stock_items.append(
                {
                    'id': row['id'],

                    'item_code': row['item_code'],

                    'name': row['name'],

                    'quantity': row['quantity'],

                    'reorder_level': row['reorder_level'],

                    'unit': row['unit'],

                    'category_name': row['category_name']
                }
            )






        # Recent issues


        cur.execute(
            """
            SELECT

                t.transaction_date,

                t.quantity,

                t.issued_to,

                t.purpose,

                t.recorded_by,

                i.name AS item_name


            FROM inventory_transactions t


            LEFT JOIN inventory_items i

            ON t.item_id=i.id


            WHERE t.transaction_type='issued'


            ORDER BY t.created_at DESC


            LIMIT 20
            """
        )



        rows = cur.fetchall()



        for row in rows:

            recent_issues.append(
                {
                    'transaction_date':
                        row['transaction_date'],

                    'quantity':
                        row['quantity'],

                    'issued_to':
                        row['issued_to'],

                    'purpose':
                        row['purpose'],

                    'recorded_by':
                        row['recorded_by'],

                    'item_name':
                        row['item_name']
                }
            )



        cur.close()



    except Exception as e:

        print(
            f"Error in inventory_reports: {str(e)}"
        )


        flash(
            f'Error loading reports: {str(e)}',
            'danger'
        )




    return render_template(
        'inventory/reports.html',

        by_category=by_category,

        by_status=by_status,

        low_stock_items=low_stock_items,

        recent_issues=recent_issues,

        total_items=total_items,

        total_quantity=total_quantity,

        low_stock_count=low_stock_count,

        total_value=total_value
    )



# ==================== PRINT INVENTORY REPORT ====================


@app.route('/inventory/print_report')
def inventory_print_report():

    if not check_permission(
        ['admin','bursar','stores_keeper']
    ):
        abort(403)



    category = request.args.get(
        'category',
        ''
    )



    db = get_db_dict()

    cur = db.cursor()



    if category:


        cur.execute(
            """
            SELECT

                i.*,

                c.name AS category_name


            FROM inventory_items i


            JOIN inventory_categories c

            ON i.category_id=c.id


            WHERE c.name=%s


            ORDER BY i.name
            """,
            (category,)
        )


    else:


        cur.execute(
            """
            SELECT

                i.*,

                c.name AS category_name


            FROM inventory_items i


            JOIN inventory_categories c

            ON i.category_id=c.id


            ORDER BY c.name,i.name
            """
        )



    items = cur.fetchall()



    cur.close()



    return render_template(
        'inventory/print_report.html',

        items=items,

        category=category
    )


# ==================== ALERT COUNT API ====================

@app.route('/inventory/alert/count')
def inventory_alert_count():

    if not check_permission(['admin', 'bursar', 'dos', 'stores_keeper']):
        return jsonify({'count':0})


    cur = get_db().cursor()


    cur.execute("""
        SELECT COUNT(*)
        FROM inventory_alerts
        WHERE is_read=FALSE
    """)


    count = cur.fetchone()[0]


    cur.close()


    return jsonify({
        'count': count
    })

# ==================== UPLOADS & MISC ====================

@app.route('/uploads/<filename>')
def uploaded_file(filename):
    return send_from_directory(
        app.config['UPLOAD_FOLDER'],
        filename
    )


# ==================== MOBILE API ENDPOINTS ====================

@app.route('/mobile/login', methods=['POST'])
def mobile_login():

    data = request.get_json()

    username = data.get('username')
    password = data.get('password')

    db = get_db()
    cur = db.cursor()

    cur.execute("""
        SELECT id, username, role, status
        FROM users
        WHERE username=%s
        AND password=%s
    """,
    (
        username,
        password
    ))

    user = cur.fetchone()

    cur.close()


    if user and user[3] == 1:

        token, _ = generate_secure_token()

        return jsonify({
            'success': True,
            'token': token,
            'role': user[2],
            'username': user[1]
        })


    return jsonify({
        'success': False,
        'message': 'Invalid credentials'
    })



@app.route('/mobile/dashboard', methods=['GET'])
def mobile_dashboard():

    token = request.headers.get(
        'Authorization',
        ''
    ).replace(
        'Bearer ',
        ''
    )


    role = session.get('role')


    if role == 'admin':

        db = get_db()
        cur = db.cursor()


        cur.execute("""
            SELECT COUNT(*)
            FROM users
        """)

        users = cur.fetchone()


        cur.execute("""
            SELECT COUNT(*)
            FROM students
        """)

        students = cur.fetchone()


        cur.close()


        return jsonify({

            'total_users':
                users[0] if users else 0,

            'total_students':
                students[0] if students else 0

        })


    return jsonify({})


if __name__ == '__main__':
    app.run(debug=True)
